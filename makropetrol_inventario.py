# ============================================================
# MAKROPETROL NEXUS v5.0 — Executive Intelligence
# Plataforma de inteligencia logística multisede
#
# Incluye:
# - Recalculo dinámico en tiempo real desde datos base
# - Persistencia local diaria con SQLite (capa de acceso corta, sin
#   conexión global compartida, pensada para concurrencia futura)
# - Inventario + ventas por corte, con lectura Excel acelerada (calamine)
# - Cortes dedicados por proveedor/marca sin sobrescribir el corte maestro
# - Centro de productos sin rotación con trazabilidad histórica
# - Ventas acumuladas / por período / diarias
# - Compras, retiros y redistribución automática entre sedes (con Sankey)
# - ABC / Pareto, rotación y cobertura
# - Punto de reorden + lead time + stock de seguridad
# - Alertas y Health Score
# - Historial de cortes
# - Modo Gerencia / Modo Operativo
# - Exportación a 6 formatos: Excel operativo, Excel gerencial, HTML,
#   PDF ejecutivo, CSV, JSON + paquete ZIP completo
# - Hub de automatización mediante webhook (CRM / n8n / Make / WhatsApp provider)
#
# Requisitos: streamlit, pandas, numpy, openpyxl, plotly, requests
# Recomendado (mejora de rendimiento de lectura Excel): python-calamine
# Opcional (PDF ejecutivo): fpdf2
# Ejecutar: streamlit run makropetrol_nexus_v4_3.py
# ============================================================

from __future__ import annotations

import io
import json
import re
import sqlite3
import unicodedata
import zipfile
from concurrent.futures import ThreadPoolExecutor

from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import requests
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio

# Tema global NEXUS: evita plantillas heredadas que puedan serializar colores
# negros al exportar HTML y mantiene una identidad visual consistente.
NEXUS_BLUE = "#3568F5"
NEXUS_BLUE_2 = "#244BD7"
NEXUS_CYAN = "#17AABE"
NEXUS_GREEN = "#14A36A"
NEXUS_ORANGE = "#EF982D"
NEXUS_RED = "#DF5555"
NEXUS_GRAY = "#7B8799"
NEXUS_PALETTE = [NEXUS_BLUE, NEXUS_CYAN, NEXUS_GREEN, NEXUS_ORANGE, NEXUS_RED, NEXUS_BLUE_2, NEXUS_GRAY]
pio.templates.default = "plotly_white"

# Motor de lectura Excel acelerado. Si python-calamine no está instalado,
# NEXUS sigue funcionando normalmente con el motor por defecto de pandas.
try:
    import python_calamine  # noqa: F401
    EXCEL_ENGINE = "calamine"
except Exception:
    EXCEL_ENGINE = None

# Generación de PDF ejecutivo (opcional). Si fpdf2 no está instalado, el
# botón de PDF se oculta automáticamente y todo lo demás sigue igual.
try:
    from fpdf import FPDF
    HAS_PDF = True
except Exception:
    HAS_PDF = False


# ============================================================
# CONFIGURACIÓN
# ============================================================

APP_VERSION = "5.0.0"
DB_PATH = Path(__file__).with_name("nexus_data.db")
CONTACTS_PATH = Path(__file__).with_name("nexus_contacts.json")

SEDES = [
    "Centurión",
    "Perrokeros",
    "Puerto La Cruz",
    "Barcelona",
    "Matriz Makropetrol",
]

SEDE_ICON = {
    "Centurión": "◉",
    "Perrokeros": "◈",
    "Puerto La Cruz": "◇",
    "Barcelona": "▣",
    "Matriz Makropetrol": "◆",
}

SALES_MODES = {
    "Ventas del período": "periodo",
    "Ventas acumuladas": "acumuladas",
    "Ventas diarias": "diarias",
}

DEFAULT_CONTACTS = {
    site: {
        "supervisor": "",
        "whatsapp": "",
        "webhook": "",
    }
    for site in SEDES
}

# ============================================================
# DOCUMENTOS OPERATIVOS — CONFIGURACIÓN DE ENCABEZADO
# ============================================================
COMPANY_INFO = {
    "name": "LUBRIGAMA ORIENTE, C.A.",
    "address": "AV FUERZAS ARMADAS, CASA S/N, BARRIO LA ADUANA, BARCELONA, ANZOATEGUI",
    "phone": "0412-033.40.40",
    "rif": "J-50555333-0",
}

DOCUMENT_ORDER_PREFIX = {
    "purchase": "OC",
    "withdrawal": "RET",
    "redistribution": "TR",
    "alerts": "ALT",
}

SUPPLIER_DEFAULTS = {
    "name": "PROVEEDOR / SEGÚN INVENTARIO",
    "address": "",
    "phone": "",
}

DOC_MIME_PDF = "application/pdf"
DOC_MIME_HTML = "text/html"
DOC_MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

WEBHOOK_EXECUTOR = ThreadPoolExecutor(max_workers=2, thread_name_prefix="nexus-webhook")


# ============================================================
# ESTILO
# ============================================================

st.set_page_config(
    page_title="Makropetrol NEXUS | Intelligence",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
<style>
:root{
 --bg:#f5f7fb; --surface:rgba(255,255,255,.92); --surface-2:#ffffff;
 --line:#e5eaf2; --text:#152038; --muted:#768399; --blue:#3867f4;
 --blue2:#2448c9; --green:#11996a; --orange:#ed8b24; --red:#dc5157;
 --cyan:#16a7b7; --violet:#7656e8; --navy:#10192c;
 --shadow-sm:0 7px 24px rgba(29,45,75,.055);
 --shadow:0 18px 46px rgba(29,45,75,.085);
 --radius:20px;
}
*{box-sizing:border-box}
html,body,[class*="css"]{font-family:Inter,ui-sans-serif,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif}
.stApp{
 background:
  radial-gradient(circle at 8% -8%,rgba(56,103,244,.12),transparent 27%),
  radial-gradient(circle at 103% 4%,rgba(22,167,183,.08),transparent 25%),
  linear-gradient(150deg,#fafcff 0%,#f2f6fb 52%,#eef3f8 100%);
 color:var(--text)
}
.block-container{max-width:1580px;padding-top:1.2rem;padding-bottom:3rem}
header[data-testid="stHeader"]{background:transparent}
section[data-testid="stSidebar"]{
 background:linear-gradient(180deg,rgba(16,25,44,.985),rgba(22,35,61,.985));
 border-right:1px solid rgba(255,255,255,.06);
}
section[data-testid="stSidebar"] *{color:#edf3ff}
section[data-testid="stSidebar"] .stCaptionContainer p,
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p{color:#aebbd0}
section[data-testid="stSidebar"] hr{border-color:rgba(255,255,255,.09)}
section[data-testid="stSidebar"] .nav-group{color:#7f91ad!important}
section[data-testid="stSidebar"] button[kind="secondary"]{
 background:transparent;border:1px solid transparent;color:#c9d5e8;box-shadow:none
}
section[data-testid="stSidebar"] button[kind="secondary"]:hover{
 background:rgba(255,255,255,.06);border-color:rgba(255,255,255,.08)
}
section[data-testid="stSidebar"] button[kind="primary"]{
 background:linear-gradient(135deg,#446ff5,#2e55d9);border:1px solid rgba(255,255,255,.12);
 box-shadow:0 10px 24px rgba(31,70,190,.30)
}
section[data-testid="stSidebar"] [data-baseweb="input"],
section[data-testid="stSidebar"] [data-baseweb="select"]>div{background:rgba(255,255,255,.075)!important;border-color:rgba(255,255,255,.10)!important}
section[data-testid="stSidebar"] input{color:#f3f7ff!important}
section[data-testid="stSidebar"] input::placeholder{color:#8293ad!important}
h1,h2,h3,h4{font-family:Inter,ui-sans-serif,sans-serif!important;color:var(--text)!important;letter-spacing:-.035em}
.hero{
 background:linear-gradient(135deg,rgba(255,255,255,.96),rgba(247,250,255,.86));
 border:1px solid rgba(255,255,255,.96);box-shadow:var(--shadow);border-radius:30px;
 padding:28px 31px 24px;position:relative;overflow:hidden;isolation:isolate
}
.hero:before{content:"";position:absolute;inset:0;background:linear-gradient(90deg,rgba(56,103,244,.035),transparent 48%);z-index:-2}
.hero:after{content:"";position:absolute;width:390px;height:390px;right:-145px;top:-210px;border-radius:50%;background:radial-gradient(circle,rgba(56,103,244,.22),transparent 67%);z-index:-1}
.brand{font-size:2.05rem;font-weight:850;letter-spacing:-.055em}.brand span{color:var(--blue)}
.hero-sub{color:var(--muted);margin-top:5px;font-size:.93rem;max-width:900px}
.status-row{display:flex;gap:9px;align-items:center;flex-wrap:wrap;margin-top:15px;color:#66758c;font-size:.76rem}
.dot{width:8px;height:8px;border-radius:50%;background:#18b879;box-shadow:0 0 0 5px rgba(24,184,121,.11);animation:pulse 2s ease-in-out infinite}
@keyframes pulse{0%,100%{box-shadow:0 0 0 5px rgba(24,184,121,.10)}50%{box-shadow:0 0 0 9px rgba(24,184,121,.025)}}
.section-title{font-size:1.06rem;font-weight:850;margin:23px 0 10px;letter-spacing:-.025em}
.section-kicker{font-size:.69rem;text-transform:uppercase;letter-spacing:.11em;font-weight:850;color:var(--blue);margin-bottom:3px}
.kpi-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-top:17px}
.kpi{
 background:linear-gradient(145deg,rgba(255,255,255,.97),rgba(250,252,255,.90));
 border:1px solid rgba(255,255,255,.98);box-shadow:var(--shadow-sm);border-radius:21px;
 padding:18px 19px;min-height:121px;position:relative;overflow:hidden;transition:transform .18s ease,box-shadow .18s ease
}
.kpi:hover{transform:translateY(-2px);box-shadow:0 16px 38px rgba(29,45,75,.10)}
.kpi:before{content:"";position:absolute;left:0;top:0;bottom:0;width:4px;background:var(--blue)}
.kpi.green:before{background:var(--green)}.kpi.orange:before{background:var(--orange)}.kpi.red:before{background:var(--red)}.kpi.cyan:before{background:var(--cyan)}
.kpi-label{font-size:.71rem;color:var(--muted);font-weight:850;letter-spacing:.055em;text-transform:uppercase}
.kpi-value{font-size:1.72rem;font-weight:850;letter-spacing:-.045em;margin-top:9px}.kpi-note{font-size:.71rem;color:#98a4b5;margin-top:4px}
.panel{background:var(--surface);border:1px solid rgba(255,255,255,.97);box-shadow:var(--shadow-sm);border-radius:22px;padding:18px}
.glass-strip{background:rgba(255,255,255,.72);backdrop-filter:blur(12px);border:1px solid rgba(255,255,255,.88);border-radius:18px;padding:12px 14px;box-shadow:var(--shadow-sm)}
.badge{display:inline-flex;align-items:center;border-radius:999px;padding:5px 10px;font-size:.69rem;font-weight:850}
.badge-blue{background:#edf2ff;color:#355bd7}.badge-green{background:#e9f8f2;color:#137a52}.badge-orange{background:#fff2df;color:#b86a0c}.badge-red{background:#ffeded;color:#c43c43}.badge-gray{background:#eef1f5;color:#68758a}.badge-cyan{background:#e9f9fb;color:#147a86}
.stButton>button,.stDownloadButton>button{border-radius:13px;font-weight:780;min-height:43px;transition:transform .15s ease,box-shadow .15s ease}
.stButton>button:hover,.stDownloadButton>button:hover{transform:translateY(-1px)}
.stDownloadButton>button{border:0;background:linear-gradient(135deg,var(--blue),var(--blue2));color:#fff;box-shadow:0 9px 22px rgba(53,104,245,.22)}
[data-testid="stFileUploaderDropzone"]{border:1.5px dashed #c7d3e4!important;border-radius:17px!important;background:rgba(255,255,255,.75)!important;transition:border-color .2s ease,background .2s ease}
[data-testid="stFileUploaderDropzone"]:hover{border-color:#7d9bf5!important;background:#f8faff!important}
.stTabs [data-baseweb="tab-list"]{gap:6px;background:rgba(255,255,255,.72);border:1px solid var(--line);padding:6px;border-radius:16px;box-shadow:var(--shadow-sm)}
.stTabs [data-baseweb="tab"]{border-radius:11px;color:#7a8799;font-weight:760;padding:9px 15px}
.stTabs [aria-selected="true"]{background:#edf2ff!important;color:var(--blue)!important}
div[data-testid="stDataFrame"]{border:1px solid var(--line);border-radius:17px;overflow:hidden;box-shadow:0 7px 20px rgba(29,45,75,.035)}
[data-baseweb="input"],[data-baseweb="select"]>div{border-radius:12px!important}
.footer{text-align:center;color:#9aa6b6;font-size:.69rem;margin-top:34px}
.empty{background:rgba(255,255,255,.74);border:1px dashed #cbd5e4;border-radius:22px;padding:36px;text-align:center;color:#758195;box-shadow:var(--shadow-sm)}
.empty-title{font-weight:850;font-size:1.05rem;color:#273550}.empty-icon{font-size:2rem;margin-bottom:7px}
.nav-group{font-size:.64rem;font-weight:850;letter-spacing:.11em;text-transform:uppercase;margin:15px 0 3px 2px}
.gov-banner{display:flex;align-items:center;gap:8px;background:linear-gradient(135deg,#121d33,#22365b);color:#fff;border-radius:15px;padding:10px 14px;font-size:.78rem;font-weight:760;margin-bottom:10px;box-shadow:0 12px 30px rgba(16,25,44,.18)}
.health-big{display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;padding:20px 10px}
.health-big .num{font-size:3rem;font-weight:850;letter-spacing:-.055em;color:var(--blue)}
.health-big .lbl{font-size:.72rem;color:var(--muted);font-weight:850;letter-spacing:.075em;text-transform:uppercase;margin-top:2px}
.op-row{display:flex;justify-content:space-between;align-items:center;padding:9px 4px;border-bottom:1px solid var(--line);font-size:.85rem;gap:8px}.op-row:last-child{border-bottom:none}
.bar-track{background:#edf1f6;border-radius:8px;height:10px;overflow:hidden;flex:1;margin:0 10px}.bar-fill{height:100%;border-radius:8px;background:linear-gradient(90deg,var(--blue),var(--cyan))}
.opp-banner{display:flex;gap:22px;flex-wrap:wrap;background:linear-gradient(135deg,#e9f9f1,#edf5ff);border:1px solid #d8e9eb;border-radius:17px;padding:14px 18px;font-size:.85rem;font-weight:760;color:#245a4b;box-shadow:var(--shadow-sm)}
.scope-banner{display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap;background:linear-gradient(135deg,#f2f5ff,#eefbfb);border:1px solid #dfe7f4;border-radius:18px;padding:13px 16px;margin:10px 0 16px}
.scope-title{font-weight:850;color:#24324a}.scope-sub{font-size:.75rem;color:#748198;margin-top:2px}
.metric-chip{display:inline-flex;align-items:center;gap:6px;padding:7px 10px;border-radius:999px;background:white;border:1px solid var(--line);font-size:.72rem;font-weight:780;color:#536177}
@media(max-width:1000px){.kpi-grid{grid-template-columns:repeat(2,1fr)}}
@media(max-width:650px){.kpi-grid{grid-template-columns:1fr}.hero{padding:21px}.opp-banner{flex-direction:column;gap:8px}.brand{font-size:1.65rem}}
</style>
""",
    unsafe_allow_html=True,
)


# ============================================================
# UTILIDADES
# ============================================================

def money(v: Any) -> str:
    try:
        return f"${float(v):,.2f}"
    except Exception:
        return "$0.00"


def integer(v: Any) -> str:
    try:
        return f"{int(round(float(v))):,}"
    except Exception:
        return "0"


def safe_float(v: Any, default: float = 0.0) -> float:
    try:
        x = float(v)
        return x if np.isfinite(x) else default
    except Exception:
        return default


def normalize_text(v: Any) -> str:
    x = str(v).strip().lower()
    x = re.sub(r"\s+", " ", x)
    return x



def normalize_key(v: Any) -> str:
    """Clave estable para búsquedas de proveedor/marca y comparaciones humanas."""
    x = normalize_text(v)
    x = unicodedata.normalize("NFKD", x)
    x = "".join(ch for ch in x if not unicodedata.combining(ch))
    x = re.sub(r"[^a-z0-9]+", " ", x)
    return re.sub(r"\s+", " ", x).strip()


def safe_slug(v: Any, fallback: str = "reporte") -> str:
    x = normalize_key(v).replace(" ", "_")
    return x[:80] or fallback


def supplier_matches(query: str, suppliers: list[str], limit: int = 12) -> list[str]:
    q = normalize_key(query)
    if not q:
        return suppliers[:limit]
    starts, contains = [], []
    for supplier in suppliers:
        k = normalize_key(supplier)
        if k.startswith(q):
            starts.append(supplier)
        elif q in k:
            contains.append(supplier)
    return (starts + contains)[:limit]


def parse_number(v: Any) -> float:
    if pd.isna(v):
        return 0.0
    if isinstance(v, (int, float, np.integer, np.floating)):
        return float(v)
    s = str(v).strip().replace("$", "").replace("€", "").replace(" ", "")
    if not s:
        return 0.0
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        parts = s.split(",")
        if len(parts[-1]) in (1, 2):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        pass
    try:
        return float(s)
    except Exception:
        return 0.0


def empty_state(title: str, message: str, icon: str = "◌") -> None:
    st.markdown(
        f"""
        <div class="empty">
            <div class="empty-icon">{icon}</div>
            <div class="empty-title">{title}</div>
            <div>{message}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def kpi(label: str, value: str, note: str = "", tone: str = "blue") -> None:
    st.markdown(
        f"""
        <div class="kpi {tone}">
            <div class="kpi-label">{label}</div>
            <div class="kpi-value">{value}</div>
            <div class="kpi-note">{note}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def chart_layout(fig: go.Figure, height: int = 330) -> go.Figure:
    fig.update_layout(
        template="plotly_white",
        height=height,
        margin=dict(l=10, r=10, t=42, b=10),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="DM Sans", color="#59677b"),
        legend=dict(orientation="h", y=1.03, x=0),
        hoverlabel=dict(bgcolor="white", font_size=13, font_family="DM Sans"),
    )
    return fig


def present_columns(df: pd.DataFrame, preferred: list[str]) -> list[str]:
    return [c for c in preferred if c in df.columns]


def build_executive_frame(all_df: pd.DataFrame) -> pd.DataFrame:
    if all_df.empty:
        return pd.DataFrame()
    preferred = [
        "Sede", "Código", "Descripción", "ABC", "Estado", "Prioridad", "Acción",
        "Ventas", "Demanda Mensual", "Existencia", "Stock Mínimo", "Stock Máximo",
        "Punto de Reorden", "Compra Ajustada", "Retiro Almacén", "Cobertura (meses)",
        "Valor Inventario ($)", "Capital Inmovilizado ($)", "Costo",
    ]
    return all_df[present_columns(all_df, preferred)].copy()


def export_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")


def export_json_bytes(df: pd.DataFrame) -> bytes:
    return df.to_json(orient="records", force_ascii=False, indent=2).encode("utf-8")


@st.cache_data(show_spinner=False)
def export_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "Datos") -> bytes:
    from openpyxl.styles import Font, PatternFill

    out = io.BytesIO()
    safe_sheet = re.sub(r"[\\/*?:\[\]]", "_", str(sheet_name))[:31] or "Datos"
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=safe_sheet)
        ws = writer.book[safe_sheet]
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="3568F5")
        for col_cells in ws.columns:
            values = [str(c.value or "") for c in list(col_cells)[:120]]
            width = min(max(max((len(v) for v in values), default=8) + 2, 11), 42)
            ws.column_dimensions[col_cells[0].column_letter].width = width
    return out.getvalue()


@st.cache_data(show_spinner=False)
def build_interactive_html_report(
    all_df: pd.DataFrame,
    transfers: pd.DataFrame,
    alerts: pd.DataFrame,
    snapshot_label: str,
    score: int,
) -> bytes:
    if all_df.empty:
        body = "<h2>Sin datos disponibles</h2><p>Procese al menos un corte diario.</p>"
        return f"<!doctype html><html lang='es'><meta charset='utf-8'><body>{body}</body></html>".encode("utf-8")

    purchase_col = "Compra Ajustada" if "Compra Ajustada" in all_df.columns else "Compra Sugerida"
    site_summary = all_df.groupby("Sede").agg(
        Inventario=("Valor Inventario ($)", "sum"),
        Críticos=("Estado", lambda x: int(x.astype(str).str.startswith("CRÍTICO").sum())),
        Compras=(purchase_col, lambda x: int((x > 0).sum())),
        Retiros=("Retiro Almacén", lambda x: int((x > 0).sum())),
    ).reset_index()

    fig1 = px.bar(site_summary, x="Sede", y="Inventario", title="Valor de inventario por sede", color="Sede", color_discrete_sequence=NEXUS_PALETTE)
    fig2 = px.bar(site_summary, x="Sede", y=["Compras", "Retiros"], barmode="group", title="Acciones por sede", color_discrete_sequence=[NEXUS_ORANGE, NEXUS_BLUE])
    state = all_df["Estado"].value_counts().reset_index()
    state.columns = ["Estado", "Cantidad"]
    fig3 = px.pie(state, names="Estado", values="Cantidad", hole=.62, title="Distribución del estado logístico", color="Estado", color_discrete_map={"ÓPTIMO":NEXUS_GREEN,"CRÍTICO — SALDO NEGATIVO":NEXUS_RED,"CRÍTICO — COMPRAR":NEXUS_ORANGE,"SOBRESTOCK — RETIRAR":NEXUS_BLUE})

    top = all_df[all_df["Capital Inmovilizado ($)"] > 0].sort_values("Capital Inmovilizado ($)", ascending=False).head(12)
    fig4 = px.bar(top.sort_values("Capital Inmovilizado ($)"), x="Capital Inmovilizado ($)", y="Descripción", orientation="h", title="Top capital inmovilizado", color_discrete_sequence=[NEXUS_RED]) if not top.empty else None

    plotly_bundle_included = False

    def fig_html(fig):
        # Incluir plotly.js una sola vez reduce el HTML de ~20 MB a ~5 MB
        # cuando el dashboard contiene varios gráficos.
        nonlocal plotly_bundle_included
        include_js = "inline" if not plotly_bundle_included else False
        plotly_bundle_included = True
        return pio.to_html(
            fig, full_html=False, include_plotlyjs=include_js,
            config={"responsive": True, "displaylogo": False, "modeBarButtonsToRemove": ["lasso2d", "select2d"]},
        )

    exec_df = build_executive_frame(all_df).head(40)
    alert_html = alerts.head(30).to_html(index=False, classes="dataframe") if not alerts.empty else "<p>Sin alertas.</p>"
    transfer_html = transfers.head(30).to_html(index=False, classes="dataframe") if not transfers.empty else "<p>Sin transferencias sugeridas.</p>"
    table_html = exec_df.to_html(index=False, classes="dataframe")

    html = f"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Makropetrol NEXUS · Reporte Ejecutivo</title>
<style>
body{{margin:0;background:#f4f7fb;color:#172137;font-family:Inter,Segoe UI,Arial,sans-serif}}
.wrap{{max-width:1400px;margin:auto;padding:38px}}
.hero{{background:linear-gradient(135deg,#fff,#f0f5ff);border:1px solid #e2e8f2;border-radius:28px;padding:30px;box-shadow:0 18px 50px rgba(30,50,80,.10)}}
.brand{{font-size:28px;font-weight:800}} .brand span{{color:#3568f5}}
.muted{{color:#748198}} .score{{font-size:42px;font-weight:800;color:#3568f5}}
.kpis{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-top:18px}}
.kpi{{background:#fff;border:1px solid #e3e9f1;border-radius:20px;padding:18px;box-shadow:0 8px 24px rgba(30,50,80,.07)}}
.kpi small{{color:#7d899b;text-transform:uppercase;font-weight:800;font-size:11px}} .kpi strong{{display:block;font-size:27px;margin-top:6px}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-top:18px}} .panel{{background:#fff;border:1px solid #e3e9f1;border-radius:20px;padding:18px;box-shadow:0 8px 24px rgba(30,50,80,.06)}}
.dataframe{{width:100%;border-collapse:collapse;font-size:12px}} .dataframe th{{background:#3568f5;color:#fff;padding:9px;text-align:left}} .dataframe td{{padding:8px;border-bottom:1px solid #e5eaf1}} .dataframe tr:nth-child(even){{background:#f9fbfd}}
@media(max-width:900px){{.kpis,.grid{{grid-template-columns:1fr 1fr}}}} @media(max-width:600px){{.kpis,.grid{{grid-template-columns:1fr}} .wrap{{padding:18px}}}}
</style>

</head>
<body><div class="wrap">
<div class="hero"><div class="brand">MAKROPETROL <span>NEXUS</span></div>
<div class="muted">Reporte ejecutivo interactivo · Corte {snapshot_label}</div>
<div style="margin-top:16px"><span class="muted">Salud logística</span><div class="score">{score}/100</div></div></div>
<div class="kpis">
<div class="kpi"><small>SKUs</small><strong>{len(all_df):,}</strong></div>
<div class="kpi"><small>Críticos</small><strong>{int(all_df['Estado'].astype(str).str.startswith('CRÍTICO').sum()):,}</strong></div>
<div class="kpi"><small>Compra neta</small><strong>{int(all_df[purchase_col].sum()):,}</strong></div>
<div class="kpi"><small>Capital inmovilizado</small><strong>${float(all_df['Capital Inmovilizado ($)'].sum()):,.2f}</strong></div>
</div>
<div class="grid">
<div class="panel">{fig_html(fig1)}</div><div class="panel">{fig_html(fig2)}</div>
<div class="panel">{fig_html(fig3)}</div>{f'<div class="panel">{fig_html(fig4)}</div>' if fig4 is not None else '<div class="panel"><p class="muted">Sin capital inmovilizado detectado.</p></div>'}
</div>
<div class="panel" style="margin-top:18px"><h2>Vista ejecutiva de productos</h2>{table_html}</div>
<div class="panel" style="margin-top:18px"><h2>Alertas</h2>{alert_html}</div>
<div class="panel" style="margin-top:18px"><h2>Redistribución</h2>{transfer_html}</div>
</div></body></html>"""
    return html.encode("utf-8")


@st.cache_data(show_spinner=False)
def build_executive_excel(all_df: pd.DataFrame, snapshot_label: str, score: int) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    out = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Gerencial"

    navy, blue, white, gray = "18243A", "3568F5", "FFFFFF", "7B8799"
    ws.merge_cells("A1:F1")
    ws["A1"] = "MAKROPETROL NEXUS · RESUMEN GERENCIAL"
    ws["A1"].font = Font(size=18, bold=True, color=white)
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Corte {snapshot_label} · Salud logística {score}/100 · {health_label(score)}"
    ws["A2"].font = Font(size=10, italic=True, color=gray)

    if not all_df.empty:
        purchase_col = "Compra Ajustada" if "Compra Ajustada" in all_df.columns else "Compra Sugerida"
        by_site = all_df.groupby("Sede").agg(
            Inventario=("Valor Inventario ($)", "sum"),
            Críticos=("Estado", lambda x: int(x.astype(str).str.startswith("CRÍTICO").sum())),
            Compras=(purchase_col, lambda x: int((x > 0).sum())),
            Inmovilizado=("Capital Inmovilizado ($)", "sum"),
        ).reset_index()
        start = 4
        for j, col in enumerate(by_site.columns, 1):
            c = ws.cell(start, j, col)
            c.font = Font(bold=True, color=white)
            c.fill = PatternFill("solid", fgColor=blue)
        for i, row in enumerate(by_site.itertuples(index=False), start + 1):
            for j, val in enumerate(row, 1):
                ws.cell(i, j, val)
        ref = f"A{start}:{get_column_letter(len(by_site.columns))}{start + len(by_site)}"
        table = Table(displayName="ResumenGerencial", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
        for j in range(1, len(by_site.columns) + 1):
            ws.column_dimensions[get_column_letter(j)].width = 20

        exec_df = build_executive_frame(all_df).head(500)
        keep_cols = [c for c in ["Sede", "Código", "Descripción", "ABC", "Estado", "Acción", "Cobertura (meses)"] if c in exec_df.columns]
        sh2 = wb.create_sheet("Acciones recomendadas")
        for j, col in enumerate(keep_cols, 1):
            c = sh2.cell(1, j, col)
            c.font = Font(bold=True, color=white)
            c.fill = PatternFill("solid", fgColor=blue)
        for i, row in enumerate(exec_df[keep_cols].itertuples(index=False), 2):
            for j, val in enumerate(row, 1):
                sh2.cell(i, j, val)
        for j in range(1, len(keep_cols) + 1):
            sh2.column_dimensions[get_column_letter(j)].width = 22

    wb.save(out)
    return out.getvalue()


@st.cache_data(show_spinner=False)
def build_executive_pdf(all_df: pd.DataFrame, transfers: pd.DataFrame, snapshot_label: str, score: int) -> bytes | None:
    if not HAS_PDF:
        return None
    purchase_col = "Compra Ajustada" if "Compra Ajustada" in all_df.columns else "Compra Sugerida"
    critical = int(all_df["Estado"].astype(str).str.startswith("CRÍTICO").sum()) if not all_df.empty else 0
    purchases = int((all_df[purchase_col] > 0).sum()) if not all_df.empty else 0
    inv_value = float(all_df["Valor Inventario ($)"].sum()) if not all_df.empty else 0.0
    immobilized = float(all_df["Capital Inmovilizado ($)"].sum()) if not all_df.empty else 0.0
    avoided = float(transfers["Compra Evitada Estimada ($)"].sum()) if not transfers.empty else 0.0

    pdf = FPDF()
    pdf.add_page()
    pdf.set_fill_color(24, 36, 58)
    pdf.rect(0, 0, 210, 28, "F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_xy(10, 8)
    pdf.cell(0, 10, "MAKROPETROL NEXUS - Reporte Ejecutivo")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_xy(10, 18)
    pdf.cell(0, 8, f"Corte {snapshot_label}  ·  Salud logistica {score}/100 ({health_label(score)})")

    pdf.set_text_color(30, 30, 30)
    pdf.set_xy(10, 36)
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, "Indicadores clave", ln=True)
    pdf.set_font("Helvetica", "", 11)
    rows = [
        ("Valor de inventario", money(inv_value)),
        ("SKUs criticos", f"{critical:,}"),
        ("SKUs a comprar", f"{purchases:,}"),
        ("Capital inmovilizado", money(immobilized)),
        ("Compra evitada por redistribucion", money(avoided)),
    ]
    for label, val in rows:
        pdf.set_x(12)
        pdf.cell(90, 8, label)
        pdf.cell(0, 8, val, ln=True)

    if not all_df.empty:
        pdf.ln(4)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 8, "Inventario por sede", ln=True)
        pdf.set_font("Helvetica", "", 10)
        by_site = all_df.groupby("Sede")["Valor Inventario ($)"].sum().sort_values(ascending=False)
        for sede, val in by_site.items():
            pdf.set_x(12)
            pdf.cell(90, 7, str(sede))
            pdf.cell(0, 7, money(val), ln=True)

    pdf.ln(4)
    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(120, 120, 120)
    pdf.multi_cell(0, 6, f"Generado automaticamente por NEXUS v{APP_VERSION} el {datetime.now().strftime('%d/%m/%Y %H:%M')}.")

    return bytes(pdf.output())


def build_readme_text(snapshot_label: str) -> bytes:
    text = f"""MAKROPETROL NEXUS · Paquete de reporte
Corte: {snapshot_label}
Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}

Contenido del paquete:
- Dashboard_Gerencial.html  → dashboard interactivo con gráficos Plotly, se abre en cualquier navegador.
- Reporte_Gerencial.xlsx    → resumen ejecutivo por sede + acciones recomendadas.
- Reporte_Operativo.xlsx    → detalle completo (inventario, compras, retiros, redistribución, alertas).
- Datos.csv / Datos.json    → datos consolidados para Power BI, APIs o automatizaciones.
- Reporte_Ejecutivo.pdf     → una página lista para enviar por WhatsApp o correo (si fpdf2 está instalado).
- Documentos Operativos      → pedidos de compra por sede, retiros por sede, alertas por sede y redistribución multisede.
- ZIP Documentos Operativos → paquete separado por tienda + redistribución global.

NEXUS v{APP_VERSION} · datos base persistidos en SQLite · recálculo dinámico.
"""
    return text.encode("utf-8")


@st.cache_data(show_spinner=False)
def export_bundle(
    excel_bytes: bytes,
    all_df: pd.DataFrame,
    transfers: pd.DataFrame,
    alerts: pd.DataFrame,
    snapshot_label: str,
    score: int,
) -> bytes:
    html = build_interactive_html_report(all_df, transfers, alerts, snapshot_label, score)
    gerencial = build_executive_excel(all_df, snapshot_label, score)
    pdf_bytes = build_executive_pdf(all_df, transfers, snapshot_label, score)
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr(f"Reporte_Operativo_{snapshot_label}.xlsx", excel_bytes)
        z.writestr(f"Reporte_Gerencial_{snapshot_label}.xlsx", gerencial)
        z.writestr(f"Datos_{snapshot_label}.csv", export_csv_bytes(all_df))
        z.writestr(f"Datos_{snapshot_label}.json", export_json_bytes(all_df))
        z.writestr(f"Dashboard_Gerencial_{snapshot_label}.html", html)
        if pdf_bytes:
            z.writestr(f"Reporte_Ejecutivo_{snapshot_label}.pdf", pdf_bytes)
        z.writestr("README.txt", build_readme_text(snapshot_label))
    return out.getvalue()


# ============================================================
# SQLite — PERSISTENCIA DIARIA
# ============================================================

def db() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH, timeout=10)
    con.execute("PRAGMA journal_mode=WAL")
    return con


def init_db() -> None:
    with db() as con:
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS inventory_snapshots (
                snapshot_date TEXT NOT NULL,
                site TEXT NOT NULL,
                code TEXT NOT NULL,
                description TEXT,
                existence REAL NOT NULL DEFAULT 0,
                cost REAL NOT NULL DEFAULT 0,
                supplier TEXT,
                category TEXT,
                PRIMARY KEY(snapshot_date, site, code)
            )
            """
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS sales_snapshots (
                snapshot_date TEXT NOT NULL,
                site TEXT NOT NULL,
                code TEXT NOT NULL,
                description TEXT,
                sales REAL NOT NULL DEFAULT 0,
                PRIMARY KEY(snapshot_date, site, code)
            )
            """
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS processing_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                processed_at TEXT NOT NULL,
                snapshot_date TEXT NOT NULL,
                site TEXT NOT NULL,
                inventory_rows INTEGER,
                sales_rows INTEGER,
                user_label TEXT DEFAULT ''
            )
            """
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS supplier_inventory_snapshots (
                snapshot_date TEXT NOT NULL,
                site TEXT NOT NULL,
                supplier_key TEXT NOT NULL,
                supplier_name TEXT NOT NULL,
                code TEXT NOT NULL,
                description TEXT,
                existence REAL NOT NULL DEFAULT 0,
                cost REAL NOT NULL DEFAULT 0,
                category TEXT,
                PRIMARY KEY(snapshot_date, site, supplier_key, code)
            )
            """
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS supplier_sales_snapshots (
                snapshot_date TEXT NOT NULL,
                site TEXT NOT NULL,
                supplier_key TEXT NOT NULL,
                supplier_name TEXT NOT NULL,
                code TEXT NOT NULL,
                description TEXT,
                sales REAL NOT NULL DEFAULT 0,
                PRIMARY KEY(snapshot_date, site, supplier_key, code)
            )
            """
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS supplier_processing_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                processed_at TEXT NOT NULL,
                snapshot_date TEXT NOT NULL,
                site TEXT NOT NULL,
                supplier_key TEXT NOT NULL,
                supplier_name TEXT NOT NULL,
                inventory_rows INTEGER,
                sales_rows INTEGER
            )
            """
        )
        # Índices para MAX(fecha), filtros por sede/proveedor y cargas históricas.
        con.execute("CREATE INDEX IF NOT EXISTS idx_inv_site_date ON inventory_snapshots(site, snapshot_date)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_sales_site_date ON sales_snapshots(site, snapshot_date)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_sup_inv_key_site_date ON supplier_inventory_snapshots(supplier_key, site, snapshot_date)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_sup_sales_key_site_date ON supplier_sales_snapshots(supplier_key, site, snapshot_date)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_sup_log_key_date ON supplier_processing_log(supplier_key, snapshot_date)")
        con.commit()


def reset_database() -> None:
    with db() as con:
        con.execute("DELETE FROM inventory_snapshots")
        con.execute("DELETE FROM sales_snapshots")
        con.execute("DELETE FROM processing_log")
        con.execute("DELETE FROM supplier_inventory_snapshots")
        con.execute("DELETE FROM supplier_sales_snapshots")
        con.execute("DELETE FROM supplier_processing_log")
        con.commit()
    st.cache_data.clear()


def db_version() -> str:
    with db() as con:
        row = con.execute(
            "SELECT COALESCE(MAX(processed_at),'') FROM processing_log"
        ).fetchone()
    return str(row[0] or "")


def latest_snapshot_date(site: str) -> str | None:
    with db() as con:
        row = con.execute(
            "SELECT MAX(snapshot_date) FROM inventory_snapshots WHERE site=?",
            (site,),
        ).fetchone()
    return row[0] if row and row[0] else None


def previous_snapshot_date(site: str, current_date: str) -> str | None:
    with db() as con:
        row = con.execute(
            "SELECT MAX(snapshot_date) FROM sales_snapshots WHERE site=? AND snapshot_date<?",
            (site, current_date),
        ).fetchone()
    return row[0] if row and row[0] else None


def list_snapshot_dates(site: str | None = None) -> list[str]:
    with db() as con:
        if site:
            rows = con.execute(
                "SELECT DISTINCT snapshot_date FROM inventory_snapshots WHERE site=? ORDER BY snapshot_date DESC",
                (site,),
            ).fetchall()
        else:
            rows = con.execute(
                "SELECT DISTINCT snapshot_date FROM inventory_snapshots ORDER BY snapshot_date DESC"
            ).fetchall()
    return [r[0] for r in rows]


def save_snapshot(
    site: str,
    snapshot_date: str,
    inventory: pd.DataFrame,
    sales: pd.DataFrame,
) -> None:
    inv = inventory
    sal = sales

    inv_shaped = pd.DataFrame({
        "snapshot_date": snapshot_date,
        "site": site,
        "code": inv.get("Código", pd.Series("", index=inv.index)).astype(str),
        "description": inv.get("Descripción", pd.Series("", index=inv.index)).fillna("").astype(str)
        if "Descripción" in inv.columns else "",
        "existence": pd.to_numeric(inv.get("Existencia", pd.Series(0, index=inv.index)), errors="coerce").fillna(0.0),
        "cost": pd.to_numeric(inv.get("Costo", pd.Series(0, index=inv.index)), errors="coerce").fillna(0.0),
        "supplier": inv.get("Proveedor", pd.Series("", index=inv.index)).fillna("").astype(str)
        if "Proveedor" in inv.columns else "",
        "category": inv.get("Categoría", pd.Series("", index=inv.index)).fillna("").astype(str)
        if "Categoría" in inv.columns else "",
    })

    sal_shaped = pd.DataFrame({
        "snapshot_date": snapshot_date,
        "site": site,
        "code": sal.get("Código", pd.Series("", index=sal.index)).astype(str),
        "description": sal.get("Descripción", pd.Series("", index=sal.index)).fillna("").astype(str)
        if "Descripción" in sal.columns else "",
        "sales": pd.to_numeric(sal.get("Ventas", pd.Series(0, index=sal.index)), errors="coerce").fillna(0.0),
    })

    with db() as con:
        con.execute(
            "DELETE FROM inventory_snapshots WHERE site=? AND snapshot_date=?",
            (site, snapshot_date),
        )
        con.execute(
            "DELETE FROM sales_snapshots WHERE site=? AND snapshot_date=?",
            (site, snapshot_date),
        )

        if not inv_shaped.empty:
            inv_shaped.to_sql(
                "inventory_snapshots",
                con,
                if_exists="append",
                index=False,
                method="multi",
                chunksize=5000,
            )

        if not sal_shaped.empty:
            sal_shaped.to_sql(
                "sales_snapshots",
                con,
                if_exists="append",
                index=False,
                method="multi",
                chunksize=5000,
            )

        con.execute(
            "INSERT INTO processing_log(processed_at,snapshot_date,site,inventory_rows,sales_rows) VALUES(?,?,?,?,?)",
            (
                datetime.now().isoformat(timespec="seconds"),
                snapshot_date,
                site,
                len(inv_shaped),
                len(sal_shaped),
            ),
        )
        con.commit()

    st.cache_data.clear()


@st.cache_data(show_spinner=False)
def load_snapshot(site: str, snapshot_date: str) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    with db() as con:
        inv = pd.read_sql_query(
            "SELECT code AS Código, description AS Descripción, existence AS Existencia, cost AS Costo, supplier AS Proveedor, category AS Categoría FROM inventory_snapshots WHERE site=? AND snapshot_date=?",
            con,
            params=(site, snapshot_date),
        )
        sales = pd.read_sql_query(
            "SELECT code AS Código, description AS Descripción, sales AS Ventas FROM sales_snapshots WHERE site=? AND snapshot_date=?",
            con,
            params=(site, snapshot_date),
        )

        prev_date = previous_snapshot_date(site, snapshot_date)
        if prev_date:
            prev_sales = pd.read_sql_query(
                "SELECT code AS Código, sales AS Ventas FROM sales_snapshots WHERE site=? AND snapshot_date=?",
                con,
                params=(site, prev_date),
            )
        else:
            prev_sales = pd.DataFrame(columns=["Código", "Ventas"])
    return inv, sales, prev_sales



def supplier_db_version(supplier_name: str = "") -> str:
    key = normalize_key(supplier_name)
    with db() as con:
        if key:
            row = con.execute(
                "SELECT COALESCE(MAX(processed_at),'') FROM supplier_processing_log WHERE supplier_key=?",
                (key,),
            ).fetchone()
        else:
            row = con.execute(
                "SELECT COALESCE(MAX(processed_at),'') FROM supplier_processing_log"
            ).fetchone()
    return str(row[0] or "")


@st.cache_data(show_spinner=False)
def list_known_suppliers(_db_version: str, _supplier_version: str) -> list[str]:
    with db() as con:
        rows = con.execute(
            """
            SELECT supplier_name FROM supplier_inventory_snapshots
            WHERE TRIM(COALESCE(supplier_name,''))<>''
            UNION
            SELECT supplier FROM inventory_snapshots
            WHERE TRIM(COALESCE(supplier,''))<>''
            """
        ).fetchall()
    names = sorted({str(r[0]).strip() for r in rows if r and str(r[0]).strip()}, key=lambda x: normalize_key(x))
    return names


def latest_supplier_snapshot_date(site: str, supplier_name: str) -> str | None:
    key = normalize_key(supplier_name)
    if not key:
        return None
    with db() as con:
        row = con.execute(
            "SELECT MAX(snapshot_date) FROM supplier_inventory_snapshots WHERE site=? AND supplier_key=?",
            (site, key),
        ).fetchone()
    return row[0] if row and row[0] else None


def previous_supplier_snapshot_date(site: str, supplier_name: str, current_date: str) -> str | None:
    key = normalize_key(supplier_name)
    with db() as con:
        row = con.execute(
            "SELECT MAX(snapshot_date) FROM supplier_sales_snapshots WHERE site=? AND supplier_key=? AND snapshot_date<?",
            (site, key, current_date),
        ).fetchone()
    return row[0] if row and row[0] else None


def save_supplier_snapshot(
    site: str,
    snapshot_date: str,
    supplier_name: str,
    inventory: pd.DataFrame,
    sales: pd.DataFrame,
) -> None:
    supplier_name = str(supplier_name).strip()
    supplier_key = normalize_key(supplier_name)
    if not supplier_key:
        raise ValueError("Debes indicar un proveedor o marca.")

    inv = inventory.copy()
    sal = sales.copy()
    inv["Proveedor"] = supplier_name

    inv_shaped = pd.DataFrame({
        "snapshot_date": snapshot_date,
        "site": site,
        "supplier_key": supplier_key,
        "supplier_name": supplier_name,
        "code": inv.get("Código", pd.Series("", index=inv.index)).astype(str),
        "description": inv.get("Descripción", pd.Series("", index=inv.index)).fillna("").astype(str),
        "existence": pd.to_numeric(inv.get("Existencia", pd.Series(0, index=inv.index)), errors="coerce").fillna(0.0),
        "cost": pd.to_numeric(inv.get("Costo", pd.Series(0, index=inv.index)), errors="coerce").fillna(0.0),
        "category": inv.get("Categoría", pd.Series("", index=inv.index)).fillna("").astype(str),
    })
    sal_shaped = pd.DataFrame({
        "snapshot_date": snapshot_date,
        "site": site,
        "supplier_key": supplier_key,
        "supplier_name": supplier_name,
        "code": sal.get("Código", pd.Series("", index=sal.index)).astype(str),
        "description": sal.get("Descripción", pd.Series("", index=sal.index)).fillna("").astype(str),
        "sales": pd.to_numeric(sal.get("Ventas", pd.Series(0, index=sal.index)), errors="coerce").fillna(0.0),
    })

    with db() as con:
        con.execute(
            "DELETE FROM supplier_inventory_snapshots WHERE site=? AND snapshot_date=? AND supplier_key=?",
            (site, snapshot_date, supplier_key),
        )
        con.execute(
            "DELETE FROM supplier_sales_snapshots WHERE site=? AND snapshot_date=? AND supplier_key=?",
            (site, snapshot_date, supplier_key),
        )
        if not inv_shaped.empty:
            inv_shaped.to_sql("supplier_inventory_snapshots", con, if_exists="append", index=False, method="multi", chunksize=5000)
        if not sal_shaped.empty:
            sal_shaped.to_sql("supplier_sales_snapshots", con, if_exists="append", index=False, method="multi", chunksize=5000)
        con.execute(
            """
            INSERT INTO supplier_processing_log(
                processed_at,snapshot_date,site,supplier_key,supplier_name,inventory_rows,sales_rows
            ) VALUES(?,?,?,?,?,?,?)
            """,
            (
                datetime.now().isoformat(timespec="seconds"), snapshot_date, site, supplier_key,
                supplier_name, len(inv_shaped), len(sal_shaped),
            ),
        )
        con.commit()
    st.cache_data.clear()


@st.cache_data(show_spinner=False)
def load_supplier_snapshot(site: str, supplier_name: str, snapshot_date: str) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    key = normalize_key(supplier_name)
    with db() as con:
        inv = pd.read_sql_query(
            """
            SELECT code AS Código, description AS Descripción, existence AS Existencia,
                   cost AS Costo, supplier_name AS Proveedor, category AS Categoría
            FROM supplier_inventory_snapshots
            WHERE site=? AND supplier_key=? AND snapshot_date=?
            """,
            con, params=(site, key, snapshot_date),
        )
        sales = pd.read_sql_query(
            """
            SELECT code AS Código, description AS Descripción, sales AS Ventas
            FROM supplier_sales_snapshots
            WHERE site=? AND supplier_key=? AND snapshot_date=?
            """,
            con, params=(site, key, snapshot_date),
        )
        prev_date = previous_supplier_snapshot_date(site, supplier_name, snapshot_date)
        if prev_date:
            prev_sales = pd.read_sql_query(
                """
                SELECT code AS Código, sales AS Ventas
                FROM supplier_sales_snapshots
                WHERE site=? AND supplier_key=? AND snapshot_date=?
                """,
                con, params=(site, key, prev_date),
            )
        else:
            prev_sales = pd.DataFrame(columns=["Código", "Ventas"])
    return inv, sales, prev_sales


@st.cache_data(show_spinner=False)
def build_rotation_history(_db_version: str, sales_mode: str) -> pd.DataFrame:
    """Última fecha con movimiento positivo por sede/SKU usando el historial guardado."""
    with db() as con:
        hist = pd.read_sql_query(
            "SELECT snapshot_date, site AS Sede, code AS Código, sales AS Ventas FROM sales_snapshots ORDER BY site, code, snapshot_date",
            con,
        )
    if hist.empty:
        return pd.DataFrame(columns=["Sede", "Código", "Primera fecha", "Última fecha", "Última venta positiva"])
    hist["snapshot_date"] = pd.to_datetime(hist["snapshot_date"], errors="coerce")
    hist["Ventas"] = pd.to_numeric(hist["Ventas"], errors="coerce").fillna(0.0)
    hist = hist.dropna(subset=["snapshot_date"])
    if sales_mode == "acumuladas":
        hist["Movimiento"] = hist.groupby(["Sede", "Código"], observed=True)["Ventas"].diff().fillna(0).clip(lower=0)
    else:
        hist["Movimiento"] = hist["Ventas"].clip(lower=0)
    base = hist.groupby(["Sede", "Código"], observed=True).agg(
        **{"Primera fecha": ("snapshot_date", "min"), "Última fecha": ("snapshot_date", "max")}
    ).reset_index()
    positive = hist[hist["Movimiento"] > 0].groupby(["Sede", "Código"], observed=True)["snapshot_date"].max().rename("Última venta positiva").reset_index()
    return base.merge(positive, on=["Sede", "Código"], how="left")


# ============================================================
# EXCEL INPUT — NORMALIZACIÓN ROBUSTA
# ============================================================

def find_header_row(raw: pd.DataFrame) -> int:
    keywords = [
        "código", "codigo", "cod", "artículo", "articulo", "sku",
        "descripción", "descripcion", "existencia", "stock", "saldo",
        "cantidad", "ventas", "costo", "proveedor", "categoria", "categoría",
    ]
    best_row, best_score = 0, 0
    for idx, row in raw.iterrows():
        vals = [normalize_text(v) for v in row.tolist() if pd.notna(v)]
        score = sum(1 for kw in keywords if any(kw == v or kw in v for v in vals))
        if score > best_score:
            best_row, best_score = idx, score
    return best_row


def standardize_columns(df: pd.DataFrame, kind: str) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    aliases = {
        "Código": ["código", "codigo", "cod", "cod_articulo", "cod. artículo", "artículo", "articulo", "sku", "codigo producto"],
        "Descripción": ["descripción", "descripcion", "desc", "producto", "nombre", "detalle"],
        "Ventas": ["cantidad", "ventas", "salidas", "venta", "unidades vendidas", "cantidad vendida", "movimiento"],
        "Existencia": ["existencia", "stock", "saldo", "inventario", "disponible", "existencias"],
        "Costo": ["costo", "precio", "costo unitario", "coste", "precio costo", "precio de costo"],
        "Proveedor": ["proveedor", "marca/proveedor", "vendor", "supplier"],
        "Categoría": ["categoría", "categoria", "familia", "rubro", "grupo", "linea", "línea"],
    }
    rename = {}
    for col in df.columns:
        c = normalize_text(col)
        for target, variants in aliases.items():
            if c in variants or any(c.startswith(v + " ") for v in variants):
                if target not in rename.values():
                    rename[col] = target
                break
    df = df.rename(columns=rename)

    for col in ["Ventas", "Existencia", "Costo"]:
        if col in df.columns:
            df[col] = df[col].map(parse_number)

    if "Código" in df.columns:
        df["Código"] = (
            df["Código"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        )
        df = df[(df["Código"] != "") & (df["Código"].str.lower() != "nan")]

    if "Descripción" not in df.columns:
        df["Descripción"] = ""
    df["Descripción"] = df["Descripción"].fillna("").astype(str).str.strip()

    if kind == "inventario":
        if "Existencia" not in df.columns:
            df["Existencia"] = 0.0
        if "Costo" not in df.columns:
            df["Costo"] = 0.0
    else:
        if "Ventas" not in df.columns:
            df["Ventas"] = 0.0

    if "Código" in df.columns:
        agg: dict[str, str] = {}
        for col in df.columns:
            if col == "Código":
                continue
            if col in ["Ventas", "Existencia"]:
                agg[col] = "sum"
            else:
                agg[col] = "first"
        df = df.groupby("Código", as_index=False).agg(agg)

    return df.reset_index(drop=True)


def _read_excel_fast(bio: io.BytesIO, **kwargs) -> pd.DataFrame:
    if EXCEL_ENGINE:
        try:
            bio.seek(0)
            return pd.read_excel(bio, engine=EXCEL_ENGINE, **kwargs)
        except Exception:
            pass
    bio.seek(0)
    return pd.read_excel(bio, **kwargs)


@st.cache_data(show_spinner=False)
def parse_uploaded_excel(raw_bytes: bytes, kind: str) -> tuple[pd.DataFrame, int]:
    bio = io.BytesIO(raw_bytes)
    # El encabezado suele estar al inicio: leer solo una muestra evita parsear
    # el Excel completo dos veces y acelera notablemente archivos grandes.
    raw = _read_excel_fast(bio, header=None, nrows=80)
    header_row = find_header_row(raw)
    df = _read_excel_fast(bio, skiprows=header_row)
    df = standardize_columns(df, kind)
    df = downcast_dataframe(df)
    return df, header_row


def downcast_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    category_like = ["Estado", "Prioridad", "ABC", "Acción", "Sede", "Categoría",
                      "Proveedor", "Tiene costo", "Stock negativo", "Sin movimiento"]
    for col in category_like:
        if col in df.columns and df[col].dtype == object:
            if df[col].nunique(dropna=False) < max(50, len(df) // 2):
                df[col] = df[col].astype("category")

    integer_like = [
        "Ventas", "Demanda Mensual", "Existencia", "Stock Mínimo", "Stock Máximo",
        "Demanda Lead Time", "Stock Seguridad", "Punto de Reorden",
        "Compra Sugerida", "Retiro Almacén", "Compra Ajustada",
        "Transferencias Recibidas", "Transferencias Enviadas",
    ]
    for col in integer_like:
        if col in df.columns:
            s = pd.to_numeric(df[col], errors="coerce")
            if s.notna().all():
                df[col] = pd.to_numeric(s.round(), downcast="integer")
    return df


# ============================================================
# MOTOR LOGÍSTICO
# ============================================================

@st.cache_data(show_spinner=False)
def calculate_site(
    inv: pd.DataFrame,
    sales: pd.DataFrame,
    prev_sales: pd.DataFrame,
    months_history: int,
    min_coverage: float,
    max_coverage: float,
    lead_time_days: int,
    safety_days: int,
    abc_basis: str,
    sales_mode: str,
    rolling_days: int,
) -> pd.DataFrame:
    inv = inv.copy()
    sales = sales.copy()
    prev_sales = prev_sales.copy()

    if "Existencia" not in inv.columns:
        inv["Existencia"] = 0.0
    if "Costo" not in inv.columns:
        inv["Costo"] = 0.0
    if "Descripción" not in inv.columns:
        inv["Descripción"] = ""
    if "Proveedor" not in inv.columns:
        inv["Proveedor"] = ""
    if "Categoría" not in inv.columns:
        inv["Categoría"] = ""
    if "Ventas" not in sales.columns:
        sales["Ventas"] = 0.0

    current = sales[["Código", "Ventas"]].copy()
    if sales_mode == "acumuladas":
        prev = prev_sales.rename(columns={"Ventas": "Ventas Anteriores"})
        current = current.merge(prev, on="Código", how="left")
        current["Ventas del Corte"] = np.maximum(
            current["Ventas"].fillna(0) - current["Ventas Anteriores"].fillna(0), 0
        )
        current = current[["Código", "Ventas del Corte"]]
        current = current.rename(columns={"Ventas del Corte": "Ventas"})

    df = inv.merge(current, on="Código", how="outer", suffixes=("_inv", "_ventas"))

    for col in ["Existencia", "Costo"]:
        if col not in df.columns:
            df[col] = 0.0
    if "Ventas" not in df.columns:
        df["Ventas"] = 0.0
    if "Descripción_inv" in df.columns or "Descripción_ventas" in df.columns:
        d1 = df.get("Descripción_inv", pd.Series("", index=df.index)).fillna("")
        d2 = df.get("Descripción_ventas", pd.Series("", index=df.index)).fillna("")
        df["Descripción"] = d1.where(d1.astype(str).str.strip() != "", d2)
    elif "Descripción" not in df.columns:
        df["Descripción"] = ""

    df["Ventas"] = pd.to_numeric(df["Ventas"], errors="coerce").fillna(0)
    df["Existencia"] = pd.to_numeric(df["Existencia"], errors="coerce").fillna(0)
    df["Costo"] = pd.to_numeric(df["Costo"], errors="coerce").fillna(0)

    months_history = max(1, int(months_history))
    min_coverage = max(0.1, float(min_coverage))
    max_coverage = max(min_coverage, float(max_coverage))
    lead_time_days = max(0, int(lead_time_days))
    safety_days = max(0, int(safety_days))
    rolling_days = max(1, int(rolling_days))

    if sales_mode == "diarias":
        monthly_demand = df["Ventas"] * 30
    else:
        monthly_demand = df["Ventas"] / months_history

    df["Demanda Mensual"] = np.ceil(monthly_demand).astype(int)
    df["Demanda Diaria"] = (monthly_demand / 30).round(2)

    df["Stock Mínimo"] = np.where(
        df["Demanda Mensual"] > 0,
        np.ceil(df["Demanda Mensual"] * min_coverage),
        1,
    ).astype(int)
    df["Stock Máximo"] = np.where(
        df["Demanda Mensual"] > 0,
        np.ceil(df["Demanda Mensual"] * max_coverage),
        2,
    ).astype(int)

    df["Demanda Lead Time"] = np.ceil(df["Demanda Diaria"] * lead_time_days).astype(int)
    df["Stock Seguridad"] = np.ceil(df["Demanda Diaria"] * safety_days).astype(int)
    df["Punto de Reorden"] = np.maximum(
        df["Stock Mínimo"], df["Demanda Lead Time"] + df["Stock Seguridad"]
    ).astype(int)

    df["Compra Sugerida"] = np.where(
        df["Existencia"] < df["Punto de Reorden"],
        np.maximum(df["Stock Máximo"] - df["Existencia"], 0),
        0,
    ).astype(int)

    df["Retiro Almacén"] = np.where(
        df["Existencia"] > df["Stock Máximo"],
        np.maximum(df["Existencia"] - df["Stock Máximo"], 0),
        0,
    ).astype(int)

    positive_stock = np.maximum(df["Existencia"], 0)
    df["Cobertura (meses)"] = np.where(
        df["Demanda Mensual"] > 0,
        (positive_stock / df["Demanda Mensual"]).round(1),
        np.nan,
    )
    df["Rotación estimada (x/mes)"] = np.where(
        positive_stock > 0,
        (df["Demanda Mensual"] / positive_stock).round(2),
        0,
    )

    df["Valor Inventario ($)"] = (df["Existencia"] * df["Costo"]).round(2)
    df["Capital Inmovilizado ($)"] = (df["Retiro Almacén"] * df["Costo"]).round(2)
    df["Costo Compra Estimada ($)"] = (df["Compra Sugerida"] * df["Costo"]).round(2)
    df["Valor Movimiento ($)"] = (df["Demanda Mensual"] * df["Costo"]).round(2)

    if abc_basis == "Inventario":
        abc_value = np.maximum(df["Valor Inventario ($)"], 0)
    elif abc_basis == "Capital inmovilizado":
        abc_value = np.maximum(df["Capital Inmovilizado ($)"], 0)
    else:
        abc_value = np.maximum(df["Valor Movimiento ($)"], 0)

    total_abc = float(abc_value.sum())
    if total_abc > 0:
        order = np.argsort(-abc_value.to_numpy())
        cumulative = np.zeros(len(df))
        ordered = abc_value.to_numpy()[order]
        cumulative[order] = np.cumsum(ordered) / total_abc
        df["ABC"] = np.select(
            [cumulative <= 0.80, cumulative <= 0.95],
            ["A", "B"],
            default="C",
        )
    else:
        df["ABC"] = "C"

    df["Tiene costo"] = np.where(df["Costo"] > 0, "Sí", "No")
    df["Stock negativo"] = np.where(df["Existencia"] < 0, "Sí", "No")
    df["Sin movimiento"] = np.where((df["Ventas"] <= 0) & (df["Existencia"] > 0), "Sí", "No")

    df["Estado"] = np.select(
        [
            df["Existencia"] < 0,
            df["Existencia"] < df["Punto de Reorden"],
            df["Retiro Almacén"] > 0,
        ],
        [
            "CRÍTICO — SALDO NEGATIVO",
            "CRÍTICO — COMPRAR",
            "SOBRESTOCK — RETIRAR",
        ],
        default="ÓPTIMO",
    )

    df["Prioridad"] = np.select(
        [
            df["Existencia"] < 0,
            (df["Compra Sugerida"] > 0) & (df["ABC"] == "A"),
            df["Compra Sugerida"] > 0,
            (df["Retiro Almacén"] > 0) & (df["ABC"] == "A"),
            df["Retiro Almacén"] > 0,
        ],
        ["CRÍTICA", "ALTA", "MEDIA", "ALTA", "MEDIA"],
        default="BAJA",
    )

    df["Acción"] = np.select(
        [
            df["Existencia"] < 0,
            df["Compra Sugerida"] > 0,
            df["Retiro Almacén"] > 0,
        ],
        ["REVISAR SALDO / ABASTECER", "COMPRAR", "RETIRAR / REDISTRIBUIR"],
        default="NO HACER NADA",
    )

    keep = [
        "Código", "Descripción", "Proveedor", "Categoría", "Ventas", "Demanda Mensual",
        "Demanda Diaria", "Existencia", "Costo", "Stock Mínimo", "Stock Máximo",
        "Demanda Lead Time", "Stock Seguridad", "Punto de Reorden", "Compra Sugerida",
        "Retiro Almacén", "Cobertura (meses)", "Rotación estimada (x/mes)", "Valor Inventario ($)",
        "Capital Inmovilizado ($)", "Costo Compra Estimada ($)", "Valor Movimiento ($)",
        "ABC", "Tiene costo", "Stock negativo", "Sin movimiento", "Estado", "Prioridad", "Acción",
    ]
    for col in keep:
        if col not in df.columns:
            df[col] = ""
    result = df[keep].copy()
    integer_cols = [
        "Ventas", "Demanda Mensual", "Existencia", "Stock Mínimo", "Stock Máximo",
        "Demanda Lead Time", "Stock Seguridad", "Punto de Reorden", "Compra Sugerida",
        "Retiro Almacén", "Transferencias Recibidas", "Transferencias Enviadas", "Compra Ajustada"
    ]
    for col in integer_cols:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors="coerce").fillna(0).round().astype(int)
    money_cols = [
        "Costo", "Valor Inventario ($)", "Capital Inmovilizado ($)",
        "Costo Compra Estimada ($)", "Valor Movimiento ($)"
    ]
    for col in money_cols:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors="coerce").fillna(0).round(2)
    if "Demanda Diaria" in result.columns:
        result["Demanda Diaria"] = pd.to_numeric(result["Demanda Diaria"], errors="coerce").fillna(0).round(2)
    if "Cobertura (meses)" in result.columns:
        result["Cobertura (meses)"] = pd.to_numeric(result["Cobertura (meses)"], errors="coerce").round(1)
    if "Rotación estimada (x/mes)" in result.columns:
        result["Rotación estimada (x/mes)"] = pd.to_numeric(result["Rotación estimada (x/mes)"], errors="coerce").fillna(0).round(2)
    result = result.sort_values(
        by=["Prioridad", "Capital Inmovilizado ($)", "Demanda Mensual"],
        ascending=[True, False, False],
    ).reset_index(drop=True)
    return downcast_dataframe(result)


# ============================================================
# REDISTRIBUCIÓN MULTISEDE
# ============================================================

def build_redistribution(site_results: dict[str, pd.DataFrame]) -> pd.DataFrame:
    frames = []
    for site, df in site_results.items():
        if df is None or df.empty:
            continue
        x = df[[
            "Código", "Descripción", "Existencia", "Stock Máximo",
            "Stock Mínimo", "Costo", "Demanda Mensual"
        ]].copy()
        x.insert(0, "Sede", site)
        frames.append(x)
    if not frames:
        return pd.DataFrame()

    all_df = pd.concat(frames, ignore_index=True)
    suggestions: list[dict[str, Any]] = []

    for code, group in all_df.groupby("Código"):
        donors = group.copy()
        receivers = group.copy()
        donors["Exceso"] = np.maximum(donors["Existencia"] - donors["Stock Máximo"], 0)
        receivers["Déficit"] = np.maximum(receivers["Stock Mínimo"] - receivers["Existencia"], 0)
        donors = donors[donors["Exceso"] > 0].sort_values("Exceso", ascending=False)
        receivers = receivers[receivers["Déficit"] > 0].sort_values("Déficit", ascending=False)

        for r in receivers.itertuples(index=False):
            need = float(r.Déficit)
            if need <= 0:
                continue
            for d in donors.itertuples(index=False):
                if d.Sede == r.Sede:
                    continue
                available = float(d.Exceso)
                if available <= 0:
                    continue
                qty = int(np.floor(min(need, available)))
                if qty <= 0:
                    continue
                cost = safe_float(r.Costo or d.Costo)
                suggestions.append(
                    {
                        "Código": code,
                        "Descripción": r.Descripción or d.Descripción,
                        "Origen": d.Sede,
                        "Destino": r.Sede,
                        "Existencia Origen": int(round(d.Existencia)),
                        "Necesidad Destino": int(round(r.Déficit)),
                        "Unidades Sugeridas": qty,
                        "Costo Unitario ($)": cost,
                        "Compra Evitada Estimada ($)": round(qty * cost, 2),
                        "Estado": "RECOMENDADA",
                    }
                )
                need -= qty
                idx = donors.index[donors["Sede"] == d.Sede][0]
                donors.loc[idx, "Exceso"] -= qty
                if need <= 0:
                    break

    if not suggestions:
        return pd.DataFrame(
            columns=[
                "Código", "Descripción", "Origen", "Destino", "Existencia Origen",
                "Necesidad Destino", "Unidades Sugeridas", "Costo Unitario ($)",
                "Compra Evitada Estimada ($)", "Estado"
            ]
        )

    return pd.DataFrame(suggestions).sort_values(
        "Compra Evitada Estimada ($)", ascending=False
    ).reset_index(drop=True)


def apply_redistribution_to_results(
    site_results: dict[str, pd.DataFrame],
    transfers: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    out = {site: df.copy() for site, df in site_results.items()}
    if transfers.empty:
        for site, df in out.items():
            df["Transferencias Recibidas"] = 0
            df["Transferencias Enviadas"] = 0
            df["Compra Ajustada"] = df["Compra Sugerida"]
        return out

    received = (
        transfers.groupby(["Destino", "Código"])["Unidades Sugeridas"]
        .sum().rename("Transferencias Recibidas").reset_index()
        .rename(columns={"Destino": "Sede"})
    )
    sent = (
        transfers.groupby(["Origen", "Código"])["Unidades Sugeridas"]
        .sum().rename("Transferencias Enviadas").reset_index()
        .rename(columns={"Origen": "Sede"})
    )

    for site, df in out.items():
        df.drop(columns=["Transferencias Recibidas", "Transferencias Enviadas"], errors="ignore", inplace=True)
        df = df.merge(received[received["Sede"] == site][["Código", "Transferencias Recibidas"]], on="Código", how="left")
        df = df.merge(sent[sent["Sede"] == site][["Código", "Transferencias Enviadas"]], on="Código", how="left")
        df["Transferencias Recibidas"] = df["Transferencias Recibidas"].fillna(0).astype(int)
        df["Transferencias Enviadas"] = df["Transferencias Enviadas"].fillna(0).astype(int)
        df["Compra Ajustada"] = (df["Compra Sugerida"] - df["Transferencias Recibidas"]).clip(lower=0)

        recibe_y_compraba = (df["Transferencias Recibidas"] > 0) & (df["Compra Sugerida"] > 0)
        envia_y_retiraba = (df["Transferencias Enviadas"] > 0) & (df["Retiro Almacén"] > 0)
        df["Acción"] = np.select(
            [recibe_y_compraba, envia_y_retiraba],
            ["TRANSFERIR DESDE OTRA SEDE", "TRANSFERIR A OTRA SEDE"],
            default=df["Acción"],
        )
        out[site] = df
    return out


# ============================================================
# MÉTRICAS GLOBALES
# ============================================================

def combine_sites(site_results: dict[str, pd.DataFrame]) -> pd.DataFrame:
    frames = []
    for site, df in site_results.items():
        if df is None or df.empty:
            continue
        x = df.copy()
        x.insert(0, "Sede", site)
        frames.append(x)
    if not frames:
        return pd.DataFrame()
    return downcast_dataframe(pd.concat(frames, ignore_index=True))


def health_score(df: pd.DataFrame) -> int:
    if df.empty:
        return 0
    n = len(df)
    critical = float((df["Estado"].str.startswith("CRÍTICO")).sum()) / n
    overstock = float((df["Estado"] == "SOBRESTOCK — RETIRAR").sum()) / n
    dead = float((df["Sin movimiento"] == "Sí").sum()) / n
    no_cost = float((df["Tiene costo"] == "No").sum()) / n
    negative = float((df["Stock negativo"] == "Sí").sum()) / n
    score = 100 - (critical * 45 + overstock * 25 + dead * 15 + no_cost * 10 + negative * 5)
    return int(np.clip(round(score), 0, 100))


def health_label(score: int) -> str:
    if score >= 85:
        return "SALUDABLE"
    if score >= 70:
        return "VIGILAR"
    if score >= 50:
        return "RIESGO"
    return "CRÍTICA"


def build_alerts(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["Tipo", "Severidad", "Sede", "Código", "Descripción", "Detalle"])

    base_cols = ["Sede", "Código", "Descripción"]
    pieces = []

    m = df["Estado"] == "CRÍTICO — SALDO NEGATIVO"
    if m.any():
        sub = df.loc[m, base_cols].copy()
        sub["Tipo"] = "Saldo negativo"
        sub["Severidad"] = "CRÍTICA"
        sub["Detalle"] = "Existencia " + df.loc[m, "Existencia"].round().astype(int).astype(str) + " · revisar origen y abastecer."
        pieces.append(sub)

    m = df["Estado"] == "CRÍTICO — COMPRAR"
    if m.any():
        sub = df.loc[m, base_cols].copy()
        sub["Tipo"] = "Stock crítico"
        sub["Severidad"] = "ALTA"
        sub["Detalle"] = (
            "Existencia " + df.loc[m, "Existencia"].round().astype(int).astype(str)
            + " · punto de reorden " + df.loc[m, "Punto de Reorden"].astype(int).astype(str) + "."
        )
        pieces.append(sub)

    m = df["Estado"] == "SOBRESTOCK — RETIRAR"
    if m.any():
        sub = df.loc[m, base_cols].copy()
        sub["Tipo"] = "Sobrestock"
        sub["Severidad"] = "MEDIA"
        sub["Detalle"] = "Retirar " + df.loc[m, "Retiro Almacén"].round().astype(int).astype(str) + " unidades."
        pieces.append(sub)

    m = (df["Sin movimiento"] == "Sí") & (df["Existencia"] > 0)
    if m.any():
        sub = df.loc[m, base_cols].copy()
        sub["Tipo"] = "Sin movimiento"
        sub["Severidad"] = "MEDIA"
        sub["Detalle"] = "Hay " + df.loc[m, "Existencia"].round().astype(int).astype(str) + " unidades sin venta en el corte."
        pieces.append(sub)

    m = df["Tiene costo"] == "No"
    if m.any():
        sub = df.loc[m, base_cols].copy()
        sub["Tipo"] = "Costo faltante"
        sub["Severidad"] = "BAJA"
        sub["Detalle"] = "Costo unitario no disponible; el análisis financiero puede estar subestimado."
        pieces.append(sub)

    if not pieces:
        return pd.DataFrame(columns=["Tipo", "Severidad", "Sede", "Código", "Descripción", "Detalle"])

    result = pd.concat(pieces, ignore_index=True)
    return result[["Tipo", "Severidad", "Sede", "Código", "Descripción", "Detalle"]]


# ============================================================
# CONTACTOS / AUTOMATIZACIÓN
# ============================================================

def load_contacts() -> dict[str, dict[str, str]]:
    if not CONTACTS_PATH.exists():
        return DEFAULT_CONTACTS.copy()
    try:
        raw = json.loads(CONTACTS_PATH.read_text(encoding="utf-8"))
        result = DEFAULT_CONTACTS.copy()
        for site in SEDES:
            result[site] = {
                "supervisor": str(raw.get(site, {}).get("supervisor", "")),
                "whatsapp": str(raw.get(site, {}).get("whatsapp", "")),
                "webhook": str(raw.get(site, {}).get("webhook", "")),
            }
        return result
    except Exception:
        return DEFAULT_CONTACTS.copy()


def save_contacts(contacts: dict[str, dict[str, str]]) -> None:
    CONTACTS_PATH.write_text(
        json.dumps(contacts, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def compose_digest(site: str, df: pd.DataFrame, transfers: pd.DataFrame) -> str:
    if df.empty:
        return f"MAKROPETROL NEXUS · {site}\nSin datos procesados."
    score = health_score(df)
    critical = int(df["Estado"].str.startswith("CRÍTICO").sum())
    purchases = int((df["Compra Ajustada"] > 0).sum()) if "Compra Ajustada" in df.columns else int((df["Compra Sugerida"] > 0).sum())
    withdrawals = int((df["Retiro Almacén"] > 0).sum())
    transfers_out = int(df.get("Transferencias Enviadas", pd.Series(dtype=float)).sum()) if "Transferencias Enviadas" in df.columns else 0
    transfers_in = int(df.get("Transferencias Recibidas", pd.Series(dtype=float)).sum()) if "Transferencias Recibidas" in df.columns else 0
    capital = float(df["Capital Inmovilizado ($)"].sum())
    parts = [
        f"*MAKROPETROL NEXUS* · {site}",
        f"Salud logística: {score}/100 · {health_label(score)}",
        f"🔴 Críticos: {critical}",
        f"🛒 Compras pendientes: {purchases}",
        f"📤 Retiros: {withdrawals}",
        f"♻️ Transferencias recibidas/enviadas: {transfers_in}/{transfers_out}",
        f"💰 Capital inmovilizado: {money(capital)}",
    ]
    if not transfers.empty:
        own = transfers[(transfers["Origen"] == site) | (transfers["Destino"] == site)].head(3)
        if not own.empty:
            parts.append("\n*Movimientos prioritarios:*")
            for r in own.itertuples(index=False):
                parts.append(f"• {r.Código} · {r.Origen} → {r.Destino} · {int(r.Unidades_Sugeridas)} uds")
    return "\n".join(parts)


def post_webhook(url: str, payload: dict[str, Any]) -> tuple[bool, str]:
    if not url.strip():
        return False, "Webhook vacío"
    try:
        r = requests.post(url.strip(), json=payload, timeout=8)
        if 200 <= r.status_code < 300:
            return True, f"HTTP {r.status_code}"
        return False, f"HTTP {r.status_code}: {r.text[:180]}"
    except Exception as exc:
        return False, str(exc)


def post_webhook_async(url: str, payload: dict[str, Any]):
    if not url.strip():
        return None
    return WEBHOOK_EXECUTOR.submit(post_webhook, url, payload)


# ============================================================
# DOCUMENTOS OPERATIVOS
# ============================================================

def _doc_safe_text(value: Any, fallback: str = "") -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return fallback
    return str(value).strip() or fallback


def _doc_order_number(doc_type: str, site: str, snapshot_label: str, index: int = 1) -> str:
    prefix = DOCUMENT_ORDER_PREFIX.get(doc_type, "DOC")
    compact_date = re.sub(r"[^0-9]", "", str(snapshot_label))
    site_code = re.sub(r"[^A-Za-z0-9]", "", site)[:4].upper() or "SEDE"
    return f"{prefix}-{compact_date}-{site_code}-{index:04d}"


def _doc_prepare_frame(
    df: pd.DataFrame,
    document_type: str,
    site: str | None = None,
) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(
            columns=["Código", "Descripción", "Cantidad", "Precio Unitario", "Total"]
        )

    x = df.copy()
    if site and "Sede" in x.columns:
        x = x[x["Sede"].astype(str) == str(site)].copy()

    if document_type == "purchase":
        qty_col = "Compra Ajustada" if "Compra Ajustada" in x.columns else "Compra Sugerida"
        x = x[pd.to_numeric(x.get(qty_col, 0), errors="coerce").fillna(0) > 0].copy()
        if x.empty:
            return pd.DataFrame(columns=["Código", "Descripción", "Cantidad", "Precio Unitario", "Total"])
        result = pd.DataFrame({
            "Código": x["Código"].astype(str),
            "Descripción": x["Descripción"].astype(str),
            "Cantidad": pd.to_numeric(x[qty_col], errors="coerce").fillna(0).round().astype(int),
            "Precio Unitario": pd.to_numeric(x["Costo"], errors="coerce").fillna(0).round(2),
        })

    elif document_type == "withdrawal":
        x = x[pd.to_numeric(x.get("Retiro Almacén", 0), errors="coerce").fillna(0) > 0].copy()
        if x.empty:
            return pd.DataFrame(columns=["Código", "Descripción", "Cantidad", "Precio Unitario", "Total"])
        result = pd.DataFrame({
            "Código": x["Código"].astype(str),
            "Descripción": x["Descripción"].astype(str),
            "Cantidad": pd.to_numeric(x["Retiro Almacén"], errors="coerce").fillna(0).round().astype(int),
            "Precio Unitario": pd.to_numeric(x["Costo"], errors="coerce").fillna(0).round(2),
        })

    elif document_type == "alerts":
        if "Tipo" not in x.columns:
            return pd.DataFrame(columns=["Código", "Descripción", "Alerta", "Severidad", "Detalle"])
        result = x[["Código", "Descripción", "Tipo", "Severidad", "Detalle"]].copy()
        result.columns = ["Código", "Descripción", "Alerta", "Severidad", "Detalle"]
        return result.reset_index(drop=True)

    elif document_type == "redistribution":
        result = x[[
            "Código", "Descripción", "Origen", "Destino",
            "Unidades Sugeridas", "Costo Unitario ($)",
            "Compra Evitada Estimada ($)", "Estado"
        ]].copy()
        return result.reset_index(drop=True)

    else:
        return pd.DataFrame()

    result["Total"] = (result["Cantidad"] * result["Precio Unitario"]).round(2)
    return result.reset_index(drop=True)


def _document_purchase_meta(
    site: str,
    snapshot_label: str,
    supplier: str = "",
    supplier_address: str = "",
    supplier_phone: str = "",
    sequence: int = 1,
) -> dict[str, str]:
    now = datetime.now()
    return {
        "company": COMPANY_INFO["name"],
        "company_address": COMPANY_INFO["address"],
        "company_phone": COMPANY_INFO["phone"],
        "company_rif": COMPANY_INFO["rif"],
        "document_title": "ORDEN DE COMPRA",
        "date": str(snapshot_label),
        "time": now.strftime("%H:%M:%S"),
        "site": site,
        "order_no": _doc_order_number("purchase", site, snapshot_label, sequence),
        "supplier": supplier or SUPPLIER_DEFAULTS["name"],
        "supplier_address": supplier_address or SUPPLIER_DEFAULTS["address"],
        "supplier_phone": supplier_phone or SUPPLIER_DEFAULTS["phone"],
        "operator": "NEXUS / ADMINISTRACION",
    }


def build_purchase_pdf(
    data: pd.DataFrame,
    site: str,
    snapshot_label: str,
    sequence: int = 1,
    supplier: str = "",
    supplier_address: str = "",
    supplier_phone: str = "",
) -> bytes | None:
    if not HAS_PDF:
        return None

    x = _doc_prepare_frame(data, "purchase", site)
    meta = _document_purchase_meta(
        site, snapshot_label, supplier, supplier_address, supplier_phone, sequence
    )
    total = float(x["Total"].sum()) if not x.empty else 0.0

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()
    pdf.set_title(f"{meta['document_title']} {meta['order_no']}")

    def header(page_no: int):
        pdf.set_fill_color(24, 36, 58)
        pdf.rect(0, 0, 210, 14, "F")
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_xy(8, 3)
        pdf.cell(130, 6, _doc_safe_text(meta["company"]))
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(0, 6, _doc_safe_text(meta["document_title"]), align="R")
        pdf.set_text_color(30, 30, 30)

        pdf.set_xy(8, 19)
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 6, _doc_safe_text(meta["company"]))
        pdf.set_font("Helvetica", "", 7)
        pdf.set_xy(8, 25)
        pdf.multi_cell(115, 4, _doc_safe_text(meta["company_address"]))
        pdf.set_xy(132, 19)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(68, 5, "ORDEN DE COMPRA", align="R")
        pdf.set_font("Helvetica", "", 7)
        pdf.set_xy(132, 25)
        pdf.cell(68, 4, f"Fecha: {meta['date']}", align="R")
        pdf.set_xy(132, 30)
        pdf.cell(68, 4, f"Hora: {meta['time']}", align="R")
        pdf.set_xy(132, 35)
        pdf.cell(68, 4, f"Pag: {page_no}", align="R")

        y = 43
        pdf.set_fill_color(238, 242, 249)
        pdf.rect(8, y, 194, 29, "F")
        pdf.set_draw_color(214, 222, 234)
        pdf.rect(8, y, 194, 29)
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_xy(11, y + 3)
        pdf.cell(24, 4, "Telefono:")
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(40, 4, meta["company_phone"])
        pdf.set_font("Helvetica", "B", 7)
        pdf.cell(16, 4, "Rif:")
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(35, 4, meta["company_rif"])
        pdf.set_font("Helvetica", "B", 7)
        pdf.cell(27, 4, "Orden:")
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(42, 4, meta["order_no"])
        pdf.ln(6)
        pdf.set_x(11)
        pdf.set_font("Helvetica", "B", 7)
        pdf.cell(24, 4, "Proveedor:")
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(117, 4, _doc_safe_text(meta["supplier"]))
        pdf.set_font("Helvetica", "B", 7)
        pdf.cell(18, 4, "Sede:")
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(30, 4, _doc_safe_text(site), align="R")
        pdf.ln(5)
        pdf.set_x(11)
        pdf.set_font("Helvetica", "B", 7)
        pdf.cell(24, 4, "Direccion:")
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(117, 4, _doc_safe_text(meta["supplier_address"]))
        pdf.set_font("Helvetica", "B", 7)
        pdf.cell(18, 4, "Telefono:")
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(30, 4, _doc_safe_text(meta["supplier_phone"]), align="R")
        pdf.set_xy(8, y + 34)

        widths = [25, 91, 18, 27, 33]
        headers = ["Código", "Descripción", "Cantidad", "Precio Unitario", "Total"]
        pdf.set_fill_color(53, 104, 245)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 7)
        for w, h in zip(widths, headers):
            pdf.cell(w, 7, h, border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_text_color(30, 30, 30)

    header(1)

    widths = [25, 91, 18, 27, 33]
    if x.empty:
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(194, 8, "Sin productos pendientes de compra para esta sede.", border=1)
        pdf.ln(8)
    else:
        pdf.set_font("Helvetica", "", 7)
        for row_idx, row in enumerate(x.to_dict("records"), start=1):
            if pdf.get_y() > 270:
                pdf.add_page()
                header(pdf.page_no())
            desc = _doc_safe_text(row["Descripción"])
            pdf.cell(25, 6, _doc_safe_text(row["Código"]), border="LR")
            pdf.cell(91, 6, desc[:70], border="LR")
            pdf.cell(18, 6, f"{int(row['Cantidad']):,}", border="LR", align="R")
            pdf.cell(27, 6, f"{float(row['Precio Unitario']):,.2f}", border="LR", align="R")
            pdf.cell(33, 6, f"{float(row['Total']):,.2f}", border="LR", align="R")
            pdf.ln()
        pdf.cell(sum(widths), 0, "", border="T")

    pdf.ln(7)
    pdf.set_font("Helvetica", "", 8)
    summary = [
        ("Total:", total),
        ("Flete:", 0.0),
        ("Descuento:", 0.0),
        ("Otro descuento:", 0.0),
        ("Total Neto:", total),
    ]
    x0 = 126
    for label, value in summary:
        pdf.set_x(x0)
        if label == "Total Neto:":
            pdf.set_font("Helvetica", "B", 9)
        pdf.cell(42, 6, label, align="R")
        pdf.cell(36, 6, f"{value:,.2f}", align="R")
        pdf.ln()
        pdf.set_font("Helvetica", "", 8)

    pdf.set_y(min(pdf.get_y() + 5, 278))
    pdf.set_draw_color(210, 218, 230)
    pdf.line(8, pdf.get_y(), 202, pdf.get_y())
    pdf.ln(3)
    pdf.set_font("Helvetica", "B", 7)
    pdf.cell(45, 5, "Operador:")
    pdf.set_font("Helvetica", "", 7)
    pdf.cell(55, 5, meta["operator"])
    pdf.set_font("Helvetica", "B", 7)
    pdf.cell(42, 5, "Total Items:")
    pdf.set_font("Helvetica", "", 7)
    pdf.cell(30, 5, f"{len(x):,}", align="R")
    if len(x) > 0 and pdf.page_no() > 1:
        pdf.ln(8)
        pdf.set_font("Helvetica", "I", 7)
        pdf.cell(0, 5, "Continua...", align="R")

    pdf.set_y(287)
    pdf.set_font("Helvetica", "", 6)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 4, f"MAKROPETROL NEXUS v{APP_VERSION} · Documento generado automaticamente", align="C")

    return bytes(pdf.output())


def build_withdrawal_pdf(
    data: pd.DataFrame,
    site: str,
    snapshot_label: str,
    sequence: int = 1,
) -> bytes | None:
    if not HAS_PDF:
        return None

    x = _doc_prepare_frame(data, "withdrawal", site)
    total = float(x["Total"].sum()) if not x.empty else 0.0
    meta = _document_purchase_meta(site, snapshot_label, sequence=sequence)
    meta["document_title"] = "RETIRO DE ALMACÉN"
    meta["order_no"] = _doc_order_number("withdrawal", site, snapshot_label, sequence)

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()

    def hdr(page_no: int):
        pdf.set_fill_color(24, 36, 58)
        pdf.rect(0, 0, 210, 14, "F")
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_xy(8, 3)
        pdf.cell(135, 6, _doc_safe_text(COMPANY_INFO["name"]))
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(0, 6, "RETIRO DE ALMACÉN", align="R")
        pdf.set_text_color(30, 30, 30)
        pdf.set_xy(8, 20)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(125, 6, "RETIRO DE ALMACÉN")
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(0, 6, f"Pag: {page_no}", align="R")
        pdf.set_xy(8, 28)
        pdf.cell(194, 5, f"Sede: {_doc_safe_text(site)}")
        pdf.set_xy(8, 34)
        pdf.cell(194, 5, f"Fecha: {snapshot_label}    Hora: {meta['time']}    Documento: {meta['order_no']}")
        pdf.set_xy(8, 44)
        widths=[25,91,18,27,33]
        heads=["Código","Descripción","Cantidad","Costo Unitario","Valor Retiro"]
        pdf.set_fill_color(53,104,245)
        pdf.set_text_color(255,255,255)
        pdf.set_font("Helvetica","B",7)
        for w,h in zip(widths,heads):
            pdf.cell(w,7,h,border=1,fill=True,align="C")
        pdf.ln()
        pdf.set_text_color(30,30,30)

    hdr(1)
    if x.empty:
        pdf.set_font("Helvetica","",8)
        pdf.cell(194,8,"Sin productos pendientes de retiro para esta sede.",border=1)
    else:
        for row in x.to_dict("records"):
            if pdf.get_y() > 270:
                pdf.add_page()
                hdr(pdf.page_no())
            pdf.set_font("Helvetica","",7)
            pdf.cell(25,6,_doc_safe_text(row["Código"]),border="LR")
            pdf.cell(91,6,_doc_safe_text(row["Descripción"])[:70],border="LR")
            pdf.cell(18,6,f"{int(row['Cantidad']):,}",border="LR",align="R")
            pdf.cell(27,6,f"{float(row['Precio Unitario']):,.2f}",border="LR",align="R")
            pdf.cell(33,6,f"{float(row['Total']):,.2f}",border="LR",align="R")
            pdf.ln()
    pdf.ln(6)
    pdf.set_font("Helvetica","B",8)
    pdf.cell(145,6,"Valor total del retiro:",align="R")
    pdf.cell(45,6,f"{total:,.2f}",align="R")
    pdf.ln(8)
    pdf.set_font("Helvetica","",7)
    pdf.cell(45,5,"Motivo:")
    pdf.cell(149,5,"SOBRESTOCK / REDISTRIBUCIÓN / AJUSTE OPERATIVO")
    pdf.ln(8)
    pdf.cell(45,5,"Responsable de salida:")
    pdf.cell(70,5,"")
    pdf.cell(38,5,"Total Items:")
    pdf.cell(41,5,f"{len(x):,}",align="R")
    pdf.set_y(287)
    pdf.set_font("Helvetica","",6)
    pdf.set_text_color(120,120,120)
    pdf.cell(0,4,f"MAKROPETROL NEXUS v{APP_VERSION} · Documento generado automaticamente",align="C")
    return bytes(pdf.output())


def build_redistribution_pdf(
    transfers: pd.DataFrame,
    snapshot_label: str,
    sequence: int = 1,
) -> bytes | None:
    if not HAS_PDF:
        return None

    x = _doc_prepare_frame(transfers, "redistribution")
    units = int(pd.to_numeric(x.get("Unidades Sugeridas", 0), errors="coerce").fillna(0).sum()) if not x.empty else 0
    avoided = float(pd.to_numeric(x.get("Compra Evitada Estimada ($)", 0), errors="coerce").fillna(0).sum()) if not x.empty else 0.0
    doc_no = _doc_order_number("redistribution", "GLOBAL", snapshot_label, sequence)

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()

    def hdr(page_no: int):
        pdf.set_fill_color(24,36,58)
        pdf.rect(0,0,210,15,"F")
        pdf.set_text_color(255,255,255)
        pdf.set_font("Helvetica","B",11)
        pdf.set_xy(8,4)
        pdf.cell(135,6,_doc_safe_text(COMPANY_INFO["name"]))
        pdf.set_font("Helvetica","",7)
        pdf.cell(0,6,"REDISTRIBUCIÓN MULTISEDE",align="R")
        pdf.set_text_color(30,30,30)
        pdf.set_xy(8,21)
        pdf.set_font("Helvetica","B",12)
        pdf.cell(125,6,"PLAN DE REDISTRIBUCIÓN")
        pdf.set_font("Helvetica","",7)
        pdf.cell(0,6,f"Pag: {page_no}",align="R")
        pdf.set_xy(8,30)
        pdf.cell(194,5,f"Fecha: {snapshot_label}    Hora: {datetime.now().strftime('%H:%M:%S')}    Documento: {doc_no}")
        pdf.set_xy(8,40)
        widths=[23,68,31,31,18,23]
        heads=["Código","Descripción","Origen","Destino","Cantidad","Ahorro Est."]
        pdf.set_fill_color(23,170,190)
        pdf.set_text_color(255,255,255)
        pdf.set_font("Helvetica","B",7)
        for w,h in zip(widths,heads):
            pdf.cell(w,7,h,border=1,fill=True,align="C")
        pdf.ln()
        pdf.set_text_color(30,30,30)

    hdr(1)
    if x.empty:
        pdf.set_font("Helvetica","",8)
        pdf.cell(194,8,"No existen transferencias sugeridas.",border=1)
    else:
        for row in x.to_dict("records"):
            if pdf.get_y() > 270:
                pdf.add_page()
                hdr(pdf.page_no())
            pdf.set_font("Helvetica","",7)
            pdf.cell(23,6,_doc_safe_text(row["Código"]),border="LR")
            pdf.cell(68,6,_doc_safe_text(row["Descripción"])[:50],border="LR")
            pdf.cell(31,6,_doc_safe_text(row["Origen"])[:20],border="LR")
            pdf.cell(31,6,_doc_safe_text(row["Destino"])[:20],border="LR")
            pdf.cell(18,6,f"{int(row['Unidades Sugeridas']):,}",border="LR",align="R")
            pdf.cell(23,6,f"{float(row['Compra Evitada Estimada ($)']):,.0f}",border="LR",align="R")
            pdf.ln()

    pdf.ln(7)
    pdf.set_font("Helvetica","B",8)
    pdf.cell(90,6,"Unidades a redistribuir:",align="R")
    pdf.cell(50,6,f"{units:,}",align="R")
    pdf.ln(6)
    pdf.cell(90,6,"Compra potencial evitada:",align="R")
    pdf.cell(50,6,f"${avoided:,.2f}",align="R")
    pdf.ln(10)
    pdf.set_font("Helvetica","I",7)
    pdf.multi_cell(194,5,"Criterio: mover excedentes existentes hacia sedes con déficit antes de generar compras nuevas.")
    pdf.set_y(287)
    pdf.set_font("Helvetica","",6)
    pdf.set_text_color(120,120,120)
    pdf.cell(0,4,f"MAKROPETROL NEXUS v{APP_VERSION} · Documento multisede generado automaticamente",align="C")
    return bytes(pdf.output())


def build_alerts_pdf(
    alerts: pd.DataFrame,
    site: str,
    snapshot_label: str,
    sequence: int = 1,
) -> bytes | None:
    if not HAS_PDF:
        return None
    x = _doc_prepare_frame(alerts, "alerts", site)
    doc_no = _doc_order_number("alerts", site, snapshot_label, sequence)
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()

    def hdr(page_no: int):
        pdf.set_fill_color(24,36,58)
        pdf.rect(0,0,210,15,"F")
        pdf.set_text_color(255,255,255)
        pdf.set_font("Helvetica","B",11)
        pdf.set_xy(8,4)
        pdf.cell(135,6,_doc_safe_text(COMPANY_INFO["name"]))
        pdf.set_font("Helvetica","",7)
        pdf.cell(0,6,"REPORTE DE ALERTAS",align="R")
        pdf.set_text_color(30,30,30)
        pdf.set_xy(8,21)
        pdf.set_font("Helvetica","B",12)
        pdf.cell(125,6,"REPORTE DE ALERTAS OPERATIVAS")
        pdf.set_font("Helvetica","",7)
        pdf.cell(0,6,f"Pag: {page_no}",align="R")
        pdf.set_xy(8,30)
        pdf.cell(194,5,f"Sede: {site}    Fecha: {snapshot_label}    Documento: {doc_no}")
        pdf.set_xy(8,40)
        widths=[24,48,33,22,67]
        heads=["Código","Descripción","Alerta","Severidad","Detalle"]
        pdf.set_fill_color(223,85,85)
        pdf.set_text_color(255,255,255)
        pdf.set_font("Helvetica","B",7)
        for w,h in zip(widths,heads):
            pdf.cell(w,7,h,border=1,fill=True,align="C")
        pdf.ln()
        pdf.set_text_color(30,30,30)
    hdr(1)
    if x.empty:
        pdf.set_font("Helvetica","",8)
        pdf.cell(194,8,"Sin alertas para esta sede.",border=1)
    else:
        for row in x.to_dict("records"):
            if pdf.get_y() > 268:
                pdf.add_page()
                hdr(pdf.page_no())
            pdf.set_font("Helvetica","",6.5)
            pdf.cell(24,7,_doc_safe_text(row["Código"]),border=1)
            pdf.cell(48,7,_doc_safe_text(row["Descripción"])[:34],border=1)
            pdf.cell(33,7,_doc_safe_text(row["Alerta"])[:25],border=1)
            pdf.cell(22,7,_doc_safe_text(row["Severidad"]),border=1,align="C")
            pdf.cell(67,7,_doc_safe_text(row["Detalle"])[:54],border=1)
            pdf.ln()
    pdf.set_y(287)
    pdf.set_font("Helvetica","",6)
    pdf.set_text_color(120,120,120)
    pdf.cell(0,4,f"MAKROPETROL NEXUS v{APP_VERSION} · Documento generado automaticamente",align="C")
    return bytes(pdf.output())


def build_operational_html(
    document_type: str,
    data: pd.DataFrame,
    site: str | None,
    snapshot_label: str,
    title: str,
) -> bytes:
    if document_type == "redistribution":
        x = _doc_prepare_frame(data, document_type)
    else:
        x = _doc_prepare_frame(data, document_type, site)

    css = """
    :root{--navy:#18243a;--blue:#3568f5;--cyan:#17aabe;--line:#dfe6ef;--bg:#f4f7fb;--muted:#748198}
    *{box-sizing:border-box} body{margin:0;background:var(--bg);color:#172137;font-family:Arial,Segoe UI,sans-serif}
    .page{max-width:1100px;margin:24px auto;padding:28px;background:#fff;border:1px solid #e1e7ef;box-shadow:0 18px 50px rgba(25,45,75,.08);border-radius:20px}
    .top{background:linear-gradient(135deg,#18243a,#263b65);color:#fff;border-radius:15px;padding:18px 20px;display:flex;justify-content:space-between;gap:20px}
    .brand{font-size:21px;font-weight:800}.doc{font-weight:800;text-align:right}.muted{color:var(--muted);font-size:12px}.meta{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin:16px 0}
    .box{border:1px solid var(--line);border-radius:12px;padding:12px;background:#fbfcfe}.label{font-size:10px;font-weight:800;text-transform:uppercase;color:var(--muted)}.value{font-size:13px;font-weight:700;margin-top:3px}
    .kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin:16px 0}.kpi{border:1px solid var(--line);border-radius:12px;padding:14px}.kpi strong{display:block;font-size:22px;margin-top:4px}
    table{width:100%;border-collapse:collapse;font-size:12px;margin-top:16px} th{background:var(--blue);color:#fff;text-align:left;padding:9px} td{padding:8px;border-bottom:1px solid var(--line)} tr:nth-child(even){background:#fafcff}
    .foot{margin-top:20px;padding-top:12px;border-top:1px solid var(--line);font-size:11px;color:var(--muted);display:flex;justify-content:space-between}
    .pill{display:inline-block;padding:5px 8px;border-radius:999px;background:#edf2ff;color:#355bd7;font-weight:800;font-size:10px}
    @media print{body{background:#fff}.page{box-shadow:none;border:0;margin:0;max-width:none;border-radius:0}.no-print{display:none!important}}
    @media(max-width:700px){.meta,.kpis{grid-template-columns:1fr 1fr}.top{flex-direction:column}.doc{text-align:left}}
    """
    headers = list(x.columns)
    rows = []
    for rec in x.to_dict("records"):
        row = "<tr>" + "".join(f"<td>{_doc_safe_text(rec.get(c,''))}</td>" for c in headers) + "</tr>"
        rows.append(row)

    if document_type == "purchase":
        qty = int(pd.to_numeric(x.get("Cantidad", 0), errors="coerce").fillna(0).sum()) if not x.empty else 0
        total = float(pd.to_numeric(x.get("Total", 0), errors="coerce").fillna(0).sum()) if not x.empty else 0.0
        kpi_html = f"""
        <div class="kpi"><div class="label">Items</div><strong>{len(x):,}</strong></div>
        <div class="kpi"><div class="label">Unidades</div><strong>{qty:,}</strong></div>
        <div class="kpi"><div class="label">Total</div><strong>${total:,.2f}</strong></div>
        <div class="kpi"><div class="label">Sede</div><strong>{_doc_safe_text(site)}</strong></div>
        """
    elif document_type == "withdrawal":
        qty = int(pd.to_numeric(x.get("Cantidad", 0), errors="coerce").fillna(0).sum()) if not x.empty else 0
        total = float(pd.to_numeric(x.get("Total", 0), errors="coerce").fillna(0).sum()) if not x.empty else 0.0
        kpi_html = f"""
        <div class="kpi"><div class="label">Items a retirar</div><strong>{len(x):,}</strong></div>
        <div class="kpi"><div class="label">Unidades</div><strong>{qty:,}</strong></div>
        <div class="kpi"><div class="label">Capital liberable</div><strong>${total:,.2f}</strong></div>
        <div class="kpi"><div class="label">Sede</div><strong>{_doc_safe_text(site)}</strong></div>
        """
    elif document_type == "redistribution":
        units = int(pd.to_numeric(x.get("Unidades Sugeridas", 0), errors="coerce").fillna(0).sum()) if not x.empty else 0
        avoided = float(pd.to_numeric(x.get("Compra Evitada Estimada ($)", 0), errors="coerce").fillna(0).sum()) if not x.empty else 0.0
        routes = int(len(x)) if not x.empty else 0
        kpi_html = f"""
        <div class="kpi"><div class="label">Movimientos</div><strong>{routes:,}</strong></div>
        <div class="kpi"><div class="label">Unidades</div><strong>{units:,}</strong></div>
        <div class="kpi"><div class="label">Compra evitada</div><strong>${avoided:,.2f}</strong></div>
        <div class="kpi"><div class="label">Ámbito</div><strong>Multisede</strong></div>
        """
    else:
        critical = int((x["Severidad"].astype(str) == "CRÍTICA").sum()) if "Severidad" in x.columns else 0
        kpi_html = f"""
        <div class="kpi"><div class="label">Alertas</div><strong>{len(x):,}</strong></div>
        <div class="kpi"><div class="label">Críticas</div><strong>{critical:,}</strong></div>
        <div class="kpi"><div class="label">Sede</div><strong>{_doc_safe_text(site)}</strong></div>
        <div class="kpi"><div class="label">Corte</div><strong>{snapshot_label}</strong></div>
        """

    meta_left = f"""
    <div class="box"><div class="label">Empresa</div><div class="value">{_doc_safe_text(COMPANY_INFO["name"])}</div>
    <div class="muted">{_doc_safe_text(COMPANY_INFO["address"])}</div></div>
    <div class="box"><div class="label">Documento</div><div class="value">{_doc_safe_text(title)}</div>
    <div class="muted">Fecha: {snapshot_label} · Hora: {datetime.now().strftime('%H:%M:%S')}</div></div>
    """
    table = "<table><thead><tr>" + "".join(f"<th>{c}</th>" for c in headers) + "</tr></thead><tbody>" + "".join(rows) + "</tbody></table>"
    html = f"""<!doctype html><html lang="es"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
    <title>{title}</title><style>{css}</style></head><body><main class="page">
    <div class="top"><div><div class="brand">{_doc_safe_text(COMPANY_INFO["name"])}</div><div style="font-size:11px;margin-top:4px">{_doc_safe_text(COMPANY_INFO["address"])}</div></div>
    <div class="doc">{_doc_safe_text(title)}<div style="font-size:11px;margin-top:4px">Corte {snapshot_label}</div></div></div>
    <div class="meta">{meta_left}</div><div class="kpis">{kpi_html}</div>{table}
    <div class="foot"><span>Makropetrol NEXUS v{APP_VERSION}</span><span>Generado automáticamente · listo para imprimir / PDF</span></div>
    </main></body></html>"""
    return html.encode("utf-8")


@st.cache_data(show_spinner=False)
def build_purchase_excel(data: pd.DataFrame, site: str, snapshot_label: str, sequence: int = 1) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    x = _doc_prepare_frame(data, "purchase", site)
    qty = int(x["Cantidad"].sum()) if not x.empty else 0
    total = float(x["Total"].sum()) if not x.empty else 0.0
    out = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Orden de Compra"
    navy, blue, line, white = "18243A", "3568F5", "DDE5EF", "FFFFFF"
    ws.merge_cells("A1:E1")
    ws["A1"] = COMPANY_INFO["name"]
    ws["A1"].font = Font(size=16, bold=True, color=white)
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:E2")
    ws["A2"] = COMPANY_INFO["address"]
    ws["A2"].font = Font(size=9, color="667287")
    ws["A3"] = "ORDEN DE COMPRA"
    ws["A3"].font = Font(size=14, bold=True, color=navy)
    ws["D3"] = f"Fecha: {snapshot_label}"
    ws["E3"] = f"Pag: 1"
    ws["A4"] = f"Telefono: {COMPANY_INFO['phone']}"
    ws["B4"] = f"Rif: {COMPANY_INFO['rif']}"
    ws["C4"] = f"Sede: {site}"
    ws["D4"] = "Operador:"
    ws["E4"] = "NEXUS / ADMINISTRACION"

    headers = ["Código","Descripción","Cantidad","Precio Unitario","Total"]
    for j,h in enumerate(headers,1):
        c=ws.cell(6,j,h)
        c.fill=PatternFill("solid",fgColor=blue)
        c.font=Font(bold=True,color=white)
        c.alignment=Alignment(horizontal="center")
    for i, row in enumerate(x.to_dict("records"), 7):
        vals = [
            row["Código"],
            row["Descripción"],
            int(row["Cantidad"]),
            float(row["Precio Unitario"]),
            float(row["Total"]),
        ]
        for j, val in enumerate(vals, 1):
            ws.cell(i, j, val)
    total_row=7+len(x)
    ws.cell(total_row+1,4,"Total Neto:")
    ws.cell(total_row+1,4).font=Font(bold=True)
    ws.cell(total_row+1,5,total)
    ws.cell(total_row+1,5).font=Font(bold=True)
    ws.cell(total_row+2,1,"Total Items:")
    ws.cell(total_row+2,2,len(x))
    ws.cell(total_row+2,3,"Unidades:")
    ws.cell(total_row+2,4,qty)
    for row in ws.iter_rows():
        for cell in row:
            cell.border=Border(bottom=Side(style="hair",color=line))
            cell.alignment=Alignment(vertical="center",wrap_text=True)
    ws.freeze_panes="A7"
    widths=[18,62,12,18,18]
    for idx,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(idx)].width=w
    for r in range(7,total_row+1):
        ws.cell(r,4).number_format='$#,##0.00'
        ws.cell(r,5).number_format='$#,##0.00'
    ws.cell(total_row+1,5).number_format='$#,##0.00'
    wb.save(out)
    return out.getvalue()


# ============================================================
# EXCEL PREMIUM
# ============================================================

@st.cache_data(show_spinner=False)
def export_excel(
    site_results: dict[str, pd.DataFrame],
    transfers: pd.DataFrame,
    months_history: int,
    min_coverage: float,
    max_coverage: float,
    lead_time_days: int,
    snapshot_label: str,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    all_df = combine_sites(site_results)
    out = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "00 Dashboard"

    navy = "18243A"
    blue = "3568F5"
    gray = "7B8799"
    line = "DDE5EF"
    white = "FFFFFF"

    thin = Side(style="thin", color=line)

    def title(ws_, title, subtitle, end_col=10):
        ws_.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        ws_.cell(1, 1).value = title
        ws_.cell(1, 1).font = Font(size=20, bold=True, color=white)
        ws_.cell(1, 1).fill = PatternFill("solid", fgColor=navy)
        ws_.cell(1, 1).alignment = Alignment(vertical="center")
        ws_.row_dimensions[1].height = 34
        ws_.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        ws_.cell(2, 1).value = subtitle
        ws_.cell(2, 1).font = Font(size=10, color=gray, italic=True)
        ws_.row_dimensions[2].height = 23

    def write_table(ws_, df_: pd.DataFrame, start_row=4, start_col=1, table_name="Table1"):
        if df_.empty:
            ws_.cell(start_row, start_col).value = "Sin registros"
            return
        for j, col in enumerate(df_.columns, start_col):
            c = ws_.cell(start_row, j, col)
            c.fill = PatternFill("solid", fgColor=blue)
            c.font = Font(bold=True, color=white)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = Border(bottom=thin)
        for i, row in enumerate(df_.itertuples(index=False), start_row + 1):
            for j, val in enumerate(row, start_col):
                cell = ws_.cell(i, j, val)
                cell.border = Border(bottom=Side(style="hair", color=line))
                cell.alignment = Alignment(vertical="center")
        end_row = start_row + len(df_)
        end_col = start_col + len(df_.columns) - 1
        ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws_.add_table(table)
        ws_.freeze_panes = f"{get_column_letter(start_col)}{start_row+1}"
        ws_.auto_filter.ref = ref
        for j in range(start_col, end_col + 1):
            max_len = max(len(str(ws_.cell(r, j).value or "")) for r in range(start_row, min(end_row, start_row + 100) + 1))
            ws_.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 11), 42)

    score = health_score(all_df)
    kpis = [
        ("Valor inventario", float(all_df["Valor Inventario ($)"].sum()) if not all_df.empty else 0),
        ("Compras", int((all_df.get("Compra Ajustada", all_df.get("Compra Sugerida", pd.Series(dtype=float))) > 0).sum()) if not all_df.empty else 0),
        ("Retiros", int((all_df.get("Retiro Almacén", pd.Series(dtype=float)) > 0).sum()) if not all_df.empty else 0),
        ("Capital inmovilizado", float(all_df["Capital Inmovilizado ($)"].sum()) if not all_df.empty else 0),
    ]
    title(ws, "MAKROPETROL NEXUS · REPORTE EJECUTIVO", f"Corte: {snapshot_label} · Historial: {months_history} meses · Lead time: {lead_time_days} días", 10)
    ws["A4"] = "SALUD LOGÍSTICA"
    ws["A4"].font = Font(bold=True, color=navy, size=12)
    ws["B4"] = f"{score}/100 · {health_label(score)}"
    ws["B4"].font = Font(bold=True, color=blue, size=16)
    for idx, (label, value) in enumerate(kpis, 1):
        col = 1 + (idx - 1) * 2
        ws.cell(6, col, label).font = Font(bold=True, color=gray, size=9)
        ws.cell(7, col, value)
        ws.cell(7, col).font = Font(bold=True, color=navy, size=15)
        if "Valor" in label or "Capital" in label:
            ws.cell(7, col).number_format = '$#,##0.00'
        else:
            ws.cell(7, col).number_format = '#,##0'

    if not all_df.empty:
        by_site = all_df.groupby("Sede").agg(
            SKUs=("Código", "count"),
            Compras=("Compra Sugerida", lambda x: int((x > 0).sum())),
            Retiros=("Retiro Almacén", lambda x: int((x > 0).sum())),
            Inventario=("Valor Inventario ($)", "sum"),
            Inmovilizado=("Capital Inmovilizado ($)", "sum"),
        ).reset_index()
        write_table(ws, by_site, start_row=10, table_name="ResumenSedes")

        chart = BarChart()
        chart.title = "Valor de inventario por sede"
        data_ref = Reference(ws, min_col=6, min_row=10, max_row=10 + len(by_site))
        cat_ref = Reference(ws, min_col=1, min_row=11, max_row=10 + len(by_site))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.height = 7
        chart.width = 12
        ws.add_chart(chart, "H10")

    sheets = [
        ("01 Inventario", all_df, "InventarioGeneral"),
        ("02 Compras", all_df[all_df.get("Compra Ajustada", all_df.get("Compra Sugerida", pd.Series(dtype=float))) > 0].copy() if not all_df.empty else all_df, "PlanCompras"),
        ("03 Retiros", all_df[all_df.get("Retiro Almacén", pd.Series(dtype=float)) > 0].copy() if not all_df.empty else all_df, "PlanRetiros"),
        ("04 Redistribución", transfers.copy(), "PlanTransferencias"),
        ("05 Alertas", build_alerts(all_df), "Alertas"),
    ]
    for sheet_name, data, table_name in sheets:
        sh = wb.create_sheet(sheet_name)
        title(sh, sheet_name.replace("01 ", "").replace("02 ", "").replace("03 ", "").replace("04 ", "").replace("05 ", ""), f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", min(max(len(data.columns), 8), 16) if not data.empty else 8)
        write_table(sh, data, start_row=4, table_name=table_name)

    crit = wb.create_sheet("06 Criterios")
    title(crit, "CRITERIOS DEL MOTOR", "Parámetros y definiciones usados en el corte", 5)
    criteria = pd.DataFrame(
        [
            ["Meses de historial", months_history, "Divide el total de ventas del período para obtener demanda mensual."],
            ["Cobertura mínima", min_coverage, "Meses de cobertura usados como umbral operativo."],
            ["Cobertura máxima", max_coverage, "Meses de cobertura objetivo antes de sugerir retiro."],
            ["Lead time", lead_time_days, "Días estimados de entrega del proveedor."],
            ["Stock seguridad", "Demanda diaria × días de seguridad", "Protección adicional ante variabilidad/entrega."],
            ["Punto de reorden", "máximo(stock mínimo, demanda lead time + stock seguridad)", "Umbral para iniciar abastecimiento."],
            ["ABC", "80% / 95% / resto", "Clasificación A/B/C según base seleccionada."],
            ["Redistribución", "Exceso de una sede → déficit de otra", "Se prioriza mover inventario existente antes de comprar."],
        ],
        columns=["Parámetro", "Valor", "Interpretación"],
    )
    write_table(crit, criteria, start_row=4, table_name="CriteriosMotor")

    for sh in wb.worksheets:
        sh.sheet_view.showGridLines = False
        for row in sh.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=False)
        headers = {cell.value: cell.column for cell in sh[4]}
        integer_headers = {"Ventas", "Demanda Mensual", "Existencia", "Stock Mínimo", "Stock Máximo", "Demanda Lead Time", "Stock Seguridad", "Punto de Reorden", "Compra Sugerida", "Compra Ajustada", "Retiro Almacén", "Unidades Sugeridas", "Transferencias Recibidas", "Transferencias Enviadas"}
        money_headers = {"Costo", "Valor Inventario ($)", "Capital Inmovilizado ($)", "Costo Compra Estimada ($)", "Valor Movimiento ($)", "Costo Unitario ($)", "Compra Evitada Estimada ($)"}
        for h, col in headers.items():
            if h in integer_headers:
                for c in sh.iter_cols(min_col=col, max_col=col, min_row=5, max_row=sh.max_row):
                    for cell in c:
                        cell.number_format = '#,##0'
            if h in money_headers:
                for c in sh.iter_cols(min_col=col, max_col=col, min_row=5, max_row=sh.max_row):
                    for cell in c:
                        cell.number_format = '$#,##0.00'

    wb.save(out)
    return out.getvalue()


# ============================================================
# INICIALIZACIÓN Y SIDEBAR
# ============================================================

init_db()
contacts = load_contacts()

NAV_GROUPS: dict[str, list[str]] = {
    "▣ INICIO": ["Centro de Control"],
    "▣ OPERACIÓN": ["Inventario", "Compras", "Proveedores", "Sin Rotación", "Redistribución"],
    "▣ INTELIGENCIA": ["Centro Visual", "Finanzas", "Alertas"],
    "▣ DATOS": ["Cortes Diarios", "Historial"],
    "▣ AUTOMATIZACIÓN": ["Automatización"],
    "▣ REPORTES": ["Exportación de Datos"],
}

with st.sidebar:
    st.markdown(
        """
        <div style="padding:8px 6px 14px">
            <div style="font-family:Manrope;font-size:1.2rem;font-weight:800">◈ MAKROPETROL <span style="color:#3568f5">NEXUS</span></div>
            <div style="font-size:.72rem;color:#7b8799;margin-top:3px">Intelligence Platform</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    modo_gerencia = st.toggle("👔 Modo Gerencia", value=st.session_state.get("modo_gerencia", False), key="modo_gerencia",
                               help="Oculta columnas técnicas y parámetros; muestra KPIs, tendencias, alertas y recomendaciones.")
    st.caption("🔧 Modo Operativo" if not modo_gerencia else "Vista simplificada para dirección")
    st.divider()

    flat_views = [v for group in NAV_GROUPS.values() for v in group]
    if "nexus_view" not in st.session_state:
        st.session_state["nexus_view"] = flat_views[0]

    for group_label, items in NAV_GROUPS.items():
        st.markdown(f'<div class="nav-group">{group_label}</div>', unsafe_allow_html=True)
        for item in items:
            active = st.session_state["nexus_view"] == item
            if st.button(item, key=f"nav_{item}", use_container_width=True, type="primary" if active else "secondary"):
                st.session_state["nexus_view"] = item
                st.rerun()
    view = st.session_state["nexus_view"]

    st.divider()
    with st.popover("⚙️ Configuración del motor", use_container_width=True):
        st.markdown("**Historial**")
        months_history = st.number_input("Meses representados por ventas", min_value=1, max_value=36, value=3, step=1)
        rolling_days = st.number_input("Ventana diaria (días)", min_value=1, max_value=365, value=30, step=1)
        st.markdown("**Cobertura**")
        min_coverage = st.number_input("Cobertura mínima (meses)", min_value=0.1, max_value=12.0, value=0.5, step=0.5)
        max_coverage = st.number_input("Cobertura máxima (meses)", min_value=0.5, max_value=24.0, value=1.0, step=0.5)
        st.markdown("**Proveedor**")
        lead_time_days = st.number_input("Lead time proveedor (días)", min_value=0, max_value=180, value=7, step=1)
        st.markdown("**Seguridad**")
        safety_days = st.number_input("Stock de seguridad (días)", min_value=0, max_value=90, value=3, step=1)
        st.markdown("**Demanda**")
        sales_mode_label = st.selectbox("Tipo de ventas", list(SALES_MODES.keys()))
        sales_mode = SALES_MODES[sales_mode_label]
        st.markdown("**ABC**")
        abc_basis = st.selectbox("Base ABC", ["Movimiento", "Inventario", "Capital inmovilizado"])

    dates = list_snapshot_dates()
    latest_global = dates[0] if dates else None
    status = f"● {latest_global}" if latest_global else "● Sin cortes"
    st.markdown(f'<span class="badge badge-green">{status}</span>', unsafe_allow_html=True)
    if st.button("🔄 Forzar recálculo", use_container_width=True, help="Limpia la caché interna y recalcula todo desde cero."):
        st.cache_data.clear()
        st.toast("Caché limpiada · recalculando motor", icon="🔄")
        st.rerun()

    with st.expander("⚠️ Resetear Sistema"):
        st.warning("Borrará todo el historial y cortes guardados.")
        if st.button("🗑️ Borrar Base de Datos", use_container_width=True, type="primary"):
            reset_database()
            st.session_state.clear()
            st.toast("Sistema reiniciado a cero exitosamente.", icon="✅")
            st.rerun()

    st.caption(f"NEXUS v{APP_VERSION} · SQLite local")


@st.cache_data(show_spinner=False)
def load_live_results(
    _db_version: str,
    months_history: int,
    min_coverage: float,
    max_coverage: float,
    lead_time_days: int,
    safety_days: int,
    abc_basis: str,
    sales_mode: str,
    rolling_days: int,
) -> tuple[dict[str, pd.DataFrame], dict[str, str]]:
    results: dict[str, pd.DataFrame] = {}
    dates_used: dict[str, str] = {}
    for site in SEDES:
        d = latest_snapshot_date(site)
        if not d:
            continue
        inv, sales, prev = load_snapshot(site, d)
        if "Código" not in inv.columns or "Código" not in sales.columns:
            continue
        results[site] = calculate_site(
            inv,
            sales,
            prev,
            months_history=int(months_history),
            min_coverage=float(min_coverage),
            max_coverage=float(max_coverage),
            lead_time_days=int(lead_time_days),
            safety_days=int(safety_days),
            abc_basis=abc_basis,
            sales_mode=sales_mode,
            rolling_days=int(rolling_days),
        )
        dates_used[site] = d
    return results, dates_used



@st.cache_data(show_spinner=False)
def load_supplier_live_results(
    _db_version: str,
    _supplier_version: str,
    supplier_name: str,
    selected_sites: tuple[str, ...],
    months_history: int,
    min_coverage: float,
    max_coverage: float,
    lead_time_days: int,
    safety_days: int,
    abc_basis: str,
    sales_mode: str,
    rolling_days: int,
) -> tuple[dict[str, pd.DataFrame], dict[str, str], dict[str, str]]:
    """
    Carga primero el corte dedicado del proveedor. Si no existe, reutiliza el
    corte maestro filtrando el inventario por proveedor y recalcula desde cero.
    """
    results: dict[str, pd.DataFrame] = {}
    dates_used: dict[str, str] = {}
    sources: dict[str, str] = {}
    key = normalize_key(supplier_name)
    if not key:
        return results, dates_used, sources

    for site in selected_sites:
        dedicated_date = latest_supplier_snapshot_date(site, supplier_name)
        if dedicated_date:
            inv, sales, prev = load_supplier_snapshot(site, supplier_name, dedicated_date)
            d = dedicated_date
            sources[site] = "Corte dedicado del proveedor"
        else:
            d = latest_snapshot_date(site)
            if not d:
                continue
            inv, sales, prev = load_snapshot(site, d)
            if "Proveedor" not in inv.columns:
                continue
            mask = inv["Proveedor"].fillna("").astype(str).map(normalize_key) == key
            inv = inv.loc[mask].copy()
            if inv.empty:
                continue
            codes = set(inv["Código"].astype(str))
            sales = sales[sales["Código"].astype(str).isin(codes)].copy()
            prev = prev[prev["Código"].astype(str).isin(codes)].copy()
            sources[site] = "Corte maestro filtrado por proveedor"

        if inv.empty or "Código" not in inv.columns:
            continue
        if sales.empty:
            sales = pd.DataFrame({"Código": inv["Código"].astype(str), "Ventas": 0})
        results[site] = calculate_site(
            inv, sales, prev,
            months_history=int(months_history), min_coverage=float(min_coverage),
            max_coverage=float(max_coverage), lead_time_days=int(lead_time_days),
            safety_days=int(safety_days), abc_basis=abc_basis, sales_mode=sales_mode,
            rolling_days=int(rolling_days),
        )
        dates_used[site] = d
    return results, dates_used, sources


@st.cache_data(show_spinner=False)
def build_dashboard_state(
    site_results: dict[str, pd.DataFrame],
) -> tuple[pd.DataFrame, dict[str, pd.DataFrame], pd.DataFrame, pd.DataFrame]:
    transfers = build_redistribution(site_results)
    site_results_adjusted = apply_redistribution_to_results(site_results, transfers)
    all_df = combine_sites(site_results_adjusted)
    alerts = build_alerts(all_df)
    return transfers, site_results_adjusted, all_df, alerts



@st.cache_data(show_spinner=False)
def load_full_dashboard_state(
    _db_version: str,
    months_history: int,
    min_coverage: float,
    max_coverage: float,
    lead_time_days: int,
    safety_days: int,
    abc_basis: str,
    sales_mode: str,
    rolling_days: int,
):
    """Cachea el estado completo con argumentos pequeños para evitar re-hashear DataFrames en cada interacción."""
    site_results, dates_used = load_live_results(
        _db_version, months_history, min_coverage, max_coverage,
        lead_time_days, safety_days, abc_basis, sales_mode, rolling_days,
    )
    transfers, adjusted, all_df, alerts = build_dashboard_state(site_results)
    return site_results, dates_used, transfers, adjusted, all_df, alerts


current_db_version = db_version()
current_supplier_version = supplier_db_version()
site_results, dates_used, transfers, site_results_adjusted, all_df, alerts = load_full_dashboard_state(
    current_db_version,
    int(months_history),
    float(min_coverage),
    float(max_coverage),
    int(lead_time_days),
    int(safety_days),
    abc_basis,
    sales_mode,
    int(rolling_days),
)

latest_display = max(dates_used.values()) if dates_used else "sin cortes"
score = health_score(all_df)

st.markdown(
    f"""
    <div class="hero">
        <div class="brand">Makropetrol <span>NEXUS</span></div>
        <div class="hero-sub">Inteligencia logística multisede · inventario · proveedores · rotación · abastecimiento · redistribución · finanzas</div>
        <div class="status-row"><span class="dot"></span> Motor operativo activo <span>•</span> Último corte: {latest_display} <span>•</span> Salud: {score}/100</div>
    </div>
    """,
    unsafe_allow_html=True,
)


def _inventory_column_config(data: pd.DataFrame) -> dict[str, Any]:
    cfg: dict[str, Any] = {}
    if "Cobertura (meses)" in data.columns:
        cfg["Cobertura (meses)"] = st.column_config.NumberColumn(
            "Cobertura", format="%.1f meses", width="small"
        )
    if "Retiro Almacén" in data.columns and not data.empty:
        cfg["Retiro Almacén"] = st.column_config.ProgressColumn(
            "Retiro", min_value=0,
            max_value=max(1, int(pd.to_numeric(data["Retiro Almacén"], errors="coerce").max())),
            format="%d uds",
        )
    if "Compra Ajustada" in data.columns and not data.empty:
        cfg["Compra Ajustada"] = st.column_config.ProgressColumn(
            "Compra neta", min_value=0,
            max_value=max(1, int(pd.to_numeric(data["Compra Ajustada"], errors="coerce").max())),
            format="%d uds",
        )
    return cfg


@st.fragment
def render_inventario_fragment(all_df: pd.DataFrame, modo_gerencia: bool) -> None:
    if all_df.empty:
        empty_state("Sin inventario", "Procesa al menos una sede desde Cortes Diarios.", "📦")
        return

    fsite, fsupplier = st.columns(2)
    with fsite:
        s = st.selectbox("Sede", ["Todas"] + sorted(all_df["Sede"].unique()), key="inv_site_filter")
    data = all_df if s == "Todas" else all_df[all_df["Sede"] == s]
    with fsupplier:
        supplier_values = sorted([x for x in data.get("Proveedor", pd.Series(dtype=str)).dropna().astype(str).unique() if x.strip()], key=normalize_key)
        supplier_pick = st.selectbox("Proveedor", ["Todos"] + supplier_values, key="inv_supplier_filter") if supplier_values else "Todos"
    if supplier_pick != "Todos" and "Proveedor" in data.columns:
        data = data[data["Proveedor"].astype(str) == supplier_pick]

    search = st.text_input(
        "Buscar código o descripción",
        placeholder="Ej. TMB03 / ACEITE / FILTRO",
        key="inv_search",
    )
    if search:
        q = normalize_text(search)
        data = data[
            data["Código"].astype(str).str.lower().str.contains(q, na=False)
            | data["Descripción"].astype(str).str.lower().str.contains(q, na=False)
        ]

    state_options = sorted(data["Estado"].astype(str).unique())
    states = st.multiselect("Estado", state_options, default=state_options, key="inv_states")
    data = data[data["Estado"].astype(str).isin(states)]

    purchase_col = "Compra Ajustada" if "Compra Ajustada" in data.columns else "Compra Sugerida"
    c1, c2, c3, c4 = st.columns(4)
    with c1: kpi("Productos", integer(len(data)), "Filtrados", "blue")
    with c2: kpi("Críticos", integer(data["Estado"].astype(str).str.startswith("CRÍTICO").sum()), "Atención", "red")
    with c3: kpi("Compras", integer(data[purchase_col].sum()), "Unidades", "orange")
    with c4: kpi("Sobrestock", integer(data["Retiro Almacén"].sum()), "Unidades", "green")

    if modo_gerencia:
        cols = ["Sede", "Código", "Descripción", "ABC", "Estado", "Prioridad", "Acción",
                purchase_col, "Retiro Almacén", "Cobertura (meses)",
                "Valor Inventario ($)", "Capital Inmovilizado ($)"]
    else:
        cols = ["Sede", "Código", "Descripción", "Proveedor", "Categoría", "ABC", "Estado",
                "Prioridad", "Acción", "Ventas", "Demanda Mensual", "Existencia",
                "Stock Mínimo", "Stock Máximo", "Punto de Reorden", purchase_col,
                "Retiro Almacén", "Cobertura (meses)", "Rotación estimada (x/mes)",
                "Valor Inventario ($)", "Capital Inmovilizado ($)"]

    cols = [c for c in cols if c in data.columns]
    shown = data[cols].head(750)
    cfg = _inventory_column_config(shown)
    for c in ["Valor Inventario ($)", "Capital Inmovilizado ($)"]:
        if c in shown.columns:
            cfg[c] = st.column_config.NumberColumn(c, format="$%.2f")
    if "Prioridad" in shown.columns:
        cfg["Prioridad"] = st.column_config.TextColumn("Prioridad", width="small")
    if "Descripción" in shown.columns:
        cfg["Descripción"] = st.column_config.TextColumn("Descripción", width="large")

    st.dataframe(
        shown,
        use_container_width=True,
        hide_index=True,
        height=620,
        column_config=cfg,
    )
    if len(data) > len(shown):
        st.caption(f"Mostrando {len(shown):,} de {len(data):,} registros. La exportación conserva el 100% de los datos.")


@st.fragment
def render_compras_fragment(all_df: pd.DataFrame, modo_gerencia: bool) -> None:
    if all_df.empty:
        empty_state("Sin plan de compras", "Carga un corte diario para calcular necesidades.", "🛒")
        return
    purchase_col = "Compra Ajustada" if "Compra Ajustada" in all_df.columns else "Compra Sugerida"
    data = all_df[all_df[purchase_col] > 0]
    if data.empty:
        st.success("No existen compras pendientes después de considerar redistribuciones.")
        return

    c1, c2, c3 = st.columns(3)
    with c1: kpi("SKUs a comprar", integer(len(data)), "Compra neta", "orange")
    with c2: kpi("Unidades", integer(data[purchase_col].sum()), "Necesidad neta", "blue")
    with c3: kpi("Costo estimado", money((data[purchase_col] * data["Costo"]).sum()), "Costo disponible", "green")

    a, b, c = st.columns(3)
    with a:
        sites_filter = st.multiselect("Sedes", sorted(data["Sede"].unique()), default=sorted(data["Sede"].unique()), key="purchase_sites")
    with b:
        priorities = st.multiselect("Prioridad", sorted(data["Prioridad"].unique()), default=sorted(data["Prioridad"].unique()), key="purchase_priorities")
    with c:
        supplier_values = sorted([x for x in data.get("Proveedor", pd.Series(dtype=str)).dropna().astype(str).unique() if x.strip()], key=normalize_key)
        purchase_supplier = st.selectbox("Proveedor", ["Todos"] + supplier_values, key="purchase_supplier") if supplier_values else "Todos"
    data = data[data["Sede"].isin(sites_filter) & data["Prioridad"].isin(priorities)]
    if purchase_supplier != "Todos" and "Proveedor" in data.columns:
        data = data[data["Proveedor"].astype(str) == purchase_supplier]

    if modo_gerencia:
        cols = ["Sede", "Código", "Descripción", "Proveedor", purchase_col, "Costo Compra Estimada ($)", "Prioridad", "Acción"]
    else:
        cols = ["Sede", "Código", "Descripción", "Proveedor", "ABC", "Ventas", "Existencia",
                "Punto de Reorden", purchase_col, "Costo", "Costo Compra Estimada ($)", "Prioridad", "Acción"]
    cols = [c for c in cols if c in data.columns]

    cfg = {
        "Costo": st.column_config.NumberColumn("Costo", format="$%.2f"),
        "Costo Compra Estimada ($)": st.column_config.NumberColumn("Costo estimado", format="$%.2f"),
    }
    if purchase_col in data.columns and not data.empty:
        cfg[purchase_col] = st.column_config.ProgressColumn(
            "Compra neta", min_value=0,
            max_value=max(1, int(data[purchase_col].max())),
            format="%d uds",
        )
    st.dataframe(data[cols].head(750), use_container_width=True, hide_index=True, height=580, column_config=cfg)
    if len(data) > 750:
        st.caption(f"Mostrando 750 de {len(data):,} registros. La exportación conserva todos.")

    supplier = data.groupby("Proveedor", dropna=False)[purchase_col].sum().reset_index().sort_values(purchase_col, ascending=False).head(10)
    fig = px.bar(
        supplier, x=purchase_col, y="Proveedor", orientation="h",
        title="Top proveedores por unidades a comprar",
        color_discrete_sequence=[NEXUS_ORANGE],
    )
    st.plotly_chart(chart_layout(fig, 320), use_container_width=True)


@st.fragment
def render_alertas_fragment(alerts: pd.DataFrame) -> None:
    if alerts.empty:
        empty_state("Todo tranquilo", "No se detectaron anomalías con los parámetros actuales.", "✓")
        return
    sev = st.multiselect("Severidad", sorted(alerts["Severidad"].unique()), default=sorted(alerts["Severidad"].unique()), key="alert_sev")
    typ = st.multiselect("Tipo", sorted(alerts["Tipo"].unique()), default=sorted(alerts["Tipo"].unique()), key="alert_type")
    data = alerts[alerts["Severidad"].isin(sev) & alerts["Tipo"].isin(typ)]
    counts = data["Severidad"].value_counts().reset_index()
    counts.columns = ["Severidad", "Cantidad"]
    fig = px.bar(
        counts, x="Severidad", y="Cantidad",
        title="Alertas por severidad",
        color="Severidad",
        color_discrete_map={"CRÍTICA": NEXUS_RED, "ALTA": NEXUS_ORANGE, "MEDIA": NEXUS_BLUE, "BAJA": NEXUS_GRAY},
    )
    st.plotly_chart(chart_layout(fig, 280), use_container_width=True)
    st.dataframe(data.head(750), use_container_width=True, hide_index=True, height=580)
    if len(data) > 750:
        st.caption(f"Mostrando 750 de {len(data):,} alertas.")


@st.fragment
def render_automation_fragment(
    site_results_adjusted: dict[str, pd.DataFrame],
    transfers: pd.DataFrame,
    all_df: pd.DataFrame,
    contacts: dict[str, dict[str, str]],
) -> None:
    contact_rows = []
    for site in SEDES:
        c = contacts.get(site, {})
        contact_rows.append({
            "Sede": site,
            "Supervisor": c.get("supervisor", ""),
            "WhatsApp": c.get("whatsapp", ""),
            "Webhook": c.get("webhook", ""),
        })
    edited = st.data_editor(
        pd.DataFrame(contact_rows),
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        key="contacts_editor",
    )
    if st.button("💾 Guardar contactos y webhooks", use_container_width=True):
        new_contacts = {}
        for row in edited.to_dict("records"):
            new_contacts[str(row["Sede"])] = {
                "supervisor": str(row.get("Supervisor", "")),
                "whatsapp": str(row.get("WhatsApp", "")),
                "webhook": str(row.get("Webhook", "")),
            }
        save_contacts(new_contacts)
        contacts.clear()
        contacts.update(new_contacts)
        st.toast("Contactos y webhooks guardados", icon="💾")

    if all_df.empty:
        empty_state("Automatización lista para conectar", "Configura los webhooks y procesa un corte diario.", "⚙")
        return

    site = st.selectbox("Sede a notificar", sorted(site_results_adjusted.keys()), key="automation_site")
    message = compose_digest(site, site_results_adjusted[site], transfers)
    st.code(message, language="text")

    supervisor = contacts.get(site, {})
    df_site = site_results_adjusted[site]
    purchase_col = "Compra Ajustada" if "Compra Ajustada" in df_site.columns else "Compra Sugerida"
    payload = {
        "event": "nexus.daily.logistics",
        "version": APP_VERSION,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "site": site,
        "supervisor": supervisor.get("supervisor", ""),
        "whatsapp": supervisor.get("whatsapp", ""),
        "health_score": health_score(df_site),
        "message": message,
        "metrics": {
            "critical": int(df_site["Estado"].astype(str).str.startswith("CRÍTICO").sum()),
            "purchase_skus": int((df_site[purchase_col] > 0).sum()),
            "purchase_units": int(df_site[purchase_col].sum()),
            "withdrawal_skus": int((df_site["Retiro Almacén"] > 0).sum()),
            "inventory_value": float(df_site["Valor Inventario ($)"].sum()),
            "capital_immobilized": float(df_site["Capital Inmovilizado ($)"].sum()),
        },
    }

    with st.expander("Vista técnica del payload", expanded=False):
        st.json(payload)

    if st.button("🚀 Enviar al webhook de esta sede", use_container_width=True):
        future = post_webhook_async(supervisor.get("webhook", ""), payload)
        if future is None:
            st.error("No se envió: webhook vacío.")
        else:
            st.toast("Webhook disparado en segundo plano", icon="🚀")
            st.session_state["last_webhook_site"] = site
    st.info("Arquitectura: NEXUS → webhook → n8n/Make/API → CRM + proveedor de WhatsApp.")


# ============================================================
# VISTAS PRINCIPALES
# ============================================================

if view == "Centro de Control":
    if modo_gerencia:
        st.markdown('<div class="gov-banner">👔 Modo Gerencia activo · vista simplificada de indicadores y recomendaciones</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Centro de Control Ejecutivo</div>', unsafe_allow_html=True)
    if all_df.empty:
        empty_state("NEXUS está listo", "Carga un corte diario de Inventario + Ventas desde 'Cortes Diarios' para activar el tablero.", "◈")
    else:
        critical = int(all_df["Estado"].str.startswith("CRÍTICO").sum())
        purchase_col = "Compra Ajustada" if "Compra Ajustada" in all_df.columns else "Compra Sugerida"
        purchases = int((all_df[purchase_col] > 0).sum())
        purchase_units = int(all_df[purchase_col].sum())
        withdrawals = int((all_df["Retiro Almacén"] > 0).sum())
        inventory_value = float(all_df["Valor Inventario ($)"].sum())
        immobilized = float(all_df["Capital Inmovilizado ($)"].sum())
        avoided = float(transfers["Compra Evitada Estimada ($)"].sum()) if not transfers.empty else 0.0
        optimal = int((all_df["Estado"] == "ÓPTIMO").sum())

        c1, c2, c3, c4 = st.columns(4)
        with c1: kpi("Inventario", money(inventory_value), f"{len(all_df):,} SKUs", "blue")
        with c2: kpi("Críticos", integer(critical), "Saldo negativo + bajo reorden", "red")
        with c3: kpi("Compra", f"{integer(purchase_units)} uds", f"{purchases} SKUs a comprar", "orange")
        with c4: kpi("Capital", money(immobilized), "Inmovilizado por sobrestock", "green")

        st.markdown('<div class="section-title">Salud logística</div>', unsafe_allow_html=True)
        st.markdown(
            f"""<div class="panel health-big">
                <div class="num">{score} / 100</div>
                <div class="lbl">{health_label(score)}</div>
            </div>""",
            unsafe_allow_html=True,
        )

        left, right = st.columns(2)
        with left:
            st.markdown('<div class="section-title">Inventario por sede</div>', unsafe_allow_html=True)
            by_site = all_df.groupby("Sede")["Valor Inventario ($)"].sum().sort_values(ascending=False)
            max_val = float(by_site.max()) if len(by_site) else 1.0
            rows_html = ""
            for sede, val in by_site.items():
                pct = max(4, round(val / max_val * 100)) if max_val else 4
                rows_html += (
                    f'<div class="op-row"><span>{SEDE_ICON.get(sede, "◆")} {sede}</span>'
                    f'<div class="bar-track"><div class="bar-fill" style="width:{pct}%"></div></div>'
                    f'<span>{money(val)}</span></div>'
                )
            st.markdown(f'<div class="panel">{rows_html}</div>', unsafe_allow_html=True)
        with right:
            st.markdown('<div class="section-title">Estado operativo</div>', unsafe_allow_html=True)
            state_rows = [
                ("Óptimo", optimal, "green"),
                ("Comprar", purchases, "orange"),
                ("Retirar", withdrawals, "blue"),
                ("Crítico", critical, "red"),
            ]
            rows_html = "".join(
                f'<div class="op-row"><span>{label}</span><span class="badge badge-{tone}">{integer(count)}</span></div>'
                for label, count, tone in state_rows
            )
            st.markdown(f'<div class="panel">{rows_html}</div>', unsafe_allow_html=True)

        n_transfers = int(len(transfers)) if not transfers.empty else 0
        st.markdown(
            f"""<div class="opp-banner">
                <span>♻️ {n_transfers} oportunidad(es) de redistribución</span>
                <span>💰 {money(avoided)} de compra potencial evitada</span>
            </div>""",
            unsafe_allow_html=True,
        )

        if not modo_gerencia:
            st.markdown('<div class="section-title">Prioridades inmediatas</div>', unsafe_allow_html=True)
            top_alerts = alerts.head(15) if not alerts.empty else pd.DataFrame()
            if top_alerts.empty:
                st.success("No hay alertas activas con los parámetros actuales.")
            else:
                st.dataframe(top_alerts, use_container_width=True, hide_index=True, height=380)

            a, b = st.columns(2)
            with a:
                action_counts = pd.DataFrame({
                    "Acción": ["Óptimo", "Comprar", "Retirar", "Crítico"],
                    "Cantidad": [optimal, purchases, withdrawals, critical],
                })
                fig = px.pie(action_counts, names="Acción", values="Cantidad", hole=.68, color="Acción", color_discrete_map={"Óptimo":NEXUS_GREEN,"Comprar":NEXUS_ORANGE,"Retirar":NEXUS_BLUE,"Crítico":NEXUS_RED})
                fig.update_traces(textposition="inside", textinfo="percent+label")
                st.plotly_chart(chart_layout(fig, 300), use_container_width=True)
            with b:
                top = all_df[all_df["Capital Inmovilizado ($)"] > 0].sort_values("Capital Inmovilizado ($)", ascending=False).head(12)
                if top.empty:
                    st.success("No existe capital inmovilizado por sobrestock.")
                else:
                    fig = px.bar(top.sort_values("Capital Inmovilizado ($)"), x="Capital Inmovilizado ($)", y="Descripción", orientation="h", hover_data=["Sede", "Código", "Retiro Almacén"], color_discrete_sequence=[NEXUS_RED])
                    st.plotly_chart(chart_layout(fig, 300), use_container_width=True)
        else:
            st.markdown('<div class="section-title">Recomendaciones para dirección</div>', unsafe_allow_html=True)
            recs = []
            if critical > 0:
                recs.append(f"🔴 Atender **{critical} SKUs críticos** (saldo negativo o bajo el punto de reorden) antes de nueva compra.")
            if n_transfers > 0:
                recs.append(f"♻️ Ejecutar las **{n_transfers} redistribuciones** sugeridas: evitan {money(avoided)} en compras nuevas.")
            if withdrawals > 0:
                recs.append(f"📤 Hay **{withdrawals} SKUs en sobrestock**; considerar retiro o promoción para liberar capital.")
            if not recs:
                recs.append("✅ La operación está dentro de parámetros saludables. No se requieren acciones urgentes.")
            st.markdown('<div class="panel">' + "".join(f"<p>{r}</p>" for r in recs) + "</div>", unsafe_allow_html=True)

elif view == "Cortes Diarios":
    st.markdown('<div class="section-title">Cortes diarios · Fuente de verdad</div>', unsafe_allow_html=True)
    st.caption("Carga cortes generales de todas las sedes o analiza una marca/proveedor de forma independiente sin tocar el inventario maestro.")

    cut_mode = st.radio(
        "Tipo de corte",
        ["🌐 Corte general por sedes", "🏷️ Corte específico por proveedor"],
        horizontal=True,
        key="daily_cut_mode",
    )
    chosen_date = st.date_input("Fecha del corte", value=date.today(), key="daily_cut_date")
    snapshot_str = chosen_date.isoformat()

    if "general" in cut_mode.lower():
        st.markdown(
            '<div class="scope-banner"><div><div class="scope-title">Corte maestro multisede</div>'
            '<div class="scope-sub">Actualiza la fuente general usada por inventario, compras, finanzas y redistribución.</div></div>'
            '<span class="badge badge-blue">GENERAL</span></div>',
            unsafe_allow_html=True,
        )
        all_sites = st.checkbox("✅ Seleccionar todas las sedes", value=True, key="select_all_sites")
        if all_sites:
            selected_sites = SEDES.copy()
        else:
            selected_sites = st.multiselect("Sedes a actualizar", SEDES, default=SEDES, key="selected_sites")

        if not selected_sites:
            empty_state("Selecciona al menos una sede", "Puedes trabajar con todas las sedes o con una selección específica.", "🏢")
        else:
            upload_map = {}
            for site in selected_sites:
                with st.expander(f"{SEDE_ICON[site]}  {site}", expanded=True):
                    u1, u2 = st.columns(2)
                    with u1:
                        inv_file = st.file_uploader(
                            f"📦 Inventario + costos · {site}", type=["xlsx", "xls"],
                            key=f"daily_inv_{site}_{snapshot_str}",
                        )
                    with u2:
                        sales_file = st.file_uploader(
                            f"📈 Ventas · {site}", type=["xlsx", "xls"],
                            key=f"daily_sales_{site}_{snapshot_str}",
                        )
                    upload_map[site] = (inv_file, sales_file)

            ready_sites = [site for site, (invf, salf) in upload_map.items() if invf and salf]
            incomplete = [site for site, (invf, salf) in upload_map.items() if bool(invf) ^ bool(salf)]
            if incomplete:
                st.warning("Sedes incompletas (falta uno de los dos archivos): " + ", ".join(incomplete))
            if ready_sites:
                st.markdown(f"**{len(ready_sites)} sede(s) listas para procesar.**")
                if st.button("💾 Guardar y procesar sedes listas", type="primary", use_container_width=True, key=f"save_all_{snapshot_str}"):
                    errors = []
                    progress = st.progress(0)
                    for idx, site in enumerate(ready_sites, 1):
                        try:
                            inv_file, sales_file = upload_map[site]
                            inv, _ = parse_uploaded_excel(inv_file.getvalue(), "inventario")
                            sales, _ = parse_uploaded_excel(sales_file.getvalue(), "ventas")
                            if "Código" not in inv.columns or "Existencia" not in inv.columns:
                                raise ValueError("Inventario requiere Código y Existencia")
                            if "Código" not in sales.columns or "Ventas" not in sales.columns:
                                raise ValueError("Ventas requiere Código y Ventas")
                            save_snapshot(site, snapshot_str, inv, sales)
                        except Exception as exc:
                            errors.append(f"{site}: {exc}")
                        progress.progress(idx / max(1, len(ready_sites)))
                    if errors:
                        for err in errors:
                            st.error(err)
                    if len(errors) < len(ready_sites):
                        st.toast(f"Corte {snapshot_str} guardado · {len(ready_sites)-len(errors)} sede(s) recalculadas", icon="🚀")
                        st.rerun()

        st.markdown('<div class="section-title">Estado de actualización</div>', unsafe_allow_html=True)
        status_rows = []
        for site in SEDES:
            last = latest_snapshot_date(site)
            status_rows.append({
                "Sede": f"{SEDE_ICON[site]} {site}", "Último corte": last or "Sin datos",
                "Estado": "ACTUALIZADA" if last == snapshot_str else "PENDIENTE",
            })
        st.dataframe(pd.DataFrame(status_rows), use_container_width=True, hide_index=True)

    else:
        st.markdown(
            '<div class="scope-banner"><div><div class="scope-title">Corte dedicado por proveedor / marca</div>'
            '<div class="scope-sub">Ideal para tus ~60 líneas: escribe la marca, carga sus archivos y NEXUS calcula/reportará únicamente ese proveedor.</div></div>'
            '<span class="badge badge-cyan">AISLADO</span></div>',
            unsafe_allow_html=True,
        )
        known_suppliers = list_known_suppliers(current_db_version, current_supplier_version)
        q_col, pick_col = st.columns([1.25, 1])
        with q_col:
            supplier_text = st.text_input(
                "Proveedor / marca",
                value=st.session_state.get("supplier_focus", ""),
                placeholder="Ej. SHELL, MOBIL, CASTROL, BOSCH...",
                key="supplier_cut_text",
                help="Puedes escribir un proveedor nuevo aunque todavía no exista en la base.",
            ).strip()
        matches = supplier_matches(supplier_text, known_suppliers) if supplier_text else known_suppliers[:12]
        with pick_col:
            suggestion_options = ["Usar exactamente el texto escrito"] + matches if supplier_text else (["Selecciona o escribe a la izquierda"] + matches)
            supplier_suggestion = st.selectbox("Coincidencias", suggestion_options, key="supplier_cut_suggestion")
        supplier_name = supplier_text
        if supplier_suggestion not in {"Usar exactamente el texto escrito", "Selecciona o escribe a la izquierda"}:
            supplier_name = supplier_suggestion

        p_all = st.checkbox("✅ Todas las sedes para este proveedor", value=True, key="supplier_all_sites")
        supplier_sites = SEDES.copy() if p_all else st.multiselect("Sedes del proveedor", SEDES, default=SEDES, key="supplier_sites")

        if not supplier_name:
            empty_state("Escribe el proveedor", "Cuando indiques la marca se habilitará la carga independiente de Inventario + Ventas.", "🏷️")
        elif not supplier_sites:
            empty_state("Selecciona al menos una sede", "El proveedor puede analizarse en una o varias sedes.", "🏢")
        else:
            st.markdown(f"<span class='metric-chip'>🏷️ Proveedor activo: <b>{supplier_name}</b></span>", unsafe_allow_html=True)
            supplier_upload_map = {}
            for site in supplier_sites:
                with st.expander(f"{SEDE_ICON[site]}  {site} · {supplier_name}", expanded=True):
                    s1, s2 = st.columns(2)
                    with s1:
                        inv_file = st.file_uploader(
                            f"📦 Inventario + costos · {supplier_name} · {site}", type=["xlsx", "xls"],
                            key=f"sup_inv_{safe_slug(supplier_name)}_{site}_{snapshot_str}",
                        )
                    with s2:
                        sales_file = st.file_uploader(
                            f"📈 Ventas · {supplier_name} · {site}", type=["xlsx", "xls"],
                            key=f"sup_sales_{safe_slug(supplier_name)}_{site}_{snapshot_str}",
                        )
                    supplier_upload_map[site] = (inv_file, sales_file)

            ready = [site for site, pair in supplier_upload_map.items() if pair[0] and pair[1]]
            incomplete = [site for site, pair in supplier_upload_map.items() if bool(pair[0]) ^ bool(pair[1])]
            if incomplete:
                st.warning("Sedes incompletas para este proveedor: " + ", ".join(incomplete))
            if ready and st.button("🏷️ Guardar corte del proveedor", type="primary", use_container_width=True, key=f"save_supplier_{snapshot_str}_{safe_slug(supplier_name)}"):
                errors = []
                progress = st.progress(0)
                for idx, site in enumerate(ready, 1):
                    try:
                        inv_file, sales_file = supplier_upload_map[site]
                        inv, _ = parse_uploaded_excel(inv_file.getvalue(), "inventario")
                        sales, _ = parse_uploaded_excel(sales_file.getvalue(), "ventas")
                        if "Código" not in inv.columns or "Existencia" not in inv.columns:
                            raise ValueError("Inventario requiere Código y Existencia")
                        if "Código" not in sales.columns or "Ventas" not in sales.columns:
                            raise ValueError("Ventas requiere Código y Ventas")
                        save_supplier_snapshot(site, snapshot_str, supplier_name, inv, sales)
                    except Exception as exc:
                        errors.append(f"{site}: {exc}")
                    progress.progress(idx / max(1, len(ready)))
                if errors:
                    for err in errors:
                        st.error(err)
                if len(errors) < len(ready):
                    st.session_state["supplier_focus"] = supplier_name
                    st.toast(f"Proveedor {supplier_name} guardado · {len(ready)-len(errors)} sede(s)", icon="🏷️")
                    st.rerun()

            st.markdown('<div class="section-title">Estado del proveedor</div>', unsafe_allow_html=True)
            supplier_status = []
            for site in supplier_sites:
                last = latest_supplier_snapshot_date(site, supplier_name)
                supplier_status.append({
                    "Sede": f"{SEDE_ICON[site]} {site}", "Proveedor": supplier_name,
                    "Último corte dedicado": last or "Sin corte dedicado",
                    "Estado": "ACTUALIZADA" if last == snapshot_str else ("DISPONIBLE" if last else "SIN DATOS"),
                })
            st.dataframe(pd.DataFrame(supplier_status), use_container_width=True, hide_index=True)
            if any(r["Último corte dedicado"] != "Sin corte dedicado" for r in supplier_status):
                if st.button("📊 Abrir análisis de este proveedor", use_container_width=True):
                    st.session_state["supplier_focus"] = supplier_name
                    st.session_state["nexus_view"] = "Proveedores"
                    st.rerun()

elif view == "Centro Visual":
    st.markdown('<div class="section-title">Centro Visual · lectura rápida para gerencia</div>', unsafe_allow_html=True)
    if all_df.empty:
        empty_state("Sin datos visuales", "Carga cortes diarios primero.", "◌")
    else:
        purchase_col = "Compra Ajustada" if "Compra Ajustada" in all_df.columns else "Compra Sugerida"
        critical_df = all_df[all_df["Estado"].astype(str).str.startswith("CRÍTICO")].copy()
        purchase_df = all_df[all_df[purchase_col] > 0].copy()
        over_df = all_df[all_df["Retiro Almacén"] > 0].copy()
        k1,k2,k3,k4 = st.columns(4)
        with k1: kpi("Salud logística", f"{score}/100", health_label(score), "blue")
        with k2: kpi("Críticos", integer(len(critical_df)), "Atención inmediata", "red")
        with k3: kpi("Compra neta", integer(purchase_df[purchase_col].sum()), "Unidades", "orange")
        with k4: kpi("Sobrestock", integer(over_df["Retiro Almacén"].sum()), "Unidades", "green")

        v1,v2 = st.columns(2)
        with v1:
            site = all_df.groupby("Sede")["Valor Inventario ($)"].sum().reset_index()
            fig=px.bar(site,x="Sede",y="Valor Inventario ($)",title="Valor de inventario por sede", color="Sede", color_discrete_sequence=NEXUS_PALETTE)
            st.plotly_chart(chart_layout(fig,360),use_container_width=True)
        with v2:
            act=pd.DataFrame({"Tipo":["Crítico","Comprar","Retirar","Óptimo"],"Cantidad":[len(critical_df),len(purchase_df),len(over_df),int((all_df["Estado"]=="ÓPTIMO").sum())]})
            fig=px.pie(act,names="Tipo",values="Cantidad",hole=.62,title="Estado operativo")
            st.plotly_chart(chart_layout(fig,360),use_container_width=True)

        v3,v4 = st.columns(2)
        with v3:
            top=all_df[all_df["Capital Inmovilizado ($)"]>0].sort_values("Capital Inmovilizado ($)",ascending=False).head(10)
            if not top.empty:
                fig=px.bar(top.sort_values("Capital Inmovilizado ($)"),x="Capital Inmovilizado ($)",y="Descripción",orientation="h",title="Capital inmovilizado")
                st.plotly_chart(chart_layout(fig,360),use_container_width=True)
            else:
                empty_state("Sin capital inmovilizado", "No hay sobrestock valorizado.", "💰")
        with v4:
            if not transfers.empty:
                routes=transfers.groupby(["Origen","Destino"])["Unidades Sugeridas"].sum().reset_index()
                fig=px.bar(routes,x="Unidades Sugeridas",y="Origen",color="Destino",orientation="h",title="Redistribución sugerida")
                st.plotly_chart(chart_layout(fig,360),use_container_width=True)
            else:
                empty_state("Sin redistribuciones", "No se detectaron cruces de excedente/déficit.", "♻")


        v5, v6 = st.columns(2)
        with v5:
            if "Proveedor" in all_df.columns:
                top_sup = all_df.groupby("Proveedor", dropna=False)["Valor Inventario ($)"].sum().reset_index()
                top_sup = top_sup[top_sup["Proveedor"].astype(str).str.strip() != ""].sort_values("Valor Inventario ($)", ascending=False).head(12)
                if not top_sup.empty:
                    fig = px.bar(top_sup.sort_values("Valor Inventario ($)"), x="Valor Inventario ($)", y="Proveedor", orientation="h", title="Top proveedores por valor de inventario")
                    st.plotly_chart(chart_layout(fig, 360), use_container_width=True)
        with v6:
            dead = all_df[(all_df["Ventas"] <= 0) & (all_df["Existencia"] > 0)].copy()
            if not dead.empty:
                dead_site = dead.groupby("Sede")["Valor Inventario ($)"].sum().reset_index()
                fig = px.bar(dead_site, x="Sede", y="Valor Inventario ($)", color="Sede", title="Capital sin rotación por sede", color_discrete_sequence=NEXUS_PALETTE)
                st.plotly_chart(chart_layout(fig, 360), use_container_width=True)
            else:
                empty_state("Sin inventario detenido", "Todos los productos con stock presentan movimiento.", "✓")

elif view == "Inventario":
    st.markdown('<div class="section-title">Explorador inteligente de inventario</div>', unsafe_allow_html=True)
    render_inventario_fragment(all_df, modo_gerencia)

elif view == "Compras":
    st.markdown('<div class="section-title">Centro de abastecimiento</div>', unsafe_allow_html=True)
    render_compras_fragment(all_df, modo_gerencia)


elif view == "Proveedores":
    st.markdown('<div class="section-title">Inteligencia por proveedor / marca</div>', unsafe_allow_html=True)
    st.caption("Busca cualquiera de tus líneas, recalcula únicamente esa marca y genera reportes aislados por proveedor.")

    known_suppliers = list_known_suppliers(current_db_version, current_supplier_version)
    if not known_suppliers and all_df.empty:
        empty_state("Aún no hay proveedores", "Carga un corte general o un corte dedicado desde Cortes Diarios.", "🏷️")
    else:
        p1, p2 = st.columns([1.2, 1])
        with p1:
            supplier_query = st.text_input(
                "Buscar proveedor",
                value=st.session_state.get("supplier_focus", ""),
                placeholder="Escribe nombre o marca...",
                key="supplier_analysis_query",
            ).strip()
        matches = supplier_matches(supplier_query, known_suppliers, 20) if supplier_query else known_suppliers[:20]
        with p2:
            if matches:
                default_index = 0
                selected_supplier = st.selectbox("Proveedor detectado", matches, index=default_index, key="supplier_analysis_pick")
            else:
                selected_supplier = supplier_query
                st.text_input("Proveedor detectado", value=selected_supplier, disabled=True, key="supplier_analysis_no_match")

        supplier_name = selected_supplier or supplier_query
        if not supplier_name:
            empty_state("Selecciona una marca", "Escribe el nombre del proveedor para abrir su tablero dedicado.", "⌕")
        else:
            st.session_state["supplier_focus"] = supplier_name
            available_sites = SEDES.copy()
            selected_supplier_sites = st.multiselect(
                "Sedes incluidas en el análisis",
                available_sites,
                default=available_sites,
                key=f"supplier_analysis_sites_{safe_slug(supplier_name)}",
            )
            if not selected_supplier_sites:
                empty_state("Sin sedes", "Selecciona al menos una sede para calcular el proveedor.", "🏢")
            else:
                supplier_results, supplier_dates, supplier_sources = load_supplier_live_results(
                    current_db_version, current_supplier_version, supplier_name,
                    tuple(selected_supplier_sites), int(months_history), float(min_coverage), float(max_coverage),
                    int(lead_time_days), int(safety_days), abc_basis, sales_mode, int(rolling_days),
                )
                sup_transfers, supplier_adjusted, supplier_df, supplier_alerts = build_dashboard_state(supplier_results)

                if supplier_df.empty:
                    empty_state(
                        f"Sin datos para {supplier_name}",
                        "No encontré esta marca en el corte maestro ni en cortes dedicados. Ve a Cortes Diarios → Corte específico por proveedor y carga sus archivos.",
                        "🏷️",
                    )
                else:
                    purchase_col = "Compra Ajustada" if "Compra Ajustada" in supplier_df.columns else "Compra Sugerida"
                    inventory_value = float(supplier_df["Valor Inventario ($)"].sum())
                    purchase_units = int(supplier_df[purchase_col].sum())
                    purchase_value = float((supplier_df[purchase_col] * supplier_df["Costo"]).sum())
                    zero_df = supplier_df[(supplier_df["Ventas"] <= 0) & (supplier_df["Existencia"] > 0)].copy()
                    zero_value = float(zero_df["Valor Inventario ($)"].sum()) if not zero_df.empty else 0.0
                    sup_score = health_score(supplier_df)
                    source_summary = " · ".join(f"{site}: {supplier_sources.get(site,'—')}" for site in supplier_dates)

                    st.markdown(
                        f"""<div class="scope-banner"><div><div class="scope-title">🏷️ {supplier_name}</div>
                        <div class="scope-sub">{source_summary or 'Proveedor cargado'} · Salud {sup_score}/100</div></div>
                        <span class="badge badge-cyan">{len(supplier_df):,} SKU</span></div>""",
                        unsafe_allow_html=True,
                    )
                    k1, k2, k3, k4 = st.columns(4)
                    with k1: kpi("Inventario marca", money(inventory_value), f"{len(supplier_df):,} SKU", "blue")
                    with k2: kpi("Compra neta", f"{integer(purchase_units)} uds", money(purchase_value), "orange")
                    with k3: kpi("Sin rotación", integer(len(zero_df)), money(zero_value), "red")
                    with k4: kpi("Salud proveedor", f"{sup_score}/100", health_label(sup_score), "green")

                    tab_res, tab_prod, tab_buy, tab_zero, tab_reports = st.tabs([
                        "📊 Resumen", "📦 Productos", "🛒 Compras", "🧊 Sin rotación", "📄 Reportes"
                    ])
                    with tab_res:
                        a, b = st.columns(2)
                        with a:
                            by_site = supplier_df.groupby("Sede")["Valor Inventario ($)"].sum().reset_index()
                            fig = px.bar(by_site, x="Sede", y="Valor Inventario ($)", color="Sede", title="Valor de inventario de la marca por sede", color_discrete_sequence=NEXUS_PALETTE)
                            st.plotly_chart(chart_layout(fig, 340), use_container_width=True)
                        with b:
                            state = supplier_df["Estado"].astype(str).value_counts().reset_index()
                            state.columns = ["Estado", "Cantidad"]
                            fig = px.pie(state, names="Estado", values="Cantidad", hole=.66, title="Estado logístico de la marca")
                            st.plotly_chart(chart_layout(fig, 340), use_container_width=True)
                        c, d = st.columns(2)
                        with c:
                            buy_site = supplier_df.groupby("Sede")[purchase_col].sum().reset_index()
                            fig = px.bar(buy_site, x="Sede", y=purchase_col, title="Unidades a comprar por sede", color="Sede", color_discrete_sequence=NEXUS_PALETTE)
                            st.plotly_chart(chart_layout(fig, 330), use_container_width=True)
                        with d:
                            top = supplier_df.sort_values("Valor Inventario ($)", ascending=False).head(12)
                            fig = px.bar(top.sort_values("Valor Inventario ($)"), x="Valor Inventario ($)", y="Descripción", orientation="h", title="Productos con mayor valor de inventario")
                            st.plotly_chart(chart_layout(fig, 330), use_container_width=True)

                    with tab_prod:
                        prod_cols = ["Sede","Código","Descripción","Categoría","ABC","Estado","Ventas","Existencia","Punto de Reorden",purchase_col,"Retiro Almacén","Cobertura (meses)","Valor Inventario ($)"]
                        prod_cols = [c for c in prod_cols if c in supplier_df.columns]
                        st.dataframe(supplier_df[prod_cols], use_container_width=True, hide_index=True, height=620)

                    with tab_buy:
                        buy_df = supplier_df[supplier_df[purchase_col] > 0].copy()
                        if buy_df.empty:
                            st.success("No hay compras pendientes para esta marca con los parámetros actuales.")
                        else:
                            cols = ["Sede","Código","Descripción","ABC","Existencia","Punto de Reorden",purchase_col,"Costo","Prioridad","Acción"]
                            cols = [c for c in cols if c in buy_df.columns]
                            st.dataframe(buy_df[cols], use_container_width=True, hide_index=True, height=540)

                    with tab_zero:
                        if zero_df.empty:
                            st.success("Todos los productos con stock registran movimiento en el corte actual.")
                        else:
                            top_zero = zero_df.sort_values("Valor Inventario ($)", ascending=False).head(20)
                            fig = px.bar(top_zero.sort_values("Valor Inventario ($)"), x="Valor Inventario ($)", y="Descripción", color="Sede", orientation="h", title="Capital de la marca sin rotación")
                            st.plotly_chart(chart_layout(fig, 380), use_container_width=True)
                            st.dataframe(zero_df[[c for c in ["Sede","Código","Descripción","Existencia","Costo","Valor Inventario ($)","ABC"] if c in zero_df.columns]], use_container_width=True, hide_index=True, height=450)

                    with tab_reports:
                        st.markdown("**Todo lo que descargues aquí contiene únicamente este proveedor.**")
                        slug = safe_slug(supplier_name)
                        sup_latest = max(supplier_dates.values()) if supplier_dates else latest_display
                        report_token = safe_slug(
                            f"{slug}_{sup_latest}_{sup_score}_{months_history}_{min_coverage}_{max_coverage}_{lead_time_days}_{safety_days}_{abc_basis}_{sales_mode}"
                        )
                        report_state_key = f"supplier_report_files_{report_token}"
                        st.info("Los archivos se preparan bajo demanda para no ralentizar filtros y gráficos.")
                        if st.button("⚙️ Preparar reportes de este proveedor", type="primary", use_container_width=True, key=f"prepare_{report_token}"):
                            with st.spinner("Generando Excel, HTML y PDF del proveedor..."):
                                st.session_state[report_state_key] = {
                                    "excel": export_excel(
                                        supplier_adjusted, sup_transfers, int(months_history), float(min_coverage),
                                        float(max_coverage), int(lead_time_days), sup_latest,
                                    ),
                                    "html": build_interactive_html_report(
                                        supplier_df, sup_transfers, supplier_alerts,
                                        f"{sup_latest} · {supplier_name}", sup_score,
                                    ),
                                    "pdf": build_executive_pdf(
                                        supplier_df, sup_transfers, f"{sup_latest} · {supplier_name}", sup_score,
                                    ),
                                }
                            st.toast("Reportes del proveedor preparados", icon="📄")

                        report_files = st.session_state.get(report_state_key)
                        if report_files:
                            r1, r2, r3 = st.columns(3)
                            with r1:
                                st.download_button("📊 Excel del proveedor", report_files["excel"], f"NEXUS_{slug}_{sup_latest}.xlsx", DOC_MIME_XLSX, use_container_width=True)
                            with r2:
                                st.download_button("🌐 Dashboard HTML", report_files["html"], f"NEXUS_{slug}_{sup_latest}.html", DOC_MIME_HTML, use_container_width=True)
                            with r3:
                                if report_files.get("pdf"):
                                    st.download_button("📄 PDF ejecutivo", report_files["pdf"], f"NEXUS_{slug}_{sup_latest}.pdf", DOC_MIME_PDF, use_container_width=True)
                                else:
                                    st.info("Instala fpdf2 para PDF.")

                        purchase_sites = sorted(supplier_df["Sede"].unique())
                        purchase_site = st.selectbox("Orden de compra para la sede", purchase_sites, key=f"supplier_purchase_site_{slug}")
                        purchase_state_key = f"supplier_purchase_pdf_{report_token}_{safe_slug(purchase_site)}"
                        if st.button("🛒 Preparar orden de compra de esta marca", use_container_width=True, key=f"prepare_oc_{report_token}_{safe_slug(purchase_site)}"):
                            site_df = supplier_df[supplier_df["Sede"] == purchase_site].copy()
                            st.session_state[purchase_state_key] = build_purchase_pdf(
                                site_df, purchase_site, sup_latest, supplier=supplier_name
                            )
                        purchase_pdf = st.session_state.get(purchase_state_key)
                        if purchase_pdf:
                            st.download_button(
                                "⬇ Descargar orden de compra",
                                purchase_pdf,
                                f"OC_{slug}_{safe_slug(purchase_site)}_{sup_latest}.pdf",
                                DOC_MIME_PDF,
                                use_container_width=True,
                            )


elif view == "Sin Rotación":
    st.markdown('<div class="section-title">Productos sin rotación</div>', unsafe_allow_html=True)
    st.caption("Detecta inventario con existencia positiva y cero venta en el corte actual; además cruza el historial para mostrar cuándo se observó la última venta positiva.")
    if all_df.empty:
        empty_state("Sin datos de rotación", "Carga al menos un corte diario para activar el análisis.", "🧊")
    else:
        zero = all_df[(pd.to_numeric(all_df["Ventas"], errors="coerce").fillna(0) <= 0) & (pd.to_numeric(all_df["Existencia"], errors="coerce").fillna(0) > 0)].copy()
        if zero.empty:
            st.success("No hay productos con stock y cero ventas en el corte actual.")
        else:
            rotation_hist = build_rotation_history(current_db_version, sales_mode)
            if not rotation_hist.empty:
                zero = zero.merge(rotation_hist, on=["Sede", "Código"], how="left")
                ref_date = pd.to_datetime(latest_display, errors="coerce")
                base_date = pd.to_datetime(zero["Última venta positiva"].fillna(zero["Primera fecha"]), errors="coerce")
                row_ref = pd.to_datetime(zero["Última fecha"], errors="coerce").fillna(ref_date)
                zero["Días sin venta detectada"] = (row_ref - base_date).dt.days.clip(lower=0)

            f1, f2, f3 = st.columns(3)
            with f1:
                site_filter = st.multiselect("Sedes", sorted(zero["Sede"].unique()), default=sorted(zero["Sede"].unique()), key="zero_sites")
            supplier_opts = sorted([x for x in zero["Proveedor"].dropna().astype(str).unique() if x.strip()], key=normalize_key) if "Proveedor" in zero.columns else []
            with f2:
                supplier_filter = st.multiselect("Proveedores", supplier_opts, default=supplier_opts, key="zero_suppliers") if supplier_opts else []
            with f3:
                zero_search = st.text_input("Buscar producto", placeholder="Código o descripción", key="zero_search")

            data = zero[zero["Sede"].isin(site_filter)].copy()
            if supplier_opts:
                data = data[data["Proveedor"].astype(str).isin(supplier_filter)]
            if zero_search:
                q = normalize_text(zero_search)
                data = data[
                    data["Código"].astype(str).str.lower().str.contains(q, na=False) |
                    data["Descripción"].astype(str).str.lower().str.contains(q, na=False)
                ]

            value = float(data["Valor Inventario ($)"].sum()) if not data.empty else 0.0
            units = int(pd.to_numeric(data["Existencia"], errors="coerce").fillna(0).sum()) if not data.empty else 0
            inv_total = float(all_df["Valor Inventario ($)"].sum()) if not all_df.empty else 0.0
            pct = (value / inv_total * 100) if inv_total else 0.0
            longest = int(pd.to_numeric(data.get("Días sin venta detectada", pd.Series(dtype=float)), errors="coerce").max()) if "Días sin venta detectada" in data.columns and data["Días sin venta detectada"].notna().any() else 0
            k1,k2,k3,k4 = st.columns(4)
            with k1: kpi("SKUs sin rotación", integer(len(data)), "Stock > 0 · ventas = 0", "red")
            with k2: kpi("Unidades detenidas", integer(units), "Existencia física", "orange")
            with k3: kpi("Capital detenido", money(value), f"{pct:.1f}% del inventario", "red")
            with k4: kpi("Mayor inactividad", f"{longest} días" if longest else "—", "Según historial cargado", "blue")

            z1, z2 = st.columns(2)
            with z1:
                if "Proveedor" in data.columns:
                    by_sup = data.groupby("Proveedor", dropna=False)["Valor Inventario ($)"].sum().reset_index().sort_values("Valor Inventario ($)", ascending=False).head(15)
                    fig = px.bar(by_sup.sort_values("Valor Inventario ($)"), x="Valor Inventario ($)", y="Proveedor", orientation="h", title="Capital sin rotación por proveedor")
                    st.plotly_chart(chart_layout(fig, 390), use_container_width=True)
            with z2:
                by_site = data.groupby("Sede")["Valor Inventario ($)"].sum().reset_index()
                fig = px.pie(by_site, names="Sede", values="Valor Inventario ($)", hole=.62, title="Distribución del capital detenido")
                st.plotly_chart(chart_layout(fig, 390), use_container_width=True)

            if not data.empty:
                tree = data.copy()
                tree["Proveedor"] = tree.get("Proveedor", pd.Series("Sin proveedor", index=tree.index)).astype(str).replace({"nan": "Sin proveedor", "": "Sin proveedor"})
                tree["Producto"] = tree["Descripción"].where(tree["Descripción"].astype(str).str.strip() != "", tree["Código"].astype(str))
                tree = tree[tree["Valor Inventario ($)"] > 0].copy()
                if not tree.empty:
                    fig = px.treemap(tree.head(500), path=["Sede", "Proveedor", "Producto"], values="Valor Inventario ($)", title="Mapa de capital sin rotación")
                    st.plotly_chart(chart_layout(fig, 480), use_container_width=True)

            cols = ["Sede","Proveedor","Código","Descripción","ABC","Existencia","Costo","Valor Inventario ($)","Última venta positiva","Días sin venta detectada","Acción"]
            cols = [c for c in cols if c in data.columns]
            st.dataframe(data.sort_values("Valor Inventario ($)", ascending=False)[cols].head(1000), use_container_width=True, hide_index=True, height=620)
            xls = io.BytesIO()
            data[cols].to_excel(xls, index=False, sheet_name="Sin Rotacion")
            st.download_button("⬇ Exportar productos sin rotación", xls.getvalue(), f"NEXUS_Sin_Rotacion_{latest_display}.xlsx", DOC_MIME_XLSX)

elif view == "Redistribución":
    st.markdown('<div class="section-title">Motor de redistribución entre sedes</div>', unsafe_allow_html=True)
    if transfers.empty:
        empty_state("No hay transferencias sugeridas", "No se detectaron al mismo tiempo excedentes y déficits compatibles.", "♻️")
    else:
        avoided = float(transfers["Compra Evitada Estimada ($)"].sum())
        units = int(transfers["Unidades Sugeridas"].sum())
        c1, c2, c3 = st.columns(3)
        with c1: kpi("Transferencias", integer(len(transfers)), "Movimientos sugeridos", "green")
        with c2: kpi("Unidades", integer(units), "Redistribución total", "blue")
        with c3: kpi("Compra evitada", money(avoided), "Ahorro estimado", "orange")
        st.dataframe(
            transfers.style.format({"Costo Unitario ($)": "${:,.2f}", "Compra Evitada Estimada ($)": "${:,.2f}"}),
            use_container_width=True,
            hide_index=True,
            height=600,
        )
        route = transfers.groupby(["Origen", "Destino"])["Unidades Sugeridas"].sum().reset_index()
        if not route.empty:
            st.markdown('<div class="section-title">Flujo entre sedes</div>', unsafe_allow_html=True)
            nodes = sorted(set(route["Origen"]) | set(route["Destino"]))
            node_index = {n: i for i, n in enumerate(nodes)}
            palette = ["#3568f5", "#17aabe", "#14a36a", "#ef982d", "#df5555", "#244bd7", "#7b8799"]
            node_colors = [palette[i % len(palette)] for i in range(len(nodes))]
            link_colors = [node_colors[node_index[o]] + "55" for o in route["Origen"]]
            sankey = go.Figure(go.Sankey(
                arrangement="snap",
                node=dict(
                    label=[f"{SEDE_ICON.get(n,'◆')} {n}" for n in nodes],
                    color=node_colors,
                    pad=18,
                    thickness=16,
                    line=dict(color="rgba(0,0,0,0)", width=0),
                ),
                link=dict(
                    source=[node_index[o] for o in route["Origen"]],
                    target=[node_index[d] for d in route["Destino"]],
                    value=route["Unidades Sugeridas"],
                    color=link_colors,
                    label=[f"{int(v):,} uds" for v in route["Unidades Sugeridas"]],
                ),
            ))
            st.plotly_chart(chart_layout(sankey, 380), use_container_width=True)
        wb = io.BytesIO()
        transfers.to_excel(wb, index=False, sheet_name="Redistribucion")
        st.download_button("⬇ Exportar plan de redistribución", wb.getvalue(), "Plan_Redistribucion_Makropetrol.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

elif view == "Finanzas":
    st.markdown('<div class="section-title">Control financiero del inventario</div>', unsafe_allow_html=True)
    if all_df.empty:
        empty_state("Sin datos financieros", "Procesa cortes para calcular inventario, compras y capital inmovilizado.", "💰")
    else:
        inv_value = float(all_df["Valor Inventario ($)"].sum())
        immobilized = float(all_df["Capital Inmovilizado ($)"].sum())
        purchase_col = "Compra Ajustada" if "Compra Ajustada" in all_df.columns else "Compra Sugerida"
        purchase_value = float((all_df[purchase_col] * all_df["Costo"]).sum())
        no_cost = int((all_df["Costo"] <= 0).sum())
        c1, c2, c3, c4 = st.columns(4)
        with c1: kpi("Valor inventario", money(inv_value), "Existencia × costo", "blue")
        with c2: kpi("Capital inmovilizado", money(immobilized), "Sobrestock", "red")
        with c3: kpi("Compra estimada", money(purchase_value), "Compra neta", "orange")
        with c4: kpi("Sin costo", integer(no_cost), "Revisar fuente", "green")
        a, b = st.columns(2)
        by_site = all_df.groupby("Sede")["Valor Inventario ($)"].sum().reset_index()
        with a:
            fig = px.bar(by_site, x="Sede", y="Valor Inventario ($)", title="Valor de inventario por sede", color="Sede", color_discrete_sequence=NEXUS_PALETTE)
            st.plotly_chart(chart_layout(fig), use_container_width=True)
        with b:
            by_site2 = all_df.groupby("Sede")["Capital Inmovilizado ($)"].sum().reset_index()
            fig = px.bar(by_site2, x="Sede", y="Capital Inmovilizado ($)", title="Capital inmovilizado por sede", color="Sede", color_discrete_sequence=NEXUS_PALETTE)
            st.plotly_chart(chart_layout(fig), use_container_width=True)
        st.markdown('<div class="section-title">Capital inmovilizado por producto</div>', unsafe_allow_html=True)
        financial = all_df[all_df["Capital Inmovilizado ($)"] > 0].sort_values("Capital Inmovilizado ($)", ascending=False).head(100)
        fin_cols = ["Sede", "Código", "Descripción", "Costo", "Existencia", "Retiro Almacén", "Capital Inmovilizado ($)"] if not modo_gerencia else ["Sede", "Descripción", "Retiro Almacén", "Capital Inmovilizado ($)"]
        max_cap = float(financial["Capital Inmovilizado ($)"].max()) if not financial.empty else 1.0
        st.dataframe(
            financial[fin_cols],
            use_container_width=True,
            hide_index=True,
            height=540,
            column_config={
                "Costo": st.column_config.NumberColumn(format="$%.2f"),
                "Capital Inmovilizado ($)": st.column_config.ProgressColumn(
                    "Capital inmovilizado", min_value=0, max_value=max(1.0, max_cap), format="$%.0f"
                ),
            },
        )

elif view == "Alertas":
    st.markdown('<div class="section-title">Centro de alertas</div>', unsafe_allow_html=True)
    render_alertas_fragment(alerts)

elif view == "Automatización":
    st.markdown('<div class="section-title">Hub de automatización</div>', unsafe_allow_html=True)
    render_automation_fragment(site_results_adjusted, transfers, all_df, contacts)

elif view == "Historial":
    st.markdown('<div class="section-title">Historial de cortes y trazabilidad</div>', unsafe_allow_html=True)
    with db() as con:
        log = pd.read_sql_query(
            "SELECT processed_at, snapshot_date, site, inventory_rows, sales_rows FROM processing_log ORDER BY processed_at DESC LIMIT 300",
            con,
        )
    if log.empty:
        empty_state("Sin historial", "Los cortes que guardes aparecerán aquí.", "◷")
    else:
        st.dataframe(log, use_container_width=True, hide_index=True, height=620)

elif view == "Exportación de Datos":
    st.markdown('<div class="section-title">Centro de documentos y exportación</div>', unsafe_allow_html=True)
    st.caption("Genera documentos listos para operación y administración. Los cálculos del dashboard ya están cacheados para que esta vista no ralentice el resto de NEXUS.")

    if all_df.empty:
        empty_state("Nada que documentar", "Carga y procesa al menos un corte diario.", "📭")
    else:
        f1, f2, f3 = st.columns(3)
        with f1:
            doc_type_label = st.selectbox(
                "Tipo de documento",
                [
                    "🛒 Pedido de compra",
                    "📤 Retiro de almacén",
                    "🚨 Alertas por sede",
                    "♻️ Redistribución multisede",
                ],
                key="doc_type_selector",
            )
        with f2:
            if "Redistribución" in doc_type_label:
                selected_doc_site = "GLOBAL / TODAS LAS SEDES"
            else:
                doc_sites = sorted(all_df["Sede"].unique())
                selected_doc_site = st.selectbox(
                    "Sede",
                    doc_sites,
                    key="doc_site_selector",
                )
        with f3:
            output_format = st.selectbox(
                "Formato",
                ["PDF", "HTML imprimible", "Excel", "CSV"],
                key="doc_format_selector",
            )

        if "Pedido de compra" in doc_type_label:
            doc_kind = "purchase"
            doc_title = "ORDEN DE COMPRA"
            doc_df = all_df if selected_doc_site == "GLOBAL / TODAS LAS SEDES" else all_df[all_df["Sede"] == selected_doc_site]
            prepared = _doc_prepare_frame(doc_df, "purchase", selected_doc_site)
        elif "Retiro" in doc_type_label:
            doc_kind = "withdrawal"
            doc_title = "RETIRO DE ALMACÉN"
            doc_df = all_df if selected_doc_site == "GLOBAL / TODAS LAS SEDES" else all_df[all_df["Sede"] == selected_doc_site]
            prepared = _doc_prepare_frame(doc_df, "withdrawal", selected_doc_site)
        elif "Alertas" in doc_type_label:
            doc_kind = "alerts"
            doc_title = "REPORTE DE ALERTAS OPERATIVAS"
            doc_df = alerts if selected_doc_site == "GLOBAL / TODAS LAS SEDES" else alerts[alerts["Sede"] == selected_doc_site]
            prepared = _doc_prepare_frame(doc_df, "alerts", selected_doc_site)
        else:
            doc_kind = "redistribution"
            doc_title = "PLAN DE REDISTRIBUCIÓN"
            doc_df = transfers
            prepared = _doc_prepare_frame(doc_df, "redistribution")

        if doc_kind in {"purchase", "withdrawal"}:
            units = int(pd.to_numeric(prepared.get("Cantidad", 0), errors="coerce").fillna(0).sum()) if not prepared.empty else 0
            value = float(pd.to_numeric(prepared.get("Total", 0), errors="coerce").fillna(0).sum()) if not prepared.empty else 0.0
            k1, k2, k3, k4 = st.columns(4)
            with k1: kpi("Items", integer(len(prepared)), "Documento", "blue")
            with k2: kpi("Unidades", integer(units), "Cantidad física", "cyan")
            with k3: kpi("Valor", money(value), "Estimado", "orange")
            with k4: kpi("Ámbito", "Por sede", selected_doc_site, "green")
        elif doc_kind == "alerts":
            crit = int((prepared["Severidad"].astype(str) == "CRÍTICA").sum()) if not prepared.empty else 0
            high = int((prepared["Severidad"].astype(str) == "ALTA").sum()) if not prepared.empty else 0
            k1, k2, k3, k4 = st.columns(4)
            with k1: kpi("Alertas", integer(len(prepared)), "Registros", "red")
            with k2: kpi("Críticas", integer(crit), "Atención inmediata", "red")
            with k3: kpi("Altas", integer(high), "Prioridad", "orange")
            with k4: kpi("Ámbito", "Por sede", selected_doc_site, "blue")
        else:
            units = int(pd.to_numeric(prepared.get("Unidades Sugeridas", 0), errors="coerce").fillna(0).sum()) if not prepared.empty else 0
            avoided = float(pd.to_numeric(prepared.get("Compra Evitada Estimada ($)", 0), errors="coerce").fillna(0).sum()) if not prepared.empty else 0.0
            routes = int(len(prepared)) if not prepared.empty else 0
            k1, k2, k3, k4 = st.columns(4)
            with k1: kpi("Movimientos", integer(routes), "Cruces origen → destino", "green")
            with k2: kpi("Unidades", integer(units), "Redistribución", "blue")
            with k3: kpi("Compra evitada", money(avoided), "Estimación", "orange")
            with k4: kpi("Ámbito", "Multisede", "Cruce global", "cyan")

        st.markdown('<div class="section-title">Vista previa del documento</div>', unsafe_allow_html=True)
        preview_cols = prepared.head(60)
        if preview_cols.empty:
            empty_state("Sin registros para este documento", "No hay productos que cumplan las reglas actuales.", "✓")
        else:
            st.dataframe(
                preview_cols,
                use_container_width=True,
                hide_index=True,
                height=430,
            )
            if len(prepared) > 60:
                st.caption(f"Vista previa limitada a 60 filas. El archivo conserva las {len(prepared):,} filas completas.")

        st.markdown('<div class="section-title">Descargar documento</div>', unsafe_allow_html=True)
        site_slug = safe_slug(selected_doc_site)
        doc_slug = safe_slug(doc_title)

        if output_format == "PDF":
            if doc_kind == "purchase":
                file_bytes = build_purchase_pdf(doc_df, selected_doc_site, latest_display, sequence=1)
            elif doc_kind == "withdrawal":
                file_bytes = build_withdrawal_pdf(doc_df, selected_doc_site, latest_display)
            elif doc_kind == "alerts":
                file_bytes = build_alerts_pdf(alerts, selected_doc_site, latest_display)
            else:
                file_bytes = build_redistribution_pdf(transfers, latest_display)
            if file_bytes:
                st.download_button(
                    "📄 Descargar PDF", file_bytes,
                    f"NEXUS_{doc_slug}_{site_slug}_{latest_display}.pdf",
                    DOC_MIME_PDF, use_container_width=True,
                )
            else:
                st.info("Instala fpdf2 para habilitar PDF.")

        elif output_format == "HTML imprimible":
            html_site = None if doc_kind == "redistribution" else selected_doc_site
            file_bytes = build_operational_html(doc_kind, doc_df, html_site, latest_display, f"{doc_title} · {selected_doc_site}")
            st.download_button(
                "🌐 Descargar HTML imprimible", file_bytes,
                f"NEXUS_{doc_slug}_{site_slug}_{latest_display}.html",
                DOC_MIME_HTML, use_container_width=True,
            )

        elif output_format == "Excel":
            if doc_kind == "purchase":
                file_bytes = build_purchase_excel(doc_df, selected_doc_site, latest_display)
            else:
                file_bytes = export_xlsx_bytes(prepared, doc_title)
            st.download_button(
                "📊 Descargar Excel", file_bytes,
                f"NEXUS_{doc_slug}_{site_slug}_{latest_display}.xlsx",
                DOC_MIME_XLSX, use_container_width=True,
            )

        else:
            file_bytes = export_csv_bytes(prepared)
            st.download_button(
                "🧾 Descargar CSV", file_bytes,
                f"NEXUS_{doc_slug}_{site_slug}_{latest_display}.csv",
                "text/csv", use_container_width=True,
            )

        st.markdown('<div class="section-title">Paquete ejecutivo completo</div>', unsafe_allow_html=True)
        st.caption("Se genera solo cuando lo solicitas. Así el Centro de documentos también permanece rápido.")
        a,b,c = st.columns(3)
        with a: kpi("Formatos", "7", "Excel×2 · HTML · PDF · CSV · JSON · ZIP", "blue")
        with b: kpi("Datos", integer(len(all_df)), "Registros consolidados", "green")
        with c: kpi("Interactividad", "Alta", "HTML Plotly optimizado", "cyan")

        package_token = safe_slug(
            f"{latest_display}_{score}_{months_history}_{min_coverage}_{max_coverage}_{lead_time_days}_{current_db_version}"
        )
        package_key = f"nexus_package_{package_token}"
        if st.button("⚙️ Preparar paquete ejecutivo completo", type="primary", use_container_width=True, key=f"prepare_package_{package_token}"):
            with st.spinner("Generando reportes ejecutivos..."):
                excel_bytes = export_excel(
                    site_results_adjusted, transfers, int(months_history), float(min_coverage),
                    float(max_coverage), int(lead_time_days), latest_display,
                )
                gerencial_bytes = build_executive_excel(all_df, latest_display, score)
                html_bytes = build_interactive_html_report(all_df, transfers, alerts, latest_display, score)
                pdf_bytes = build_executive_pdf(all_df, transfers, latest_display, score)
                bundle_bytes = export_bundle(excel_bytes, all_df, transfers, alerts, latest_display, score)
                st.session_state[package_key] = {
                    "operativo": excel_bytes,
                    "gerencial": gerencial_bytes,
                    "html": html_bytes,
                    "pdf": pdf_bytes,
                    "zip": bundle_bytes,
                }
            st.toast("Paquete ejecutivo preparado", icon="📦")

        package = st.session_state.get(package_key)
        if package:
            r1,r2 = st.columns(2)
            with r1:
                st.download_button("⬇ Excel operativo", package["operativo"], f"NEXUS_Reporte_Operativo_{latest_display}.xlsx", DOC_MIME_XLSX, use_container_width=True)
            with r2:
                st.download_button("⬇ Excel gerencial", package["gerencial"], f"NEXUS_Reporte_Gerencial_{latest_display}.xlsx", DOC_MIME_XLSX, use_container_width=True)
            r3,r4 = st.columns(2)
            with r3:
                st.download_button("⬇ Dashboard HTML", package["html"], f"NEXUS_Dashboard_{latest_display}.html", DOC_MIME_HTML, use_container_width=True)
            with r4:
                if package.get("pdf"):
                    st.download_button("⬇ PDF ejecutivo", package["pdf"], f"NEXUS_Ejecutivo_{latest_display}.pdf", DOC_MIME_PDF, use_container_width=True)
                else:
                    st.info("Instala fpdf2 para habilitar PDF.")
            d1,d2,d3 = st.columns(3)
            with d1: st.download_button("⬇ CSV consolidado", export_csv_bytes(all_df), f"NEXUS_Datos_{latest_display}.csv", "text/csv", use_container_width=True)
            with d2: st.download_button("⬇ JSON estructurado", export_json_bytes(all_df), f"NEXUS_Datos_{latest_display}.json", "application/json", use_container_width=True)
            with d3: st.download_button("📦 ZIP completo", package["zip"], f"NEXUS_Paquete_{latest_display}.zip", "application/zip", use_container_width=True)

if not all_df.empty and view != "Exportación de Datos":
    st.markdown('<div class="section-title">Reportes</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="panel"><b>Centro de documentos</b><br><span style="color:#7b8799;font-size:.82rem">Los archivos pesados se generan únicamente cuando entras al Centro de documentos, evitando recalcular Excel/PDF en cada clic del dashboard.</span></div>',
        unsafe_allow_html=True,
    )
    if st.button("📄 Ir a Exportación de Datos", key="go_exports", use_container_width=False):
        st.session_state["nexus_view"] = "Exportación de Datos"
        st.rerun()


st.markdown(
    f"<div class='footer'>Makropetrol NEXUS v{APP_VERSION} · SQLite + caché inteligente · análisis por proveedor · rotación · automatización</div>",
    unsafe_allow_html=True,
)
