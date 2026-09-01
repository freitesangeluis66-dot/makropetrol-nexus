# ============================================================
# MAKROPETROL NEXUS v4.2 — Executive Intelligence
# Plataforma de inteligencia logística multisede
#
# Incluye:
# - Recalculo dinámico en tiempo real desde datos base
# - Persistencia local diaria con SQLite (capa de acceso corta, sin
#   conexión global compartida, pensada para concurrencia futura)
# - Inventario + ventas por corte, con lectura Excel acelerada (calamine)
# - Ventas acumuladas / por período / diarias
# - Compras, retiros y redistribución automática entre sedes (con Sankey)
# - ABC / Pareto, rotación y cobertura
# - Punto de reorden + lead time + stock de seguridad
# - Alertas y Health Score
# - Historial de cortes
# - Modo Gerencia / Modo Operativo
# - Exportación a 6 formatos: Excel operativo, Excel gerencial, HTML,
#   PDF ejecutivo, CSV, JSON + paquete ZIP completo
# - Hub de automatización mediante webhook (CRM / n8n / Make / WhatsApp provider)
#
# Requisitos: streamlit, pandas, numpy, openpyxl, plotly, requests
# Recomendado (mejora de rendimiento de lectura Excel): python-calamine
# Opcional (PDF ejecutivo): fpdf2
# Ejecutar: streamlit run makropetrol_nexus_v4_2.py
# ============================================================

from __future__ import annotations

import io
import json
import re
import sqlite3
import zipfile

from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import requests
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio

# Motor de lectura Excel acelerado. Si python-calamine no está instalado,
# NEXUS sigue funcionando normalmente con el motor por defecto de pandas.
try:
    import python_calamine  # noqa: F401
    EXCEL_ENGINE = "calamine"
except Exception:
    EXCEL_ENGINE = None

# Generación de PDF ejecutivo (opcional). Si fpdf2 no está instalado, el
# botón de PDF se oculta automáticamente y todo lo demás sigue igual.
try:
    from fpdf import FPDF
    HAS_PDF = True
except Exception:
    HAS_PDF = False


# ============================================================
# CONFIGURACIÓN
# ============================================================

APP_VERSION = "4.2.0"
DB_PATH = Path(__file__).with_name("nexus_data.db")
CONTACTS_PATH = Path(__file__).with_name("nexus_contacts.json")

SEDES = [
    "Centurión",
    "Perrokeros",
    "Puerto La Cruz",
    "Barcelona",
    "Matriz Makropetrol",
]

SEDE_ICON = {
    "Centurión": "◉",
    "Perrokeros": "◈",
    "Puerto La Cruz": "◇",
    "Barcelona": "▣",
    "Matriz Makropetrol": "◆",
}

SALES_MODES = {
    "Ventas del período": "periodo",
    "Ventas acumuladas": "acumuladas",
    "Ventas diarias": "diarias",
}

DEFAULT_CONTACTS = {
    site: {
        "supervisor": "",
        "whatsapp": "",
        "webhook": "",
    }
    for site in SEDES
}


# ============================================================
# ESTILO
# ============================================================

st.set_page_config(
    page_title="Makropetrol NEXUS | Intelligence",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=Manrope:wght@500;600;700;800&display=swap');
:root{
 --bg:#f4f7fb; --surface:rgba(255,255,255,.88); --line:#e4e9f2;
 --text:#172137; --muted:#7b8799; --blue:#3568f5; --blue2:#244bd7;
 --green:#14a36a; --orange:#ef982d; --red:#df5555; --cyan:#17aabe;
 --shadow:0 16px 40px rgba(34,52,82,.08);
}
html,body,[class*="css"]{font-family:'DM Sans',sans-serif}
.stApp{background:radial-gradient(circle at 8% 0%,rgba(53,104,245,.08),transparent 28%),radial-gradient(circle at 100% 0%,rgba(23,170,190,.06),transparent 26%),linear-gradient(140deg,#f8fafc,#edf3f8);color:var(--text)}
.block-container{max-width:1530px;padding-top:1.5rem;padding-bottom:3rem}
section[data-testid="stSidebar"]{background:rgba(249,251,254,.96);border-right:1px solid var(--line)}
h1,h2,h3,h4{font-family:'Manrope',sans-serif!important;color:var(--text)!important;letter-spacing:-.03em}
.hero{background:linear-gradient(135deg,rgba(255,255,255,.94),rgba(255,255,255,.68));border:1px solid rgba(255,255,255,.9);box-shadow:var(--shadow);border-radius:26px;padding:25px 29px 22px;position:relative;overflow:hidden}
.hero:after{content:"";position:absolute;width:320px;height:320px;right:-130px;top:-170px;border-radius:50%;background:radial-gradient(circle,rgba(53,104,245,.18),transparent 68%)}
.brand{font-family:'Manrope',sans-serif;font-size:1.9rem;font-weight:800;letter-spacing:-.045em}.brand span{color:var(--blue)}
.hero-sub{color:var(--muted);margin-top:4px;font-size:.92rem}.status-row{display:flex;gap:9px;align-items:center;margin-top:14px;color:#667287;font-size:.77rem}.dot{width:8px;height:8px;border-radius:50%;background:#19bd77;box-shadow:0 0 0 5px rgba(25,189,119,.11)}
.section-title{font-family:'Manrope',sans-serif;font-size:1.07rem;font-weight:800;margin:23px 0 9px}
.kpi-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-top:17px}.kpi{background:var(--surface);border:1px solid rgba(255,255,255,.95);box-shadow:var(--shadow);border-radius:20px;padding:17px 19px;min-height:120px;position:relative;overflow:hidden}.kpi:before{content:"";position:absolute;left:0;top:0;bottom:0;width:4px;background:var(--blue)}.kpi.green:before{background:var(--green)}.kpi.orange:before{background:var(--orange)}.kpi.red:before{background:var(--red)}.kpi.cyan:before{background:var(--cyan)}.kpi-label{font-size:.74rem;color:var(--muted);font-weight:800;letter-spacing:.04em;text-transform:uppercase}.kpi-value{font-family:'Manrope',sans-serif;font-size:1.7rem;font-weight:800;letter-spacing:-.04em;margin-top:8px}.kpi-note{font-size:.71rem;color:#9aa5b5;margin-top:3px}
.panel{background:var(--surface);border:1px solid rgba(255,255,255,.96);box-shadow:var(--shadow);border-radius:21px;padding:18px}
.badge{display:inline-block;border-radius:999px;padding:5px 10px;font-size:.7rem;font-weight:800}.badge-blue{background:#edf2ff;color:#355bd7}.badge-green{background:#eafaf3;color:#168154}.badge-orange{background:#fff3df;color:#ba6b10}.badge-red{background:#ffefef;color:#c33b3b}.badge-gray{background:#f0f2f6;color:#68758a}
.stButton>button,.stDownloadButton>button{border-radius:12px;font-weight:750;min-height:42px}.stDownloadButton>button{border:0;background:linear-gradient(135deg,var(--blue),var(--blue2));color:#fff;box-shadow:0 8px 20px rgba(53,104,245,.2)}
[data-testid="stFileUploaderDropzone"]{border:1.5px dashed #cbd6e6!important;border-radius:15px!important;background:rgba(255,255,255,.72)!important}
.stTabs [data-baseweb="tab-list"]{gap:6px;background:rgba(255,255,255,.72);border:1px solid var(--line);padding:6px;border-radius:15px}.stTabs [data-baseweb="tab"]{border-radius:11px;color:#7d899b;font-weight:700;padding:9px 15px}.stTabs [aria-selected="true"]{background:#edf2ff!important;color:var(--blue)!important}
div[data-testid="stDataFrame"]{border:1px solid var(--line);border-radius:15px;overflow:hidden}
.footer{text-align:center;color:#99a4b5;font-size:.7rem;margin-top:34px}
.empty{background:rgba(255,255,255,.72);border:1px dashed #ccd6e5;border-radius:20px;padding:34px;text-align:center;color:#758195}.empty-title{font-family:'Manrope';font-weight:800;font-size:1.05rem;color:#273550}.empty-icon{font-size:2rem;margin-bottom:7px}
.nav-group{font-size:.68rem;font-weight:800;letter-spacing:.08em;text-transform:uppercase;color:#93a0b4;margin:14px 0 2px 2px}
.gov-banner{display:flex;align-items:center;gap:8px;background:linear-gradient(135deg,#111a2e,#1c2b4a);color:#fff;border-radius:14px;padding:9px 14px;font-size:.78rem;font-weight:700;margin-bottom:10px}
.health-big{display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;padding:18px 10px}
.health-big .num{font-family:'Manrope',sans-serif;font-size:3rem;font-weight:800;letter-spacing:-.05em;color:var(--blue)}
.health-big .lbl{font-size:.74rem;color:var(--muted);font-weight:800;letter-spacing:.06em;text-transform:uppercase;margin-top:2px}
.op-row{display:flex;justify-content:space-between;align-items:center;padding:8px 4px;border-bottom:1px solid var(--line);font-size:.86rem}
.op-row:last-child{border-bottom:none}
.bar-track{background:#eef1f6;border-radius:8px;height:10px;overflow:hidden;flex:1;margin:0 10px}
.bar-fill{height:100%;border-radius:8px;background:linear-gradient(90deg,var(--blue),var(--cyan))}
.opp-banner{display:flex;gap:22px;flex-wrap:wrap;background:linear-gradient(135deg,#eafaf3,#edf7ff);border:1px solid #d9ecec;border-radius:16px;padding:14px 18px;font-size:.86rem;font-weight:700;color:#245a4b}
@media(max-width:1000px){.kpi-grid{grid-template-columns:repeat(2,1fr)}}@media(max-width:650px){.kpi-grid{grid-template-columns:1fr}.hero{padding:20px}.opp-banner{flex-direction:column;gap:8px}}
</style>
""",
    unsafe_allow_html=True,
)


# ============================================================
# UTILIDADES
# ============================================================

def money(v: Any) -> str:
    try:
        return f"${float(v):,.2f}"
    except Exception:
        return "$0.00"


def integer(v: Any) -> str:
    try:
        return f"{int(round(float(v))):,}"
    except Exception:
        return "0"


def safe_float(v: Any, default: float = 0.0) -> float:
    try:
        x = float(v)
        return x if np.isfinite(x) else default
    except Exception:
        return default


def normalize_text(v: Any) -> str:
    x = str(v).strip().lower()
    x = re.sub(r"\s+", " ", x)
    return x


def parse_number(v: Any) -> float:
    """Admite formatos 1.234,56 / 1,234.56 / 1234,56 / 1234.56."""
    if pd.isna(v):
        return 0.0
    if isinstance(v, (int, float, np.integer, np.floating)):
        return float(v)
    s = str(v).strip().replace("$", "").replace("€", "").replace(" ", "")
    if not s:
        return 0.0
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        parts = s.split(",")
        if len(parts[-1]) in (1, 2):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        # Punto decimal estándar. No tocarlo.
        pass
    try:
        return float(s)
    except Exception:
        return 0.0


def empty_state(title: str, message: str, icon: str = "◌") -> None:
    st.markdown(
        f"""
        <div class="empty">
            <div class="empty-icon">{icon}</div>
            <div class="empty-title">{title}</div>
            <div>{message}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def kpi(label: str, value: str, note: str = "", tone: str = "blue") -> None:
    st.markdown(
        f"""
        <div class="kpi {tone}">
            <div class="kpi-label">{label}</div>
            <div class="kpi-value">{value}</div>
            <div class="kpi-note">{note}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def chart_layout(fig: go.Figure, height: int = 330) -> go.Figure:
    fig.update_layout(
        height=height,
        margin=dict(l=10, r=10, t=42, b=10),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="DM Sans", color="#59677b"),
        legend=dict(orientation="h", y=1.03, x=0),
    )
    return fig


def present_columns(df: pd.DataFrame, preferred: list[str]) -> list[str]:
    """Devuelve columnas existentes en el orden indicado."""
    return [c for c in preferred if c in df.columns]


def build_executive_frame(all_df: pd.DataFrame) -> pd.DataFrame:
    """Versión legible del dataset para usuarios gerenciales."""
    if all_df.empty:
        return pd.DataFrame()
    preferred = [
        "Sede", "Código", "Descripción", "ABC", "Estado", "Prioridad", "Acción",
        "Ventas", "Demanda Mensual", "Existencia", "Stock Mínimo", "Stock Máximo",
        "Punto de Reorden", "Compra Ajustada", "Retiro Almacén", "Cobertura (meses)",
        "Valor Inventario ($)", "Capital Inmovilizado ($)", "Costo",
    ]
    return all_df[present_columns(all_df, preferred)].copy()


def export_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")


def export_json_bytes(df: pd.DataFrame) -> bytes:
    return df.to_json(orient="records", force_ascii=False, indent=2).encode("utf-8")


def build_interactive_html_report(
    all_df: pd.DataFrame,
    transfers: pd.DataFrame,
    alerts: pd.DataFrame,
    snapshot_label: str,
    score: int,
) -> bytes:
    """Genera un reporte HTML autosuficiente con gráficos Plotly y tablas resumidas."""
    if all_df.empty:
        body = "<h2>Sin datos disponibles</h2><p>Procese al menos un corte diario.</p>"
        return f"<!doctype html><html lang='es'><meta charset='utf-8'><body>{body}</body></html>".encode("utf-8")

    purchase_col = "Compra Ajustada" if "Compra Ajustada" in all_df.columns else "Compra Sugerida"
    site_summary = all_df.groupby("Sede").agg(
        Inventario=("Valor Inventario ($)", "sum"),
        Críticos=("Estado", lambda x: int(x.astype(str).str.startswith("CRÍTICO").sum())),
        Compras=(purchase_col, lambda x: int((x > 0).sum())),
        Retiros=("Retiro Almacén", lambda x: int((x > 0).sum())),
    ).reset_index()

    fig1 = px.bar(site_summary, x="Sede", y="Inventario", title="Valor de inventario por sede")
    fig2 = px.bar(site_summary, x="Sede", y=["Compras", "Retiros"], barmode="group", title="Acciones por sede")
    state = all_df["Estado"].value_counts().reset_index()
    state.columns = ["Estado", "Cantidad"]
    fig3 = px.pie(state, names="Estado", values="Cantidad", hole=.62, title="Distribución del estado logístico")

    top = all_df[all_df["Capital Inmovilizado ($)"] > 0].sort_values("Capital Inmovilizado ($)", ascending=False).head(12)
    fig4 = px.bar(top.sort_values("Capital Inmovilizado ($)"), x="Capital Inmovilizado ($)", y="Descripción", orientation="h", title="Top capital inmovilizado") if not top.empty else None

    def fig_html(fig):
        return pio.to_html(fig, full_html=False, include_plotlyjs=False)

    exec_df = build_executive_frame(all_df).head(40)
    alert_html = alerts.head(30).to_html(index=False, classes="dataframe") if not alerts.empty else "<p>Sin alertas.</p>"
    transfer_html = transfers.head(30).to_html(index=False, classes="dataframe") if not transfers.empty else "<p>Sin transferencias sugeridas.</p>"
    table_html = exec_df.to_html(index=False, classes="dataframe")

    html = f"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Makropetrol NEXUS · Reporte Ejecutivo</title>
<style>
body{{margin:0;background:#f4f7fb;color:#172137;font-family:Inter,Segoe UI,Arial,sans-serif}}
.wrap{{max-width:1400px;margin:auto;padding:38px}}
.hero{{background:linear-gradient(135deg,#fff,#f0f5ff);border:1px solid #e2e8f2;border-radius:28px;padding:30px;box-shadow:0 18px 50px rgba(30,50,80,.10)}}
.brand{{font-size:28px;font-weight:800}} .brand span{{color:#3568f5}}
.muted{{color:#748198}} .score{{font-size:42px;font-weight:800;color:#3568f5}}
.kpis{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-top:18px}}
.kpi{{background:#fff;border:1px solid #e3e9f1;border-radius:20px;padding:18px;box-shadow:0 8px 24px rgba(30,50,80,.07)}}
.kpi small{{color:#7d899b;text-transform:uppercase;font-weight:800;font-size:11px}} .kpi strong{{display:block;font-size:27px;margin-top:6px}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-top:18px}} .panel{{background:#fff;border:1px solid #e3e9f1;border-radius:20px;padding:18px;box-shadow:0 8px 24px rgba(30,50,80,.06)}}
.dataframe{{width:100%;border-collapse:collapse;font-size:12px}} .dataframe th{{background:#3568f5;color:#fff;padding:9px;text-align:left}} .dataframe td{{padding:8px;border-bottom:1px solid #e5eaf1}} .dataframe tr:nth-child(even){{background:#f9fbfd}}
@media(max-width:900px){{.kpis,.grid{{grid-template-columns:1fr 1fr}}}} @media(max-width:600px){{.kpis,.grid{{grid-template-columns:1fr}} .wrap{{padding:18px}}}}
</style>
<script src="https://cdn.plot.ly/plotly-2.35.2.min.js"></script>
</head>
<body><div class="wrap">
<div class="hero"><div class="brand">MAKROPETROL <span>NEXUS</span></div>
<div class="muted">Reporte ejecutivo interactivo · Corte {snapshot_label}</div>
<div style="margin-top:16px"><span class="muted">Salud logística</span><div class="score">{score}/100</div></div></div>
<div class="kpis">
<div class="kpi"><small>SKUs</small><strong>{len(all_df):,}</strong></div>
<div class="kpi"><small>Críticos</small><strong>{int(all_df['Estado'].astype(str).str.startswith('CRÍTICO').sum()):,}</strong></div>
<div class="kpi"><small>Compra neta</small><strong>{int(all_df[purchase_col].sum()):,}</strong></div>
<div class="kpi"><small>Capital inmovilizado</small><strong>${float(all_df['Capital Inmovilizado ($)'].sum()):,.2f}</strong></div>
</div>
<div class="grid">
<div class="panel">{fig_html(fig1)}</div><div class="panel">{fig_html(fig2)}</div>
<div class="panel">{fig_html(fig3)}</div>{f'<div class="panel">{fig_html(fig4)}</div>' if fig4 is not None else '<div class="panel"><p class="muted">Sin capital inmovilizado detectado.</p></div>'}
</div>
<div class="panel" style="margin-top:18px"><h2>Vista ejecutiva de productos</h2>{table_html}</div>
<div class="panel" style="margin-top:18px"><h2>Alertas</h2>{alert_html}</div>
<div class="panel" style="margin-top:18px"><h2>Redistribución</h2>{transfer_html}</div>
</div></body></html>"""
    return html.encode("utf-8")


def build_executive_excel(all_df: pd.DataFrame, snapshot_label: str, score: int) -> bytes:
    """Excel Gerencial: solo indicadores y acciones, sin columnas técnicas.
    Pensado para enviarse directo a dirección, sin necesidad de filtrar."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    out = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Gerencial"

    navy, blue, white, gray = "18243A", "3568F5", "FFFFFF", "7B8799"
    ws.merge_cells("A1:F1")
    ws["A1"] = "MAKROPETROL NEXUS · RESUMEN GERENCIAL"
    ws["A1"].font = Font(size=18, bold=True, color=white)
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Corte {snapshot_label} · Salud logística {score}/100 · {health_label(score)}"
    ws["A2"].font = Font(size=10, italic=True, color=gray)

    if not all_df.empty:
        purchase_col = "Compra Ajustada" if "Compra Ajustada" in all_df.columns else "Compra Sugerida"
        by_site = all_df.groupby("Sede").agg(
            Inventario=("Valor Inventario ($)", "sum"),
            Críticos=("Estado", lambda x: int(x.astype(str).str.startswith("CRÍTICO").sum())),
            Compras=(purchase_col, lambda x: int((x > 0).sum())),
            Inmovilizado=("Capital Inmovilizado ($)", "sum"),
        ).reset_index()
        start = 4
        for j, col in enumerate(by_site.columns, 1):
            c = ws.cell(start, j, col)
            c.font = Font(bold=True, color=white)
            c.fill = PatternFill("solid", fgColor=blue)
        for i, row in enumerate(by_site.itertuples(index=False), start + 1):
            for j, val in enumerate(row, 1):
                ws.cell(i, j, val)
        ref = f"A{start}:{get_column_letter(len(by_site.columns))}{start + len(by_site)}"
        table = Table(displayName="ResumenGerencial", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
        for j in range(1, len(by_site.columns) + 1):
            ws.column_dimensions[get_column_letter(j)].width = 20

        exec_df = build_executive_frame(all_df).head(500)
        keep_cols = [c for c in ["Sede", "Código", "Descripción", "ABC", "Estado", "Acción", "Cobertura (meses)"] if c in exec_df.columns]
        sh2 = wb.create_sheet("Acciones recomendadas")
        for j, col in enumerate(keep_cols, 1):
            c = sh2.cell(1, j, col)
            c.font = Font(bold=True, color=white)
            c.fill = PatternFill("solid", fgColor=blue)
        for i, row in enumerate(exec_df[keep_cols].itertuples(index=False), 2):
            for j, val in enumerate(row, 1):
                sh2.cell(i, j, val)
        for j in range(1, len(keep_cols) + 1):
            sh2.column_dimensions[get_column_letter(j)].width = 22

    wb.save(out)
    return out.getvalue()


def build_executive_pdf(all_df: pd.DataFrame, transfers: pd.DataFrame, snapshot_label: str, score: int) -> bytes | None:
    """PDF ejecutivo de una página, listo para enviar por WhatsApp o correo.
    Requiere fpdf2 (pip install fpdf2); si no está instalado, retorna None
    y el botón correspondiente simplemente no se muestra."""
    if not HAS_PDF:
        return None
    purchase_col = "Compra Ajustada" if "Compra Ajustada" in all_df.columns else "Compra Sugerida"
    critical = int(all_df["Estado"].astype(str).str.startswith("CRÍTICO").sum()) if not all_df.empty else 0
    purchases = int((all_df[purchase_col] > 0).sum()) if not all_df.empty else 0
    inv_value = float(all_df["Valor Inventario ($)"].sum()) if not all_df.empty else 0.0
    immobilized = float(all_df["Capital Inmovilizado ($)"].sum()) if not all_df.empty else 0.0
    avoided = float(transfers["Compra Evitada Estimada ($)"].sum()) if not transfers.empty else 0.0

    pdf = FPDF()
    pdf.add_page()
    pdf.set_fill_color(24, 36, 58)
    pdf.rect(0, 0, 210, 28, "F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_xy(10, 8)
    pdf.cell(0, 10, "MAKROPETROL NEXUS - Reporte Ejecutivo")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_xy(10, 18)
    pdf.cell(0, 8, f"Corte {snapshot_label}  ·  Salud logistica {score}/100 ({health_label(score)})")

    pdf.set_text_color(30, 30, 30)
    pdf.set_xy(10, 36)
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, "Indicadores clave", ln=True)
    pdf.set_font("Helvetica", "", 11)
    rows = [
        ("Valor de inventario", money(inv_value)),
        ("SKUs criticos", f"{critical:,}"),
        ("SKUs a comprar", f"{purchases:,}"),
        ("Capital inmovilizado", money(immobilized)),
        ("Compra evitada por redistribucion", money(avoided)),
    ]
    for label, val in rows:
        pdf.set_x(12)
        pdf.cell(90, 8, label)
        pdf.cell(0, 8, val, ln=True)

    if not all_df.empty:
        pdf.ln(4)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 8, "Inventario por sede", ln=True)
        pdf.set_font("Helvetica", "", 10)
        by_site = all_df.groupby("Sede")["Valor Inventario ($)"].sum().sort_values(ascending=False)
        for sede, val in by_site.items():
            pdf.set_x(12)
            pdf.cell(90, 7, str(sede))
            pdf.cell(0, 7, money(val), ln=True)

    pdf.ln(4)
    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(120, 120, 120)
    pdf.multi_cell(0, 6, f"Generado automaticamente por NEXUS v{APP_VERSION} el {datetime.now().strftime('%d/%m/%Y %H:%M')}.")

    return bytes(pdf.output())


def build_readme_text(snapshot_label: str) -> bytes:
    text = f"""MAKROPETROL NEXUS · Paquete de reporte
Corte: {snapshot_label}
Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}

Contenido del paquete:
- Dashboard_Gerencial.html  → dashboard interactivo con gráficos Plotly, se abre en cualquier navegador.
- Reporte_Gerencial.xlsx    → resumen ejecutivo por sede + acciones recomendadas.
- Reporte_Operativo.xlsx    → detalle completo (inventario, compras, retiros, redistribución, alertas).
- Datos.csv / Datos.json    → datos consolidados para Power BI, APIs o automatizaciones.
- Reporte_Ejecutivo.pdf     → una página lista para enviar por WhatsApp o correo (si fpdf2 está instalado).

NEXUS v{APP_VERSION} · datos base persistidos en SQLite · recálculo dinámico.
"""
    return text.encode("utf-8")


def export_bundle(
    excel_bytes: bytes,
    all_df: pd.DataFrame,
    transfers: pd.DataFrame,
    alerts: pd.DataFrame,
    snapshot_label: str,
    score: int,
) -> bytes:
    """Paquete ZIP con las 6 salidas: Excel operativo + gerencial, HTML,
    PDF ejecutivo (si está disponible), CSV, JSON y un README."""
    html = build_interactive_html_report(all_df, transfers, alerts, snapshot_label, score)
    gerencial = build_executive_excel(all_df, snapshot_label, score)
    pdf_bytes = build_executive_pdf(all_df, transfers, snapshot_label, score)
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr(f"Reporte_Operativo_{snapshot_label}.xlsx", excel_bytes)
        z.writestr(f"Reporte_Gerencial_{snapshot_label}.xlsx", gerencial)
        z.writestr(f"Datos_{snapshot_label}.csv", export_csv_bytes(all_df))
        z.writestr(f"Datos_{snapshot_label}.json", export_json_bytes(all_df))
        z.writestr(f"Dashboard_Gerencial_{snapshot_label}.html", html)
        if pdf_bytes:
            z.writestr(f"Reporte_Ejecutivo_{snapshot_label}.pdf", pdf_bytes)
        z.writestr("README.txt", build_readme_text(snapshot_label))
    return out.getvalue()


# ============================================================
# SQLite — PERSISTENCIA DIARIA
# ============================================================

def db() -> sqlite3.Connection:
    """Conexión corta por operación (no una única conexión global con
    @st.cache_resource). NEXUS está pensado para crecer a varios usuarios
    concurrentes, así que priorizamos integridad de datos sobre ahorrarnos
    unos milisegundos de apertura de conexión. WAL permite lecturas
    concurrentes mientras hay una escritura en curso."""
    con = sqlite3.connect(DB_PATH, timeout=10)
    con.execute("PRAGMA journal_mode=WAL")
    return con


def init_db() -> None:
    with db() as con:
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS inventory_snapshots (
                snapshot_date TEXT NOT NULL,
                site TEXT NOT NULL,
                code TEXT NOT NULL,
                description TEXT,
                existence REAL NOT NULL DEFAULT 0,
                cost REAL NOT NULL DEFAULT 0,
                supplier TEXT,
                category TEXT,
                PRIMARY KEY(snapshot_date, site, code)
            )
            """
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS sales_snapshots (
                snapshot_date TEXT NOT NULL,
                site TEXT NOT NULL,
                code TEXT NOT NULL,
                description TEXT,
                sales REAL NOT NULL DEFAULT 0,
                PRIMARY KEY(snapshot_date, site, code)
            )
            """
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS processing_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                processed_at TEXT NOT NULL,
                snapshot_date TEXT NOT NULL,
                site TEXT NOT NULL,
                inventory_rows INTEGER,
                sales_rows INTEGER,
                user_label TEXT DEFAULT ''
            )
            """
        )
        con.commit()


def db_version() -> str:
    with db() as con:
        row = con.execute(
            "SELECT COALESCE(MAX(processed_at),'') FROM processing_log"
        ).fetchone()
    return str(row[0] or "")


def latest_snapshot_date(site: str) -> str | None:
    with db() as con:
        row = con.execute(
            "SELECT MAX(snapshot_date) FROM inventory_snapshots WHERE site=?",
            (site,),
        ).fetchone()
    return row[0] if row and row[0] else None


def previous_snapshot_date(site: str, current_date: str) -> str | None:
    with db() as con:
        row = con.execute(
            "SELECT MAX(snapshot_date) FROM sales_snapshots WHERE site=? AND snapshot_date<?",
            (site, current_date),
        ).fetchone()
    return row[0] if row and row[0] else None


def list_snapshot_dates(site: str | None = None) -> list[str]:
    with db() as con:
        if site:
            rows = con.execute(
                "SELECT DISTINCT snapshot_date FROM inventory_snapshots WHERE site=? ORDER BY snapshot_date DESC",
                (site,),
            ).fetchall()
        else:
            rows = con.execute(
                "SELECT DISTINCT snapshot_date FROM inventory_snapshots ORDER BY snapshot_date DESC"
            ).fetchall()
    return [r[0] for r in rows]


def save_snapshot(
    site: str,
    snapshot_date: str,
    inventory: pd.DataFrame,
    sales: pd.DataFrame,
) -> None:
    # Bulk insert vectorizado: se arma la tabla completa con pandas/numpy
    # (columnas ya alineadas al esquema de SQLite) y se convierte a tuplas
    # en un solo paso con itertuples, en vez de reconstruir fila a fila con
    # getattr. Mantiene el esquema limpio (code/description/... en inglés)
    # sin depender de to_sql, que no resuelve el mapeo de nombres por sí solo.
    inv = inventory.copy()
    sal = sales.copy()

    inv_shaped = pd.DataFrame({
        "snapshot_date": snapshot_date,
        "site": site,
        "code": inv.get("Código", "").astype(str),
        "description": inv.get("Descripción", "").fillna("").astype(str) if "Descripción" in inv.columns else "",
        "existence": pd.to_numeric(inv.get("Existencia", 0), errors="coerce").fillna(0.0),
        "cost": pd.to_numeric(inv.get("Costo", 0), errors="coerce").fillna(0.0),
        "supplier": inv.get("Proveedor", "").fillna("").astype(str) if "Proveedor" in inv.columns else "",
        "category": inv.get("Categoría", "").fillna("").astype(str) if "Categoría" in inv.columns else "",
    })
    inv_records = list(inv_shaped.itertuples(index=False, name=None))

    sal_shaped = pd.DataFrame({
        "snapshot_date": snapshot_date,
        "site": site,
        "code": sal.get("Código", "").astype(str),
        "description": sal.get("Descripción", "").fillna("").astype(str) if "Descripción" in sal.columns else "",
        "sales": pd.to_numeric(sal.get("Ventas", 0), errors="coerce").fillna(0.0),
    })
    sal_records = list(sal_shaped.itertuples(index=False, name=None))

    with db() as con:
        con.execute(
            "DELETE FROM inventory_snapshots WHERE site=? AND snapshot_date=?",
            (site, snapshot_date),
        )
        con.execute(
            "DELETE FROM sales_snapshots WHERE site=? AND snapshot_date=?",
            (site, snapshot_date),
        )
        con.executemany(
            "INSERT INTO inventory_snapshots VALUES (?,?,?,?,?,?,?,?)",
            inv_records,
        )
        con.executemany(
            "INSERT INTO sales_snapshots VALUES (?,?,?,?,?)",
            sal_records,
        )
        con.execute(
            "INSERT INTO processing_log(processed_at,snapshot_date,site,inventory_rows,sales_rows) VALUES(?,?,?,?,?)",
            (
                datetime.now().isoformat(timespec="seconds"),
                snapshot_date,
                site,
                len(inv_records),
                len(sal_records),
            ),
        )
        con.commit()

    # Limpia caché porque cambió la fuente de verdad.
    st.cache_data.clear()


@st.cache_data(show_spinner=False)
def load_snapshot(site: str, snapshot_date: str) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    with db() as con:
        inv = pd.read_sql_query(
            "SELECT code AS Código, description AS Descripción, existence AS Existencia, cost AS Costo, supplier AS Proveedor, category AS Categoría FROM inventory_snapshots WHERE site=? AND snapshot_date=?",
            con,
            params=(site, snapshot_date),
        )
        sales = pd.read_sql_query(
            "SELECT code AS Código, description AS Descripción, sales AS Ventas FROM sales_snapshots WHERE site=? AND snapshot_date=?",
            con,
            params=(site, snapshot_date),
        )

        prev_date = previous_snapshot_date(site, snapshot_date)
        if prev_date:
            prev_sales = pd.read_sql_query(
                "SELECT code AS Código, sales AS Ventas FROM sales_snapshots WHERE site=? AND snapshot_date=?",
                con,
                params=(site, prev_date),
            )
        else:
            prev_sales = pd.DataFrame(columns=["Código", "Ventas"])
    return inv, sales, prev_sales


# ============================================================
# EXCEL INPUT — NORMALIZACIÓN ROBUSTA
# ============================================================

def find_header_row(raw: pd.DataFrame) -> int:
    keywords = [
        "código", "codigo", "cod", "artículo", "articulo", "sku",
        "descripción", "descripcion", "existencia", "stock", "saldo",
        "cantidad", "ventas", "costo", "proveedor", "categoria", "categoría",
    ]
    best_row, best_score = 0, 0
    for idx, row in raw.iterrows():
        vals = [normalize_text(v) for v in row.tolist() if pd.notna(v)]
        score = sum(1 for kw in keywords if any(kw == v or kw in v for v in vals))
        if score > best_score:
            best_row, best_score = idx, score
    return best_row


def standardize_columns(df: pd.DataFrame, kind: str) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    aliases = {
        "Código": ["código", "codigo", "cod", "cod_articulo", "cod. artículo", "artículo", "articulo", "sku", "codigo producto"],
        "Descripción": ["descripción", "descripcion", "desc", "producto", "nombre", "detalle"],
        "Ventas": ["cantidad", "ventas", "salidas", "venta", "unidades vendidas", "cantidad vendida", "movimiento"],
        "Existencia": ["existencia", "stock", "saldo", "inventario", "disponible", "existencias"],
        "Costo": ["costo", "precio", "costo unitario", "coste", "precio costo", "precio de costo"],
        "Proveedor": ["proveedor", "marca/proveedor", "vendor", "supplier"],
        "Categoría": ["categoría", "categoria", "familia", "rubro", "grupo", "linea", "línea"],
    }
    rename = {}
    for col in df.columns:
        c = normalize_text(col)
        for target, variants in aliases.items():
            if c in variants or any(c.startswith(v + " ") for v in variants):
                if target not in rename.values():
                    rename[col] = target
                break
    df = df.rename(columns=rename)

    for col in ["Ventas", "Existencia", "Costo"]:
        if col in df.columns:
            df[col] = df[col].map(parse_number)

    if "Código" in df.columns:
        df["Código"] = (
            df["Código"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        )
        df = df[(df["Código"] != "") & (df["Código"].str.lower() != "nan")]

    if "Descripción" not in df.columns:
        df["Descripción"] = ""
    df["Descripción"] = df["Descripción"].fillna("").astype(str).str.strip()

    if kind == "inventario":
        if "Existencia" not in df.columns:
            df["Existencia"] = 0.0
        if "Costo" not in df.columns:
            df["Costo"] = 0.0
    else:
        if "Ventas" not in df.columns:
            df["Ventas"] = 0.0

    if "Código" in df.columns:
        agg: dict[str, str] = {}
        for col in df.columns:
            if col == "Código":
                continue
            if col in ["Ventas", "Existencia"]:
                agg[col] = "sum"
            else:
                agg[col] = "first"
        df = df.groupby("Código", as_index=False).agg(agg)

    return df.reset_index(drop=True)


def _read_excel_fast(bio: io.BytesIO, **kwargs) -> pd.DataFrame:
    """Lee Excel con calamine si está disponible (más rápido), con fallback
    automático y silencioso al motor por defecto de pandas/openpyxl."""
    if EXCEL_ENGINE:
        try:
            bio.seek(0)
            return pd.read_excel(bio, engine=EXCEL_ENGINE, **kwargs)
        except Exception:
            pass
    bio.seek(0)
    return pd.read_excel(bio, **kwargs)


@st.cache_data(show_spinner=False)
def parse_uploaded_excel(raw_bytes: bytes, kind: str) -> tuple[pd.DataFrame, int]:
    bio = io.BytesIO(raw_bytes)
    raw = _read_excel_fast(bio, header=None)
    header_row = find_header_row(raw)
    df = _read_excel_fast(bio, skiprows=header_row)
    df = standardize_columns(df, kind)
    df = downcast_dataframe(df)
    return df, header_row


def downcast_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Reduce el uso de memoria de forma segura: solo columnas categóricas
    de baja cardinalidad y enteros que realmente caben en tipos chicos.
    No se toca nada indiscriminadamente para no perder precisión."""
    if df.empty:
        return df
    df = df.copy()
    category_like = ["Estado", "Prioridad", "ABC", "Acción", "Sede", "Categoría",
                      "Proveedor", "Tiene costo", "Stock negativo", "Sin movimiento"]
    for col in category_like:
        if col in df.columns and df[col].dtype == object:
            if df[col].nunique(dropna=False) < max(50, len(df) // 2):
                df[col] = df[col].astype("category")
    return df


# ============================================================
# MOTOR LOGÍSTICO
# ============================================================

@st.cache_data(show_spinner=False)
def calculate_site(
    inv: pd.DataFrame,
    sales: pd.DataFrame,
    prev_sales: pd.DataFrame,
    months_history: int,
    min_coverage: float,
    max_coverage: float,
    lead_time_days: int,
    safety_days: int,
    abc_basis: str,
    sales_mode: str,
    rolling_days: int,
) -> pd.DataFrame:
    inv = inv.copy()
    sales = sales.copy()
    prev_sales = prev_sales.copy()

    if "Existencia" not in inv.columns:
        inv["Existencia"] = 0.0
    if "Costo" not in inv.columns:
        inv["Costo"] = 0.0
    if "Descripción" not in inv.columns:
        inv["Descripción"] = ""
    if "Proveedor" not in inv.columns:
        inv["Proveedor"] = ""
    if "Categoría" not in inv.columns:
        inv["Categoría"] = ""
    if "Ventas" not in sales.columns:
        sales["Ventas"] = 0.0

    current = sales[["Código", "Ventas"]].copy()
    if sales_mode == "acumuladas":
        prev = prev_sales.rename(columns={"Ventas": "Ventas Anteriores"})
        current = current.merge(prev, on="Código", how="left")
        current["Ventas del Corte"] = np.maximum(
            current["Ventas"].fillna(0) - current["Ventas Anteriores"].fillna(0), 0
        )
        current = current[["Código", "Ventas del Corte"]]
        current = current.rename(columns={"Ventas del Corte": "Ventas"})

    df = inv.merge(current, on="Código", how="outer", suffixes=("_inv", "_ventas"))

    for col in ["Existencia", "Costo"]:
        if col not in df.columns:
            df[col] = 0.0
    if "Ventas" not in df.columns:
        df["Ventas"] = 0.0
    if "Descripción_inv" in df.columns or "Descripción_ventas" in df.columns:
        d1 = df.get("Descripción_inv", pd.Series("", index=df.index)).fillna("")
        d2 = df.get("Descripción_ventas", pd.Series("", index=df.index)).fillna("")
        df["Descripción"] = d1.where(d1.astype(str).str.strip() != "", d2)
    elif "Descripción" not in df.columns:
        df["Descripción"] = ""

    # Tipo diario: el reporte actual representa movimiento diario.
    # La demanda mensual se anualiza desde un promedio del corte diario.
    # Para minimizar saltos, el usuario puede aumentar los días de ventana.
    df["Ventas"] = pd.to_numeric(df["Ventas"], errors="coerce").fillna(0)
    df["Existencia"] = pd.to_numeric(df["Existencia"], errors="coerce").fillna(0)
    df["Costo"] = pd.to_numeric(df["Costo"], errors="coerce").fillna(0)

    months_history = max(1, int(months_history))
    min_coverage = max(0.1, float(min_coverage))
    max_coverage = max(min_coverage, float(max_coverage))
    lead_time_days = max(0, int(lead_time_days))
    safety_days = max(0, int(safety_days))
    rolling_days = max(1, int(rolling_days))

    if sales_mode == "diarias":
        monthly_demand = df["Ventas"] * 30
    else:
        monthly_demand = df["Ventas"] / months_history

    df["Demanda Mensual"] = np.ceil(monthly_demand).astype(int)
    df["Demanda Diaria"] = (monthly_demand / 30).round(2)

    # Cuando no existe demanda, usamos la política mínima de protección.
    df["Stock Mínimo"] = np.where(
        df["Demanda Mensual"] > 0,
        np.ceil(df["Demanda Mensual"] * min_coverage),
        1,
    ).astype(int)
    df["Stock Máximo"] = np.where(
        df["Demanda Mensual"] > 0,
        np.ceil(df["Demanda Mensual"] * max_coverage),
        2,
    ).astype(int)

    df["Demanda Lead Time"] = np.ceil(df["Demanda Diaria"] * lead_time_days).astype(int)
    df["Stock Seguridad"] = np.ceil(df["Demanda Diaria"] * safety_days).astype(int)
    df["Punto de Reorden"] = np.maximum(
        df["Stock Mínimo"], df["Demanda Lead Time"] + df["Stock Seguridad"]
    ).astype(int)

    df["Compra Sugerida"] = np.where(
        df["Existencia"] < df["Punto de Reorden"],
        np.maximum(df["Stock Máximo"] - df["Existencia"], 0),
        0,
    ).astype(int)

    df["Retiro Almacén"] = np.where(
        df["Existencia"] > df["Stock Máximo"],
        np.maximum(df["Existencia"] - df["Stock Máximo"], 0),
        0,
    ).astype(int)

    positive_stock = np.maximum(df["Existencia"], 0)
    df["Cobertura (meses)"] = np.where(
        df["Demanda Mensual"] > 0,
        (positive_stock / df["Demanda Mensual"]).round(1),
        np.nan,
    )
    df["Rotación estimada (x/mes)"] = np.where(
        positive_stock > 0,
        (df["Demanda Mensual"] / positive_stock).round(2),
        0,
    )

    df["Valor Inventario ($)"] = (df["Existencia"] * df["Costo"]).round(2)
    df["Capital Inmovilizado ($)"] = (df["Retiro Almacén"] * df["Costo"]).round(2)
    df["Costo Compra Estimada ($)"] = (df["Compra Sugerida"] * df["Costo"]).round(2)
    df["Valor Movimiento ($)"] = (df["Demanda Mensual"] * df["Costo"]).round(2)

    # ABC configurable.
    if abc_basis == "Inventario":
        abc_value = np.maximum(df["Valor Inventario ($)"], 0)
    elif abc_basis == "Capital inmovilizado":
        abc_value = np.maximum(df["Capital Inmovilizado ($)"], 0)
    else:
        abc_value = np.maximum(df["Valor Movimiento ($)"], 0)

    total_abc = float(abc_value.sum())
    if total_abc > 0:
        order = np.argsort(-abc_value.to_numpy())
        cumulative = np.zeros(len(df))
        ordered = abc_value.to_numpy()[order]
        cumulative[order] = np.cumsum(ordered) / total_abc
        df["ABC"] = np.select(
            [cumulative <= 0.80, cumulative <= 0.95],
            ["A", "B"],
            default="C",
        )
    else:
        df["ABC"] = "C"

    df["Tiene costo"] = np.where(df["Costo"] > 0, "Sí", "No")
    df["Stock negativo"] = np.where(df["Existencia"] < 0, "Sí", "No")
    df["Sin movimiento"] = np.where((df["Ventas"] <= 0) & (df["Existencia"] > 0), "Sí", "No")

    df["Estado"] = np.select(
        [
            df["Existencia"] < 0,
            df["Existencia"] < df["Punto de Reorden"],
            df["Retiro Almacén"] > 0,
        ],
        [
            "CRÍTICO — SALDO NEGATIVO",
            "CRÍTICO — COMPRAR",
            "SOBRESTOCK — RETIRAR",
        ],
        default="ÓPTIMO",
    )

    df["Prioridad"] = np.select(
        [
            df["Existencia"] < 0,
            (df["Compra Sugerida"] > 0) & (df["ABC"] == "A"),
            df["Compra Sugerida"] > 0,
            (df["Retiro Almacén"] > 0) & (df["ABC"] == "A"),
            df["Retiro Almacén"] > 0,
        ],
        ["CRÍTICA", "ALTA", "MEDIA", "ALTA", "MEDIA"],
        default="BAJA",
    )

    # Acción inicial. Se ajusta luego con redistribución global.
    df["Acción"] = np.select(
        [
            df["Existencia"] < 0,
            df["Compra Sugerida"] > 0,
            df["Retiro Almacén"] > 0,
        ],
        ["REVISAR SALDO / ABASTECER", "COMPRAR", "RETIRAR / REDISTRIBUIR"],
        default="NO HACER NADA",
    )

    keep = [
        "Código", "Descripción", "Proveedor", "Categoría", "Ventas", "Demanda Mensual",
        "Demanda Diaria", "Existencia", "Costo", "Stock Mínimo", "Stock Máximo",
        "Demanda Lead Time", "Stock Seguridad", "Punto de Reorden", "Compra Sugerida",
        "Retiro Almacén", "Cobertura (meses)", "Rotación estimada (x/mes)", "Valor Inventario ($)",
        "Capital Inmovilizado ($)", "Costo Compra Estimada ($)", "Valor Movimiento ($)",
        "ABC", "Tiene costo", "Stock negativo", "Sin movimiento", "Estado", "Prioridad", "Acción",
    ]
    for col in keep:
        if col not in df.columns:
            df[col] = ""
    result = df[keep].copy()
    # Presentación operativa: cantidades físicas siempre como enteros;
    # dinero a dos decimales y métricas continuas con precisión controlada.
    integer_cols = [
        "Ventas", "Demanda Mensual", "Existencia", "Stock Mínimo", "Stock Máximo",
        "Demanda Lead Time", "Stock Seguridad", "Punto de Reorden", "Compra Sugerida",
        "Retiro Almacén", "Transferencias Recibidas", "Transferencias Enviadas", "Compra Ajustada"
    ]
    for col in integer_cols:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors="coerce").fillna(0).round().astype(int)
    money_cols = [
        "Costo", "Valor Inventario ($)", "Capital Inmovilizado ($)",
        "Costo Compra Estimada ($)", "Valor Movimiento ($)"
    ]
    for col in money_cols:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors="coerce").fillna(0).round(2)
    if "Demanda Diaria" in result.columns:
        result["Demanda Diaria"] = pd.to_numeric(result["Demanda Diaria"], errors="coerce").fillna(0).round(2)
    if "Cobertura (meses)" in result.columns:
        result["Cobertura (meses)"] = pd.to_numeric(result["Cobertura (meses)"], errors="coerce").round(1)
    if "Rotación estimada (x/mes)" in result.columns:
        result["Rotación estimada (x/mes)"] = pd.to_numeric(result["Rotación estimada (x/mes)"], errors="coerce").fillna(0).round(2)
    result = result.sort_values(
        by=["Prioridad", "Capital Inmovilizado ($)", "Demanda Mensual"],
        ascending=[True, False, False],
    ).reset_index(drop=True)
    return downcast_dataframe(result)


# ============================================================
# REDISTRIBUCIÓN MULTISEDE
# ============================================================

def build_redistribution(site_results: dict[str, pd.DataFrame]) -> pd.DataFrame:
    frames = []
    for site, df in site_results.items():
        if df is None or df.empty:
            continue
        x = df[[
            "Código", "Descripción", "Existencia", "Stock Máximo",
            "Stock Mínimo", "Costo", "Demanda Mensual"
        ]].copy()
        x.insert(0, "Sede", site)
        frames.append(x)
    if not frames:
        return pd.DataFrame()

    all_df = pd.concat(frames, ignore_index=True)
    suggestions: list[dict[str, Any]] = []

    for code, group in all_df.groupby("Código"):
        donors = group.copy()
        receivers = group.copy()
        donors["Exceso"] = np.maximum(donors["Existencia"] - donors["Stock Máximo"], 0)
        receivers["Déficit"] = np.maximum(receivers["Stock Mínimo"] - receivers["Existencia"], 0)
        donors = donors[donors["Exceso"] > 0].sort_values("Exceso", ascending=False)
        receivers = receivers[receivers["Déficit"] > 0].sort_values("Déficit", ascending=False)

        for r in receivers.itertuples(index=False):
            need = float(r.Déficit)
            if need <= 0:
                continue
            for d in donors.itertuples(index=False):
                if d.Sede == r.Sede:
                    continue
                available = float(d.Exceso)
                if available <= 0:
                    continue
                qty = int(np.floor(min(need, available)))
                if qty <= 0:
                    continue
                cost = safe_float(r.Costo or d.Costo)
                suggestions.append(
                    {
                        "Código": code,
                        "Descripción": r.Descripción or d.Descripción,
                        "Origen": d.Sede,
                        "Destino": r.Sede,
                        "Existencia Origen": int(round(d.Existencia)),
                        "Necesidad Destino": int(round(r.Déficit)),
                        "Unidades Sugeridas": qty,
                        "Costo Unitario ($)": cost,
                        "Compra Evitada Estimada ($)": round(qty * cost, 2),
                        "Estado": "RECOMENDADA",
                    }
                )
                need -= qty
                # Mutamos la copia lógica del donor para permitir varios destinos.
                idx = donors.index[donors["Sede"] == d.Sede][0]
                donors.loc[idx, "Exceso"] -= qty
                if need <= 0:
                    break

    if not suggestions:
        return pd.DataFrame(
            columns=[
                "Código", "Descripción", "Origen", "Destino", "Existencia Origen",
                "Necesidad Destino", "Unidades Sugeridas", "Costo Unitario ($)",
                "Compra Evitada Estimada ($)", "Estado"
            ]
        )

    return pd.DataFrame(suggestions).sort_values(
        "Compra Evitada Estimada ($)", ascending=False
    ).reset_index(drop=True)


def apply_redistribution_to_results(
    site_results: dict[str, pd.DataFrame],
    transfers: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    out = {site: df.copy() for site, df in site_results.items()}
    if transfers.empty:
        return out

    received: dict[tuple[str, str], int] = {}
    sent: dict[tuple[str, str], int] = {}
    for _, row in transfers.iterrows():
        qty = int(row["Unidades Sugeridas"])
        received[(row["Destino"], row["Código"])] = received.get((row["Destino"], row["Código"]), 0) + qty
        sent[(row["Origen"], row["Código"])] = sent.get((row["Origen"], row["Código"]), 0) + qty

    for site, df in out.items():
        for idx, row in df.iterrows():
            code = row["Código"]
            rec = received.get((site, code), 0)
            snd = sent.get((site, code), 0)
            df.at[idx, "Transferencias Recibidas"] = rec
            df.at[idx, "Transferencias Enviadas"] = snd
            df.at[idx, "Compra Ajustada"] = max(int(row["Compra Sugerida"]) - rec, 0)
            if rec > 0 and int(row["Compra Sugerida"]) > 0:
                df.at[idx, "Acción"] = "TRANSFERIR DESDE OTRA SEDE"
            if snd > 0 and int(row["Retiro Almacén"]) > 0:
                df.at[idx, "Acción"] = "TRANSFERIR A OTRA SEDE"
    return out


# ============================================================
# MÉTRICAS GLOBALES
# ============================================================

def combine_sites(site_results: dict[str, pd.DataFrame]) -> pd.DataFrame:
    frames = []
    for site, df in site_results.items():
        if df is None or df.empty:
            continue
        x = df.copy()
        x.insert(0, "Sede", site)
        frames.append(x)
    if not frames:
        return pd.DataFrame()
    return downcast_dataframe(pd.concat(frames, ignore_index=True))


def health_score(df: pd.DataFrame) -> int:
    if df.empty:
        return 0
    n = len(df)
    critical = float((df["Estado"].str.startswith("CRÍTICO")).sum()) / n
    overstock = float((df["Estado"] == "SOBRESTOCK — RETIRAR").sum()) / n
    dead = float((df["Sin movimiento"] == "Sí").sum()) / n
    no_cost = float((df["Tiene costo"] == "No").sum()) / n
    negative = float((df["Stock negativo"] == "Sí").sum()) / n
    score = 100 - (critical * 45 + overstock * 25 + dead * 15 + no_cost * 10 + negative * 5)
    return int(np.clip(round(score), 0, 100))


def health_label(score: int) -> str:
    if score >= 85:
        return "SALUDABLE"
    if score >= 70:
        return "VIGILAR"
    if score >= 50:
        return "RIESGO"
    return "CRÍTICA"


def build_alerts(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["Tipo", "Severidad", "Sede", "Código", "Descripción", "Detalle"])
    alerts = []
    for _, row in df.iterrows():
        if row["Estado"] == "CRÍTICO — SALDO NEGATIVO":
            alerts.append({
                "Tipo": "Saldo negativo", "Severidad": "CRÍTICA", "Sede": row["Sede"],
                "Código": row["Código"], "Descripción": row["Descripción"],
                "Detalle": f"Existencia {int(round(row['Existencia']))} · revisar origen y abastecer.",
            })
        elif row["Estado"] == "CRÍTICO — COMPRAR":
            alerts.append({
                "Tipo": "Stock crítico", "Severidad": "ALTA", "Sede": row["Sede"],
                "Código": row["Código"], "Descripción": row["Descripción"],
                "Detalle": f"Existencia {int(round(row['Existencia']))} · punto de reorden {int(row['Punto de Reorden'])}.",
            })
        elif row["Estado"] == "SOBRESTOCK — RETIRAR":
            alerts.append({
                "Tipo": "Sobrestock", "Severidad": "MEDIA", "Sede": row["Sede"],
                "Código": row["Código"], "Descripción": row["Descripción"],
                "Detalle": f"Retirar {int(round(row['Retiro Almacén']))} unidades.",
            })
        if row["Sin movimiento"] == "Sí" and row["Existencia"] > 0:
            alerts.append({
                "Tipo": "Sin movimiento", "Severidad": "MEDIA", "Sede": row["Sede"],
                "Código": row["Código"], "Descripción": row["Descripción"],
                "Detalle": f"Hay {int(round(row['Existencia']))} unidades sin venta en el corte.",
            })
        if row["Tiene costo"] == "No":
            alerts.append({
                "Tipo": "Costo faltante", "Severidad": "BAJA", "Sede": row["Sede"],
                "Código": row["Código"], "Descripción": row["Descripción"],
                "Detalle": "Costo unitario no disponible; el análisis financiero puede estar subestimado.",
            })
    return pd.DataFrame(alerts, columns=["Tipo", "Severidad", "Sede", "Código", "Descripción", "Detalle"])


# ============================================================
# CONTACTOS / AUTOMATIZACIÓN
# ============================================================

def load_contacts() -> dict[str, dict[str, str]]:
    if not CONTACTS_PATH.exists():
        return DEFAULT_CONTACTS.copy()
    try:
        raw = json.loads(CONTACTS_PATH.read_text(encoding="utf-8"))
        result = DEFAULT_CONTACTS.copy()
        for site in SEDES:
            result[site] = {
                "supervisor": str(raw.get(site, {}).get("supervisor", "")),
                "whatsapp": str(raw.get(site, {}).get("whatsapp", "")),
                "webhook": str(raw.get(site, {}).get("webhook", "")),
            }
        return result
    except Exception:
        return DEFAULT_CONTACTS.copy()


def save_contacts(contacts: dict[str, dict[str, str]]) -> None:
    CONTACTS_PATH.write_text(
        json.dumps(contacts, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def compose_digest(site: str, df: pd.DataFrame, transfers: pd.DataFrame) -> str:
    if df.empty:
        return f"MAKROPETROL NEXUS · {site}\nSin datos procesados."
    score = health_score(df)
    critical = int(df["Estado"].str.startswith("CRÍTICO").sum())
    purchases = int((df["Compra Ajustada"] > 0).sum()) if "Compra Ajustada" in df.columns else int((df["Compra Sugerida"] > 0).sum())
    withdrawals = int((df["Retiro Almacén"] > 0).sum())
    transfers_out = int(df.get("Transferencias Enviadas", pd.Series(dtype=float)).sum()) if "Transferencias Enviadas" in df.columns else 0
    transfers_in = int(df.get("Transferencias Recibidas", pd.Series(dtype=float)).sum()) if "Transferencias Recibidas" in df.columns else 0
    capital = float(df["Capital Inmovilizado ($)"].sum())
    parts = [
        f"*MAKROPETROL NEXUS* · {site}",
        f"Salud logística: {score}/100 · {health_label(score)}",
        f"🔴 Críticos: {critical}",
        f"🛒 Compras pendientes: {purchases}",
        f"📤 Retiros: {withdrawals}",
        f"♻️ Transferencias recibidas/enviadas: {transfers_in}/{transfers_out}",
        f"💰 Capital inmovilizado: {money(capital)}",
    ]
    if not transfers.empty:
        own = transfers[(transfers["Origen"] == site) | (transfers["Destino"] == site)].head(3)
        if not own.empty:
            parts.append("\n*Movimientos prioritarios:*")
            for r in own.itertuples(index=False):
                parts.append(f"• {r.Código} · {r.Origen} → {r.Destino} · {int(r.Unidades_Sugeridas)} uds")
    return "\n".join(parts)


def post_webhook(url: str, payload: dict[str, Any]) -> tuple[bool, str]:
    if not url.strip():
        return False, "Webhook vacío"
    try:
        r = requests.post(url.strip(), json=payload, timeout=15)
        if 200 <= r.status_code < 300:
            return True, f"HTTP {r.status_code}"
        return False, f"HTTP {r.status_code}: {r.text[:180]}"
    except Exception as exc:
        return False, str(exc)


# ============================================================
# EXCEL PREMIUM
# ============================================================

def export_excel(
    site_results: dict[str, pd.DataFrame],
    transfers: pd.DataFrame,
    months_history: int,
    min_coverage: float,
    max_coverage: float,
    lead_time_days: int,
    snapshot_label: str,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, DoughnutChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    all_df = combine_sites(site_results)
    out = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "00 Dashboard"

    navy = "18243A"
    blue = "3568F5"
    blue_light = "EDF2FF"
    green = "EAF8F1"
    orange = "FFF2DE"
    red = "FCEEEE"
    gray = "7B8799"
    line = "DDE5EF"
    white = "FFFFFF"

    thin = Side(style="thin", color=line)

    def title(ws_, title, subtitle, end_col=10):
        ws_.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        ws_.cell(1, 1).value = title
        ws_.cell(1, 1).font = Font(size=20, bold=True, color=white)
        ws_.cell(1, 1).fill = PatternFill("solid", fgColor=navy)
        ws_.cell(1, 1).alignment = Alignment(vertical="center")
        ws_.row_dimensions[1].height = 34
        ws_.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        ws_.cell(2, 1).value = subtitle
        ws_.cell(2, 1).font = Font(size=10, color=gray, italic=True)
        ws_.row_dimensions[2].height = 23

    def write_table(ws_, df_: pd.DataFrame, start_row=4, start_col=1, table_name="Table1"):
        if df_.empty:
            ws_.cell(start_row, start_col).value = "Sin registros"
            return
        for j, col in enumerate(df_.columns, start_col):
            c = ws_.cell(start_row, j, col)
            c.fill = PatternFill("solid", fgColor=blue)
            c.font = Font(bold=True, color=white)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = Border(bottom=thin)
        for i, row in enumerate(df_.itertuples(index=False), start_row + 1):
            for j, val in enumerate(row, start_col):
                cell = ws_.cell(i, j, val)
                cell.border = Border(bottom=Side(style="hair", color=line))
                cell.alignment = Alignment(vertical="center")
        end_row = start_row + len(df_)
        end_col = start_col + len(df_.columns) - 1
        ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws_.add_table(table)
        ws_.freeze_panes = f"{get_column_letter(start_col)}{start_row+1}"
        ws_.auto_filter.ref = ref
        for j in range(start_col, end_col + 1):
            max_len = max(len(str(ws_.cell(r, j).value or "")) for r in range(start_row, min(end_row, start_row + 100) + 1))
            ws_.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 11), 42)

    score = health_score(all_df)
    kpis = [
        ("Valor inventario", float(all_df["Valor Inventario ($)"].sum()) if not all_df.empty else 0),
        ("Compras", int((all_df.get("Compra Ajustada", all_df.get("Compra Sugerida", pd.Series(dtype=float))) > 0).sum()) if not all_df.empty else 0),
        ("Retiros", int((all_df.get("Retiro Almacén", pd.Series(dtype=float)) > 0).sum()) if not all_df.empty else 0),
        ("Capital inmovilizado", float(all_df["Capital Inmovilizado ($)"].sum()) if not all_df.empty else 0),
    ]
    title(ws, "MAKROPETROL NEXUS · REPORTE EJECUTIVO", f"Corte: {snapshot_label} · Historial: {months_history} meses · Lead time: {lead_time_days} días", 10)
    ws["A4"] = "SALUD LOGÍSTICA"
    ws["A4"].font = Font(bold=True, color=navy, size=12)
    ws["B4"] = f"{score}/100 · {health_label(score)}"
    ws["B4"].font = Font(bold=True, color=blue, size=16)
    for idx, (label, value) in enumerate(kpis, 1):
        col = 1 + (idx - 1) * 2
        ws.cell(6, col, label).font = Font(bold=True, color=gray, size=9)
        ws.cell(7, col, value)
        ws.cell(7, col).font = Font(bold=True, color=navy, size=15)
        if "Valor" in label or "Capital" in label:
            ws.cell(7, col).number_format = '$#,##0.00'
        else:
            ws.cell(7, col).number_format = '#,##0'

    if not all_df.empty:
        by_site = all_df.groupby("Sede").agg(
            SKUs=("Código", "count"),
            Compras=("Compra Sugerida", lambda x: int((x > 0).sum())),
            Retiros=("Retiro Almacén", lambda x: int((x > 0).sum())),
            Inventario=("Valor Inventario ($)", "sum"),
            Inmovilizado=("Capital Inmovilizado ($)", "sum"),
        ).reset_index()
        write_table(ws, by_site, start_row=10, table_name="ResumenSedes")

        chart = BarChart()
        chart.title = "Valor de inventario por sede"
        data_ref = Reference(ws, min_col=6, min_row=10, max_row=10 + len(by_site))
        cat_ref = Reference(ws, min_col=1, min_row=11, max_row=10 + len(by_site))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.height = 7
        chart.width = 12
        ws.add_chart(chart, "H10")

    # Hojas operativas.
    sheets = [
        ("01 Inventario", all_df, "InventarioGeneral"),
        ("02 Compras", all_df[all_df.get("Compra Ajustada", all_df.get("Compra Sugerida", pd.Series(dtype=float))) > 0].copy() if not all_df.empty else all_df, "PlanCompras"),
        ("03 Retiros", all_df[all_df.get("Retiro Almacén", pd.Series(dtype=float)) > 0].copy() if not all_df.empty else all_df, "PlanRetiros"),
        ("04 Redistribución", transfers.copy(), "PlanTransferencias"),
        ("05 Alertas", build_alerts(all_df), "Alertas"),
    ]
    for sheet_name, data, table_name in sheets:
        sh = wb.create_sheet(sheet_name)
        title(sh, sheet_name.replace("01 ", "").replace("02 ", "").replace("03 ", "").replace("04 ", "").replace("05 ", ""), f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", min(max(len(data.columns), 8), 16) if not data.empty else 8)
        write_table(sh, data, start_row=4, table_name=table_name)

    # Criterios.
    crit = wb.create_sheet("06 Criterios")
    title(crit, "CRITERIOS DEL MOTOR", "Parámetros y definiciones usados en el corte", 5)
    criteria = pd.DataFrame(
        [
            ["Meses de historial", months_history, "Divide el total de ventas del período para obtener demanda mensual."],
            ["Cobertura mínima", min_coverage, "Meses de cobertura usados como umbral operativo."],
            ["Cobertura máxima", max_coverage, "Meses de cobertura objetivo antes de sugerir retiro."],
            ["Lead time", lead_time_days, "Días estimados de entrega del proveedor."],
            ["Stock seguridad", "Demanda diaria × días de seguridad", "Protección adicional ante variabilidad/entrega."],
            ["Punto de reorden", "máximo(stock mínimo, demanda lead time + stock seguridad)", "Umbral para iniciar abastecimiento."],
            ["ABC", "80% / 95% / resto", "Clasificación A/B/C según base seleccionada."],
            ["Redistribución", "Exceso de una sede → déficit de otra", "Se prioriza mover inventario existente antes de comprar."],
        ],
        columns=["Parámetro", "Valor", "Interpretación"],
    )
    write_table(crit, criteria, start_row=4, table_name="CriteriosMotor")

    # Formatos por hoja.
    for sh in wb.worksheets:
        sh.sheet_view.showGridLines = False
        for row in sh.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=False)
        # Formatos de moneda y enteros por encabezado.
        headers = {cell.value: cell.column for cell in sh[4]}
        integer_headers = {"Ventas", "Demanda Mensual", "Existencia", "Stock Mínimo", "Stock Máximo", "Demanda Lead Time", "Stock Seguridad", "Punto de Reorden", "Compra Sugerida", "Compra Ajustada", "Retiro Almacén", "Unidades Sugeridas", "Transferencias Recibidas", "Transferencias Enviadas"}
        money_headers = {"Costo", "Valor Inventario ($)", "Capital Inmovilizado ($)", "Costo Compra Estimada ($)", "Valor Movimiento ($)", "Costo Unitario ($)", "Compra Evitada Estimada ($)"}
        for h, col in headers.items():
            if h in integer_headers:
                for c in sh.iter_cols(min_col=col, max_col=col, min_row=5, max_row=sh.max_row):
                    for cell in c:
                        cell.number_format = '#,##0'
            if h in money_headers:
                for c in sh.iter_cols(min_col=col, max_col=col, min_row=5, max_row=sh.max_row):
                    for cell in c:
                        cell.number_format = '$#,##0.00'

    wb.save(out)
    return out.getvalue()


# ============================================================
# INICIALIZACIÓN
# ============================================================

init_db()
contacts = load_contacts()


# ============================================================
# SIDEBAR
# ============================================================

# Grupos de navegación · jerarquía visual pensada para un jefe muy visual:
# menos "módulos técnicos sueltos" y más categorías de negocio.
NAV_GROUPS: dict[str, list[str]] = {
    "▣ INICIO": ["Centro de Control"],
    "▣ OPERACIÓN": ["Inventario", "Compras", "Redistribución"],
    "▣ INTELIGENCIA": ["Centro Visual", "Finanzas", "Alertas"],
    "▣ DATOS": ["Cortes Diarios", "Historial"],
    "▣ AUTOMATIZACIÓN": ["Automatización"],
    "▣ REPORTES": ["Exportación de Datos"],
}

with st.sidebar:
    st.markdown(
        """
        <div style="padding:8px 6px 14px">
            <div style="font-family:Manrope;font-size:1.2rem;font-weight:800">◈ MAKROPETROL <span style="color:#3568f5">NEXUS</span></div>
            <div style="font-size:.72rem;color:#7b8799;margin-top:3px">Intelligence Platform</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    modo_gerencia = st.toggle("👔 Modo Gerencia", value=st.session_state.get("modo_gerencia", False), key="modo_gerencia",
                               help="Oculta columnas técnicas y parámetros; muestra KPIs, tendencias, alertas y recomendaciones.")
    st.caption("🔧 Modo Operativo" if not modo_gerencia else "Vista simplificada para dirección")
    st.divider()

    flat_views = [v for group in NAV_GROUPS.values() for v in group]
    if "nexus_view" not in st.session_state:
        st.session_state["nexus_view"] = flat_views[0]

    for group_label, items in NAV_GROUPS.items():
        st.markdown(f'<div class="nav-group">{group_label}</div>', unsafe_allow_html=True)
        for item in items:
            active = st.session_state["nexus_view"] == item
            if st.button(item, key=f"nav_{item}", use_container_width=True, type="primary" if active else "secondary"):
                st.session_state["nexus_view"] = item
                st.rerun()
    view = st.session_state["nexus_view"]

    st.divider()
    with st.popover("⚙️ Configuración del motor", use_container_width=True):
        st.markdown("**Historial**")
        months_history = st.number_input("Meses representados por ventas", min_value=1, max_value=36, value=3, step=1)
        rolling_days = st.number_input("Ventana diaria (días)", min_value=1, max_value=365, value=30, step=1)
        st.markdown("**Cobertura**")
        min_coverage = st.number_input("Cobertura mínima (meses)", min_value=0.1, max_value=12.0, value=0.5, step=0.5)
        max_coverage = st.number_input("Cobertura máxima (meses)", min_value=0.5, max_value=24.0, value=1.0, step=0.5)
        st.markdown("**Proveedor**")
        lead_time_days = st.number_input("Lead time proveedor (días)", min_value=0, max_value=180, value=7, step=1)
        st.markdown("**Seguridad**")
        safety_days = st.number_input("Stock de seguridad (días)", min_value=0, max_value=90, value=3, step=1)
        st.markdown("**Demanda**")
        sales_mode_label = st.selectbox("Tipo de ventas", list(SALES_MODES.keys()))
        sales_mode = SALES_MODES[sales_mode_label]
        st.markdown("**ABC**")
        abc_basis = st.selectbox("Base ABC", ["Movimiento", "Inventario", "Capital inmovilizado"])

    dates = list_snapshot_dates()
    latest_global = dates[0] if dates else None
    status = f"● {latest_global}" if latest_global else "● Sin cortes"
    st.markdown(f'<span class="badge badge-green">{status}</span>', unsafe_allow_html=True)
    st.caption(f"NEXUS v{APP_VERSION} · SQLite local")


# ============================================================
# CARGAR RESULTADOS LIVE DESDE DATOS BASE
# ============================================================

@st.cache_data(show_spinner=False)
def load_live_results(
    _db_version: str,
    months_history: int,
    min_coverage: float,
    max_coverage: float,
    lead_time_days: int,
    safety_days: int,
    abc_basis: str,
    sales_mode: str,
    rolling_days: int,
) -> tuple[dict[str, pd.DataFrame], dict[str, str]]:
    """Caché global del recálculo de todas las sedes. Se invalida sola
    cuando cambia `_db_version` (nuevo corte guardado) o cualquiera de los
    parámetros del motor — así no se recalcula la empresa entera cada vez
    que solo cambias de módulo dentro de la app."""
    results: dict[str, pd.DataFrame] = {}
    dates_used: dict[str, str] = {}
    for site in SEDES:
        d = latest_snapshot_date(site)
        if not d:
            continue
        inv, sales, prev = load_snapshot(site, d)
        if "Código" not in inv.columns or "Código" not in sales.columns:
            continue
        results[site] = calculate_site(
            inv,
            sales,
            prev,
            months_history=int(months_history),
            min_coverage=float(min_coverage),
            max_coverage=float(max_coverage),
            lead_time_days=int(lead_time_days),
            safety_days=int(safety_days),
            abc_basis=abc_basis,
            sales_mode=sales_mode,
            rolling_days=int(rolling_days),
        )
        dates_used[site] = d
    return results, dates_used


site_results, dates_used = load_live_results(
    db_version(),
    int(months_history),
    float(min_coverage),
    float(max_coverage),
    int(lead_time_days),
    int(safety_days),
    abc_basis,
    sales_mode,
    int(rolling_days),
)
raw_all = combine_sites(site_results)
transfers = build_redistribution(site_results)
site_results_adjusted = apply_redistribution_to_results(site_results, transfers)
all_df = combine_sites(site_results_adjusted)
alerts = build_alerts(all_df)


# ============================================================
# HEADER
# ============================================================

latest_display = max(dates_used.values()) if dates_used else "sin cortes"
score = health_score(all_df)

st.markdown(
    f"""
    <div class="hero">
        <div class="brand">Makropetrol <span>NEXUS</span></div>
        <div class="hero-sub">Centro de inteligencia logística · inventario · abastecimiento · redistribución · finanzas · automatización</div>
        <div class="status-row"><span class="dot"></span> Motor operativo activo <span>•</span> Último corte: {latest_display} <span>•</span> Salud: {score}/100</div>
    </div>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# CENTRO DE CONTROL
# ============================================================

if view == "Centro de Control":
    if modo_gerencia:
        st.markdown('<div class="gov-banner">👔 Modo Gerencia activo · vista simplificada de indicadores y recomendaciones</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Centro de Control Ejecutivo</div>', unsafe_allow_html=True)
    if all_df.empty:
        empty_state("NEXUS está listo", "Carga un corte diario de Inventario + Ventas desde 'Cortes Diarios' para activar el tablero.", "◈")
    else:
        critical = int(all_df["Estado"].str.startswith("CRÍTICO").sum())
        purchase_col = "Compra Ajustada" if "Compra Ajustada" in all_df.columns else "Compra Sugerida"
        purchases = int((all_df[purchase_col] > 0).sum())
        purchase_units = int(all_df[purchase_col].sum())
        withdrawals = int((all_df["Retiro Almacén"] > 0).sum())
        inventory_value = float(all_df["Valor Inventario ($)"].sum())
        immobilized = float(all_df["Capital Inmovilizado ($)"].sum())
        avoided = float(transfers["Compra Evitada Estimada ($)"].sum()) if not transfers.empty else 0.0
        optimal = int((all_df["Estado"] == "ÓPTIMO").sum())

        # Fila de KPIs — igual a la propuesta: INVENTARIO · CRÍTICOS · COMPRA · CAPITAL
        c1, c2, c3, c4 = st.columns(4)
        with c1: kpi("Inventario", money(inventory_value), f"{len(all_df):,} SKUs", "blue")
        with c2: kpi("Críticos", integer(critical), "Saldo negativo + bajo reorden", "red")
        with c3: kpi("Compra", f"{integer(purchase_units)} uds", f"{purchases} SKUs a comprar", "orange")
        with c4: kpi("Capital", money(immobilized), "Inmovilizado por sobrestock", "green")

        # Salud logística — número grande centrado, como en el mockup.
        st.markdown('<div class="section-title">Salud logística</div>', unsafe_allow_html=True)
        st.markdown(
            f"""<div class="panel health-big">
                <div class="num">{score} / 100</div>
                <div class="lbl">{health_label(score)}</div>
            </div>""",
            unsafe_allow_html=True,
        )

        left, right = st.columns(2)
        with left:
            st.markdown('<div class="section-title">Inventario por sede</div>', unsafe_allow_html=True)
            by_site = all_df.groupby("Sede")["Valor Inventario ($)"].sum().sort_values(ascending=False)
            max_val = float(by_site.max()) if len(by_site) else 1.0
            rows_html = ""
            for sede, val in by_site.items():
                pct = max(4, round(val / max_val * 100)) if max_val else 4
                rows_html += (
                    f'<div class="op-row"><span>{SEDE_ICON.get(sede, "◆")} {sede}</span>'
                    f'<div class="bar-track"><div class="bar-fill" style="width:{pct}%"></div></div>'
                    f'<span>{money(val)}</span></div>'
                )
            st.markdown(f'<div class="panel">{rows_html}</div>', unsafe_allow_html=True)
        with right:
            st.markdown('<div class="section-title">Estado operativo</div>', unsafe_allow_html=True)
            state_rows = [
                ("Óptimo", optimal, "green"),
                ("Comprar", purchases, "orange"),
                ("Retirar", withdrawals, "blue"),
                ("Crítico", critical, "red"),
            ]
            rows_html = "".join(
                f'<div class="op-row"><span>{label}</span><span class="badge badge-{tone}">{integer(count)}</span></div>'
                for label, count, tone in state_rows
            )
            st.markdown(f'<div class="panel">{rows_html}</div>', unsafe_allow_html=True)

        # Banner de oportunidades — redistribución y ahorro, visible siempre.
        n_transfers = int(len(transfers)) if not transfers.empty else 0
        st.markdown(
            f"""<div class="opp-banner">
                <span>♻️ {n_transfers} oportunidad(es) de redistribución</span>
                <span>💰 {money(avoided)} de compra potencial evitada</span>
            </div>""",
            unsafe_allow_html=True,
        )

        if not modo_gerencia:
            st.markdown('<div class="section-title">Prioridades inmediatas</div>', unsafe_allow_html=True)
            top_alerts = alerts.head(15) if not alerts.empty else pd.DataFrame()
            if top_alerts.empty:
                st.success("No hay alertas activas con los parámetros actuales.")
            else:
                st.dataframe(top_alerts, use_container_width=True, hide_index=True, height=380)

            a, b = st.columns(2)
            with a:
                action_counts = pd.DataFrame({
                    "Acción": ["Óptimo", "Comprar", "Retirar", "Crítico"],
                    "Cantidad": [optimal, purchases, withdrawals, critical],
                })
                fig = px.pie(action_counts, names="Acción", values="Cantidad", hole=.68)
                fig.update_traces(textposition="inside", textinfo="percent+label")
                st.plotly_chart(chart_layout(fig, 300), use_container_width=True)
            with b:
                top = all_df[all_df["Capital Inmovilizado ($)"] > 0].sort_values("Capital Inmovilizado ($)", ascending=False).head(12)
                if top.empty:
                    st.success("No existe capital inmovilizado por sobrestock.")
                else:
                    fig = px.bar(top.sort_values("Capital Inmovilizado ($)"), x="Capital Inmovilizado ($)", y="Descripción", orientation="h", hover_data=["Sede", "Código", "Retiro Almacén"])
                    st.plotly_chart(chart_layout(fig, 300), use_container_width=True)
        else:
            st.markdown('<div class="section-title">Recomendaciones para dirección</div>', unsafe_allow_html=True)
            recs = []
            if critical > 0:
                recs.append(f"🔴 Atender **{critical} SKUs críticos** (saldo negativo o bajo el punto de reorden) antes de nueva compra.")
            if n_transfers > 0:
                recs.append(f"♻️ Ejecutar las **{n_transfers} redistribuciones** sugeridas: evitan {money(avoided)} en compras nuevas.")
            if withdrawals > 0:
                recs.append(f"📤 Hay **{withdrawals} SKUs en sobrestock**; considerar retiro o promoción para liberar capital.")
            if not recs:
                recs.append("✅ La operación está dentro de parámetros saludables. No se requieren acciones urgentes.")
            st.markdown('<div class="panel">' + "".join(f"<p>{r}</p>" for r in recs) + "</div>", unsafe_allow_html=True)


# ============================================================
# CORTES DIARIOS — FUENTE DE VERDAD
# ============================================================

elif view == "Cortes Diarios":
    st.markdown('<div class="section-title">Cortes diarios · Fuente de verdad</div>', unsafe_allow_html=True)
    st.caption("Carga una o varias sedes en el mismo corte. La información se guarda por sede y fecha y luego se recalculan todos los módulos desde los datos base.")

    chosen_date = st.date_input("Fecha del corte", value=date.today(), key="daily_cut_date")
    snapshot_str = chosen_date.isoformat()

    all_sites = st.checkbox("✅ Seleccionar todas las sedes", value=True, key="select_all_sites")
    if all_sites:
        selected_sites = SEDES.copy()
        st.caption(f"{len(selected_sites)} sedes seleccionadas: " + " · ".join(selected_sites))
    else:
        selected_sites = st.multiselect("Sedes a actualizar", SEDES, default=SEDES, key="selected_sites")

    if not selected_sites:
        empty_state("Selecciona al menos una sede", "Puedes trabajar con todas las sedes o con una selección específica.", "🏢")
    else:
        upload_map = {}
        for site in selected_sites:
            with st.expander(f"{SEDE_ICON[site]}  {site}", expanded=True):
                u1, u2 = st.columns(2)
                with u1:
                    inv_file = st.file_uploader(
                        f"📦 Inventario + costos · {site}",
                        type=["xlsx", "xls"],
                        key=f"daily_inv_{site}_{snapshot_str}",
                    )
                with u2:
                    sales_file = st.file_uploader(
                        f"📈 Ventas · {site}",
                        type=["xlsx", "xls"],
                        key=f"daily_sales_{site}_{snapshot_str}",
                    )
                upload_map[site] = (inv_file, sales_file)

        ready_sites = [site for site, (invf, salf) in upload_map.items() if invf and salf]
        incomplete = [site for site, (invf, salf) in upload_map.items() if bool(invf) ^ bool(salf)]

        if incomplete:
            st.warning("Sedes incompletas (falta uno de los dos archivos): " + ", ".join(incomplete))
        if ready_sites:
            st.markdown(f"**{len(ready_sites)} sede(s) listas para procesar.**")
            if st.button("💾 Guardar y procesar todas las sedes listas", type="primary", use_container_width=True, key=f"save_all_{snapshot_str}"):
                errors=[]
                progress=st.progress(0)
                for idx, site in enumerate(ready_sites, 1):
                    try:
                        inv_file, sales_file = upload_map[site]
                        inv, inv_header = parse_uploaded_excel(inv_file.getvalue(), "inventario")
                        sales, sales_header = parse_uploaded_excel(sales_file.getvalue(), "ventas")
                        if "Código" not in inv.columns or "Existencia" not in inv.columns:
                            raise ValueError("Inventario requiere Código y Existencia")
                        if "Código" not in sales.columns or "Ventas" not in sales.columns:
                            raise ValueError("Ventas requiere Código y Ventas")
                        save_snapshot(site, snapshot_str, inv, sales)
                    except Exception as exc:
                        errors.append(f"{site}: {exc}")
                    progress.progress(idx / max(1, len(ready_sites)))
                if errors:
                    for err in errors: st.error(err)
                if len(errors) < len(ready_sites):
                    st.toast(f"Corte {snapshot_str} guardado · {len(ready_sites)-len(errors)} sede(s) recalculadas", icon="🚀")
                    st.rerun()

    st.markdown('<div class="section-title">Estado de actualización</div>', unsafe_allow_html=True)
    status_rows=[]
    for site in SEDES:
        last=latest_snapshot_date(site)
        status_rows.append({"Sede": f"{SEDE_ICON[site]} {site}", "Último corte": last or "Sin datos", "Estado": "ACTUALIZADA" if last == snapshot_str else "PENDIENTE"})
    st.dataframe(pd.DataFrame(status_rows), use_container_width=True, hide_index=True)


# ============================================================
# INVENTARIO
# ============================================================

elif view == "Centro Visual":
    st.markdown('<div class="section-title">Centro Visual · lectura rápida para gerencia</div>', unsafe_allow_html=True)
    if all_df.empty:
        empty_state("Sin datos visuales", "Carga cortes diarios primero.", "◌")
    else:
        purchase_col = "Compra Ajustada" if "Compra Ajustada" in all_df.columns else "Compra Sugerida"
        critical_df = all_df[all_df["Estado"].astype(str).str.startswith("CRÍTICO")].copy()
        purchase_df = all_df[all_df[purchase_col] > 0].copy()
        over_df = all_df[all_df["Retiro Almacén"] > 0].copy()
        k1,k2,k3,k4 = st.columns(4)
        with k1: kpi("Salud logística", f"{score}/100", health_label(score), "blue")
        with k2: kpi("Críticos", integer(len(critical_df)), "Atención inmediata", "red")
        with k3: kpi("Compra neta", integer(purchase_df[purchase_col].sum()), "Unidades", "orange")
        with k4: kpi("Sobrestock", integer(over_df["Retiro Almacén"].sum()), "Unidades", "green")

        v1,v2 = st.columns(2)
        with v1:
            site = all_df.groupby("Sede")["Valor Inventario ($)"].sum().reset_index()
            fig=px.bar(site,x="Sede",y="Valor Inventario ($)",title="Valor de inventario por sede")
            st.plotly_chart(chart_layout(fig,360),use_container_width=True)
        with v2:
            act=pd.DataFrame({"Tipo":["Crítico","Comprar","Retirar","Óptimo"],"Cantidad":[len(critical_df),len(purchase_df),len(over_df),int((all_df["Estado"]=="ÓPTIMO").sum())]})
            fig=px.pie(act,names="Tipo",values="Cantidad",hole=.62,title="Estado operativo")
            st.plotly_chart(chart_layout(fig,360),use_container_width=True)

        v3,v4 = st.columns(2)
        with v3:
            top=all_df[all_df["Capital Inmovilizado ($)"]>0].sort_values("Capital Inmovilizado ($)",ascending=False).head(10)
            if not top.empty:
                fig=px.bar(top.sort_values("Capital Inmovilizado ($)"),x="Capital Inmovilizado ($)",y="Descripción",orientation="h",title="Capital inmovilizado")
                st.plotly_chart(chart_layout(fig,360),use_container_width=True)
            else:
                empty_state("Sin capital inmovilizado", "No hay sobrestock valorizado.", "💰")
        with v4:
            if not transfers.empty:
                routes=transfers.groupby(["Origen","Destino"])["Unidades Sugeridas"].sum().reset_index()
                fig=px.bar(routes,x="Unidades Sugeridas",y="Origen",color="Destino",orientation="h",title="Redistribución sugerida")
                st.plotly_chart(chart_layout(fig,360),use_container_width=True)
            else:
                empty_state("Sin redistribuciones", "No se detectaron cruces de excedente/déficit.", "♻")

elif view == "Inventario":
    st.markdown('<div class="section-title">Explorador inteligente de inventario</div>', unsafe_allow_html=True)
    if all_df.empty:
        empty_state("Sin inventario", "Procesa al menos una sede desde Cortes Diarios.", "📦")
    else:
        s = st.selectbox("Sede", ["Todas"] + sorted(all_df["Sede"].unique()))
        data = all_df.copy() if s == "Todas" else all_df[all_df["Sede"] == s].copy()
        search = st.text_input("Buscar código o descripción", placeholder="Ej. TMB03 / ACEITE / FILTRO")
        if search:
            q = normalize_text(search)
            data = data[
                data["Código"].astype(str).str.lower().str.contains(q, na=False)
                | data["Descripción"].astype(str).str.lower().str.contains(q, na=False)
            ]
        states = st.multiselect("Estado", sorted(data["Estado"].unique()), default=sorted(data["Estado"].unique()))
        data = data[data["Estado"].isin(states)]
        c1,c2,c3,c4 = st.columns(4)
        purchase_col = "Compra Ajustada" if "Compra Ajustada" in data.columns else "Compra Sugerida"
        with c1: kpi("Productos", integer(len(data)), "Filtrados", "blue")
        with c2: kpi("Críticos", integer(data["Estado"].astype(str).str.startswith("CRÍTICO").sum()), "Atención", "red")
        with c3: kpi("Compras", integer(data[purchase_col].sum()), "Unidades", "orange")
        with c4: kpi("Sobrestock", integer(data["Retiro Almacén"].sum()), "Unidades", "green")

        if modo_gerencia:
            cols = ["Sede", "Código", "Descripción", "ABC", "Estado", "Acción", purchase_col, "Retiro Almacén", "Cobertura (meses)"]
        else:
            cols = ["Sede", "Código", "Descripción", "ABC", "Estado", "Prioridad", "Acción", "Ventas", "Existencia", "Stock Mínimo", "Stock Máximo", "Punto de Reorden", purchase_col, "Retiro Almacén", "Cobertura (meses)"]
        cols = [c for c in cols if c in data.columns]

        # Carga diferida: se muestran las primeras filas y un botón para ver
        # todo. El archivo completo (sin límite) sigue disponible al exportar.
        PAGE_SIZE = 300
        st.session_state.setdefault("inv_show_all", False)
        total_rows = len(data)
        if total_rows > PAGE_SIZE and not st.session_state["inv_show_all"]:
            shown = data[cols].head(PAGE_SIZE)
            st.caption(f"Mostrando {PAGE_SIZE:,} de {total_rows:,} filas · usa la búsqueda/filtros para acotar, o carga el resto.")
        else:
            shown = data[cols]

        inv_col_config: dict[str, Any] = {}
        if "Cobertura (meses)" in shown.columns:
            inv_col_config["Cobertura (meses)"] = st.column_config.NumberColumn(format="%.1f")
        if purchase_col in shown.columns:
            inv_col_config[purchase_col] = st.column_config.ProgressColumn(
                "Compra", min_value=0, max_value=max(1, int(data[purchase_col].max())), format="%d uds"
            )
        if "Retiro Almacén" in shown.columns:
            inv_col_config["Retiro Almacén"] = st.column_config.ProgressColumn(
                "Retiro Almacén", min_value=0, max_value=max(1, int(data["Retiro Almacén"].max())), format="%d uds"
            )
        st.dataframe(
            shown,
            use_container_width=True,
            hide_index=True,
            height=620,
            column_config=inv_col_config,
        )
        if total_rows > PAGE_SIZE and not st.session_state["inv_show_all"]:
            if st.button(f"⬇ Cargar las {total_rows - PAGE_SIZE:,} filas restantes", use_container_width=True):
                st.session_state["inv_show_all"] = True
                st.rerun()

        st.download_button(
            "⬇ Exportar vista filtrada completa (CSV)",
            export_csv_bytes(data[cols]),
            "Inventario_Filtrado_Makropetrol.csv",
            "text/csv",
            use_container_width=True,
        )


# ============================================================
# COMPRAS
# ============================================================

elif view == "Compras":
    st.markdown('<div class="section-title">Centro de abastecimiento</div>', unsafe_allow_html=True)
    if all_df.empty:
        empty_state("Sin plan de compras", "Carga un corte diario para que NEXUS calcule las necesidades.", "🛒")
    else:
        purchase_col = "Compra Ajustada" if "Compra Ajustada" in all_df.columns else "Compra Sugerida"
        data = all_df[all_df[purchase_col] > 0].copy()
        if data.empty:
            st.success("No existen compras pendientes después de considerar redistribuciones.")
        else:
            c1, c2, c3 = st.columns(3)
            with c1: kpi("SKUs a comprar", integer(len(data)), "Compra neta", "orange")
            with c2: kpi("Unidades", integer(data[purchase_col].sum()), "Necesidad neta", "blue")
            with c3: kpi("Costo estimado", money((data[purchase_col] * data["Costo"]).sum()), "Costo disponible", "green")

            a, b = st.columns(2)
            with a: sites_filter = st.multiselect("Sedes", sorted(data["Sede"].unique()), default=sorted(data["Sede"].unique()))
            with b: priorities = st.multiselect("Prioridad", sorted(data["Prioridad"].unique()), default=sorted(data["Prioridad"].unique()))
            data = data[data["Sede"].isin(sites_filter) & data["Prioridad"].isin(priorities)]

            if modo_gerencia:
                cols = ["Sede", "Código", "Descripción", "Proveedor", purchase_col, "Costo Compra Estimada ($)", "Prioridad"]
            else:
                cols = ["Sede", "Código", "Descripción", "Proveedor", "ABC", "Ventas", "Existencia", "Punto de Reorden", purchase_col, "Costo", "Costo Compra Estimada ($)", "Prioridad", "Acción"]
            cols = [c for c in cols if c in data.columns]
            st.dataframe(
                data[cols],
                use_container_width=True,
                hide_index=True,
                height=580,
                column_config={
                    "Costo": st.column_config.NumberColumn(format="$%.2f"),
                    "Costo Compra Estimada ($)": st.column_config.NumberColumn(format="$%.2f"),
                },
            )
            if "Proveedor" in data.columns:
                supplier = data.groupby("Proveedor", dropna=False)[purchase_col].sum().reset_index().sort_values(purchase_col, ascending=False).head(10)
                fig = px.bar(supplier, x=purchase_col, y="Proveedor", orientation="h", title="Top proveedores por unidades a comprar")
                st.plotly_chart(chart_layout(fig, 320), use_container_width=True)

            wb = io.BytesIO()
            data[cols].to_excel(wb, index=False, sheet_name="Plan Compras")
            st.download_button("⬇ Exportar plan de compras", wb.getvalue(), "Plan_Compras_Makropetrol.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ============================================================
# REDISTRIBUCIÓN
# ============================================================

elif view == "Redistribución":
    st.markdown('<div class="section-title">Motor de redistribución entre sedes</div>', unsafe_allow_html=True)
    st.caption("NEXUS cruza excedentes y déficits del mismo código para intentar resolver la necesidad usando inventario existente antes de generar una compra.")
    if transfers.empty:
        empty_state("No hay transferencias sugeridas", "No se detectaron al mismo tiempo excedentes y déficits compatibles.", "♻️")
    else:
        avoided = float(transfers["Compra Evitada Estimada ($)"].sum())
        units = int(transfers["Unidades Sugeridas"].sum())
        c1, c2, c3 = st.columns(3)
        with c1: kpi("Transferencias", integer(len(transfers)), "Movimientos sugeridos", "green")
        with c2: kpi("Unidades", integer(units), "Redistribución total", "blue")
        with c3: kpi("Compra evitada", money(avoided), "Ahorro estimado", "orange")
        st.dataframe(
            transfers.style.format({"Costo Unitario ($)": "${:,.2f}", "Compra Evitada Estimada ($)": "${:,.2f}"}),
            use_container_width=True,
            hide_index=True,
            height=600,
        )
        route = transfers.groupby(["Origen", "Destino"])["Unidades Sugeridas"].sum().reset_index()
        if not route.empty:
            st.markdown('<div class="section-title">Flujo entre sedes</div>', unsafe_allow_html=True)
            nodes = sorted(set(route["Origen"]) | set(route["Destino"]))
            node_index = {n: i for i, n in enumerate(nodes)}
            palette = ["#3568f5", "#17aabe", "#14a36a", "#ef982d", "#df5555", "#244bd7", "#7b8799"]
            node_colors = [palette[i % len(palette)] for i in range(len(nodes))]
            link_colors = [node_colors[node_index[o]] + "55" for o in route["Origen"]]
            sankey = go.Figure(go.Sankey(
                arrangement="snap",
                node=dict(
                    label=[f"{SEDE_ICON.get(n,'◆')} {n}" for n in nodes],
                    color=node_colors,
                    pad=18,
                    thickness=16,
                    line=dict(color="rgba(0,0,0,0)", width=0),
                ),
                link=dict(
                    source=[node_index[o] for o in route["Origen"]],
                    target=[node_index[d] for d in route["Destino"]],
                    value=route["Unidades Sugeridas"],
                    color=link_colors,
                    label=[f"{int(v):,} uds" for v in route["Unidades Sugeridas"]],
                ),
            ))
            st.plotly_chart(chart_layout(sankey, 380), use_container_width=True)
        wb = io.BytesIO()
        transfers.to_excel(wb, index=False, sheet_name="Redistribucion")
        st.download_button("⬇ Exportar plan de redistribución", wb.getvalue(), "Plan_Redistribucion_Makropetrol.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ============================================================
# FINANZAS
# ============================================================

elif view == "Finanzas":
    st.markdown('<div class="section-title">Control financiero del inventario</div>', unsafe_allow_html=True)
    if all_df.empty:
        empty_state("Sin datos financieros", "Procesa cortes para calcular inventario, compras y capital inmovilizado.", "💰")
    else:
        inv_value = float(all_df["Valor Inventario ($)"].sum())
        immobilized = float(all_df["Capital Inmovilizado ($)"].sum())
        purchase_col = "Compra Ajustada" if "Compra Ajustada" in all_df.columns else "Compra Sugerida"
        purchase_value = float((all_df[purchase_col] * all_df["Costo"]).sum())
        no_cost = int((all_df["Costo"] <= 0).sum())
        c1, c2, c3, c4 = st.columns(4)
        with c1: kpi("Valor inventario", money(inv_value), "Existencia × costo", "blue")
        with c2: kpi("Capital inmovilizado", money(immobilized), "Sobrestock", "red")
        with c3: kpi("Compra estimada", money(purchase_value), "Compra neta", "orange")
        with c4: kpi("Sin costo", integer(no_cost), "Revisar fuente", "green")
        a, b = st.columns(2)
        by_site = all_df.groupby("Sede")["Valor Inventario ($)"].sum().reset_index()
        with a:
            fig = px.bar(by_site, x="Sede", y="Valor Inventario ($)", title="Valor de inventario por sede")
            st.plotly_chart(chart_layout(fig), use_container_width=True)
        with b:
            by_site2 = all_df.groupby("Sede")["Capital Inmovilizado ($)"].sum().reset_index()
            fig = px.bar(by_site2, x="Sede", y="Capital Inmovilizado ($)", title="Capital inmovilizado por sede")
            st.plotly_chart(chart_layout(fig), use_container_width=True)
        st.markdown('<div class="section-title">Capital inmovilizado por producto</div>', unsafe_allow_html=True)
        financial = all_df[all_df["Capital Inmovilizado ($)"] > 0].sort_values("Capital Inmovilizado ($)", ascending=False).head(100)
        fin_cols = ["Sede", "Código", "Descripción", "Costo", "Existencia", "Retiro Almacén", "Capital Inmovilizado ($)"] if not modo_gerencia else ["Sede", "Descripción", "Retiro Almacén", "Capital Inmovilizado ($)"]
        max_cap = float(financial["Capital Inmovilizado ($)"].max()) if not financial.empty else 1.0
        st.dataframe(
            financial[fin_cols],
            use_container_width=True,
            hide_index=True,
            height=540,
            column_config={
                "Costo": st.column_config.NumberColumn(format="$%.2f"),
                "Capital Inmovilizado ($)": st.column_config.ProgressColumn(
                    "Capital inmovilizado", min_value=0, max_value=max(1.0, max_cap), format="$%.0f"
                ),
            },
        )


# ============================================================
# ALERTAS
# ============================================================

elif view == "Alertas":
    st.markdown('<div class="section-title">Centro de alertas</div>', unsafe_allow_html=True)
    if alerts.empty:
        empty_state("Todo tranquilo", "No se detectaron anomalías con los parámetros actuales.", "✓")
    else:
        sev = st.multiselect("Severidad", sorted(alerts["Severidad"].unique()), default=sorted(alerts["Severidad"].unique()))
        typ = st.multiselect("Tipo", sorted(alerts["Tipo"].unique()), default=sorted(alerts["Tipo"].unique()))
        data = alerts[alerts["Severidad"].isin(sev) & alerts["Tipo"].isin(typ)]
        counts = data["Severidad"].value_counts().reset_index()
        counts.columns = ["Severidad", "Cantidad"]
        fig = px.bar(counts, x="Severidad", y="Cantidad", title="Alertas por severidad")
        st.plotly_chart(chart_layout(fig, 280), use_container_width=True)
        st.dataframe(data, use_container_width=True, hide_index=True, height=580)


# ============================================================
# AUTOMATIZACIÓN
# ============================================================

elif view == "Automatización":
    st.markdown('<div class="section-title">Hub de automatización</div>', unsafe_allow_html=True)
    st.caption("El ERP genera payloads para conectarse con tu CRM y tu proveedor de WhatsApp por webhook. Así puedes orquestar avisos, tickets, pedidos y flujos externos sin meter credenciales dentro del código.")

    contact_rows = []
    for site in SEDES:
        c = contacts.get(site, {})
        contact_rows.append({
            "Sede": site,
            "Supervisor": c.get("supervisor", ""),
            "WhatsApp": c.get("whatsapp", ""),
            "Webhook": c.get("webhook", ""),
        })
    edited = st.data_editor(pd.DataFrame(contact_rows), use_container_width=True, hide_index=True, num_rows="fixed", key="contacts_editor")
    if st.button("💾 Guardar contactos y webhooks", use_container_width=True):
        new_contacts = {}
        for row in edited.to_dict("records"):
            new_contacts[str(row["Sede"])] = {
                "supervisor": str(row.get("Supervisor", "")),
                "whatsapp": str(row.get("WhatsApp", "")),
                "webhook": str(row.get("Webhook", "")),
            }
        contacts = new_contacts
        save_contacts(contacts)
        st.toast("Contactos y webhooks guardados", icon="💾")

    if all_df.empty:
        empty_state("Automatización lista para conectar", "Configura los webhooks y procesa un corte diario para generar mensajes y payloads.", "⚙")
    else:
        st.markdown('<div class="section-title">Envío por sede</div>', unsafe_allow_html=True)
        site = st.selectbox("Sede a notificar", sorted(site_results.keys()))
        message = compose_digest(site, site_results_adjusted[site], transfers)
        st.code(message, language="text")

        supervisor = contacts.get(site, {})
        payload = {
            "event": "nexus.daily.logistics",
            "version": APP_VERSION,
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "site": site,
            "supervisor": supervisor.get("supervisor", ""),
            "whatsapp": supervisor.get("whatsapp", ""),
            "health_score": health_score(site_results_adjusted[site]),
            "message": message,
            "metrics": {
                "critical": int(site_results_adjusted[site]["Estado"].str.startswith("CRÍTICO").sum()),
                "purchase_skus": int((site_results_adjusted[site].get("Compra Ajustada", site_results_adjusted[site]["Compra Sugerida"]) > 0).sum()),
                "withdrawal_skus": int((site_results_adjusted[site]["Retiro Almacén"] > 0).sum()),
                "inventory_value": float(site_results_adjusted[site]["Valor Inventario ($)"].sum()),
                "capital_immobilized": float(site_results_adjusted[site]["Capital Inmovilizado ($)"].sum()),
            },
        }

        st.json(payload)
        if st.button("🚀 Enviar al webhook de esta sede", use_container_width=True):
            ok, detail = post_webhook(supervisor.get("webhook", ""), payload)
            if ok:
                st.toast(f"Webhook enviado · {detail}", icon="🚀")
            else:
                st.error(f"No se envió · {detail}")

        st.info("Arquitectura recomendada: NEXUS → webhook → n8n/Make/tu API → CRM + proveedor WhatsApp. El proveedor de WhatsApp decide el formato, plantillas y reglas de envío.")


# ============================================================
# HISTORIAL
# ============================================================

elif view == "Historial":
    st.markdown('<div class="section-title">Historial de cortes y trazabilidad</div>', unsafe_allow_html=True)
    with db() as con:
        log = pd.read_sql_query(
            "SELECT processed_at, snapshot_date, site, inventory_rows, sales_rows FROM processing_log ORDER BY processed_at DESC LIMIT 300",
            con,
        )
    if log.empty:
        empty_state("Sin historial", "Los cortes que guardes aparecerán aquí.", "◷")
    else:
        st.dataframe(log, use_container_width=True, hide_index=True, height=620)


# ============================================================
# EXPORTACIÓN DE DATOS — MULTIFORMATO
# ============================================================

elif view == "Exportación de Datos":
    st.markdown('<div class="section-title">Centro de exportación y presentación</div>', unsafe_allow_html=True)
    st.caption("Elige cómo quieres consumir la información: Excel operativo, Excel gerencial, HTML interactivo, PDF ejecutivo, CSV o JSON.")
    if all_df.empty:
        empty_state("Nada que exportar", "Carga y procesa al menos un corte diario.", "📭")
    else:
        purchase_col = "Compra Ajustada" if "Compra Ajustada" in all_df.columns else "Compra Sugerida"
        executive = build_executive_frame(all_df)
        excel_bytes = export_excel(site_results_adjusted, transfers, int(months_history), float(min_coverage), float(max_coverage), int(lead_time_days), latest_display)
        gerencial_bytes = build_executive_excel(all_df, latest_display, score)
        html_bytes = build_interactive_html_report(all_df, transfers, alerts, latest_display, score)
        pdf_bytes = build_executive_pdf(all_df, transfers, latest_display, score)
        bundle_bytes = export_bundle(excel_bytes, all_df, transfers, alerts, latest_display, score)

        a,b,c = st.columns(3)
        with a:
            kpi("Formatos disponibles", "6", "Excel×2 · HTML · PDF · CSV · JSON", "blue")
        with b:
            kpi("Vista ejecutiva", integer(min(40, len(executive))), "Filas visibles del resumen", "green")
        with c:
            kpi("Interactividad", "Alta", "HTML con gráficos Plotly", "cyan")

        st.markdown('<div class="section-title">1 · Reporte gerencial</div>', unsafe_allow_html=True)
        r1,r2 = st.columns(2)
        with r1:
            st.markdown('<div class="panel"><h4>📊 Excel operativo</h4><p>Todo el detalle: inventario, compras, retiros, redistribución y alertas.</p></div>', unsafe_allow_html=True)
            st.download_button("⬇ Descargar Excel operativo", excel_bytes, f"NEXUS_Reporte_Operativo_{latest_display}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with r2:
            st.markdown('<div class="panel"><h4>👔 Excel gerencial</h4><p>Solo indicadores y acciones recomendadas, listo para dirección.</p></div>', unsafe_allow_html=True)
            st.download_button("⬇ Descargar Excel gerencial", gerencial_bytes, f"NEXUS_Reporte_Gerencial_{latest_display}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

        r3, r4 = st.columns(2)
        with r3:
            st.markdown('<div class="panel"><h4>🌐 Dashboard HTML interactivo</h4><p>Gráficos y tablas para presentar desde Chrome. Puede imprimirse como PDF desde el navegador.</p></div>', unsafe_allow_html=True)
            st.download_button("⬇ Descargar Dashboard HTML", html_bytes, f"NEXUS_Dashboard_{latest_display}.html", "text/html", use_container_width=True)
        with r4:
            st.markdown('<div class="panel"><h4>📄 PDF ejecutivo</h4><p>Una página lista para enviar por WhatsApp o correo.</p></div>', unsafe_allow_html=True)
            if pdf_bytes:
                st.download_button("⬇ Descargar PDF ejecutivo", pdf_bytes, f"NEXUS_Ejecutivo_{latest_display}.pdf", "application/pdf", use_container_width=True)
            else:
                st.info("Instala `fpdf2` (`pip install fpdf2`) para habilitar el PDF ejecutivo.")

        st.markdown('<div class="section-title">2 · Datos para análisis e integración</div>', unsafe_allow_html=True)
        d1,d2,d3 = st.columns(3)
        with d1:
            st.download_button("⬇ CSV consolidado", export_csv_bytes(all_df), f"NEXUS_Datos_{latest_display}.csv", "text/csv", use_container_width=True)
        with d2:
            st.download_button("⬇ JSON estructurado", export_json_bytes(all_df), f"NEXUS_Datos_{latest_display}.json", "application/json", use_container_width=True)
        with d3:
            st.download_button("📦 Paquete ZIP completo", bundle_bytes, f"NEXUS_Paquete_{latest_display}.zip", "application/zip", use_container_width=True)

        st.markdown('<div class="section-title">3 · Vista gerencial dentro de NEXUS</div>', unsafe_allow_html=True)
        exec_col_config: dict[str, Any] = {}
        if "Valor Inventario ($)" in executive.columns:
            exec_col_config["Valor Inventario ($)"] = st.column_config.NumberColumn(format="$%.2f")
        if "Capital Inmovilizado ($)" in executive.columns and not executive.empty:
            exec_col_config["Capital Inmovilizado ($)"] = st.column_config.ProgressColumn(
                "Capital Inmovilizado ($)", min_value=0,
                max_value=max(1.0, float(executive["Capital Inmovilizado ($)"].max())),
                format="$%.0f",
            )
        if "Costo" in executive.columns:
            exec_col_config["Costo"] = st.column_config.NumberColumn(format="$%.2f")
        if "Cobertura (meses)" in executive.columns:
            exec_col_config["Cobertura (meses)"] = st.column_config.NumberColumn(format="%.1f")
        st.dataframe(
            executive.head(40),
            use_container_width=True,
            hide_index=True,
            height=520,
            column_config=exec_col_config,
        )


# ============================================================
# EXPORTACIÓN EJECUTIVA GLOBAL
# ============================================================

if not all_df.empty and view != "Exportación de Datos":
    st.markdown('<div class="section-title">Reporte ejecutivo</div>', unsafe_allow_html=True)
    export_data = export_excel(
        site_results_adjusted,
        transfers,
        int(months_history),
        float(min_coverage),
        float(max_coverage),
        int(lead_time_days),
        latest_display,
    )
    st.download_button(
        "⬇ Descargar reporte ejecutivo NEXUS",
        data=export_data,
        file_name=f"Makropetrol_Nexus_Executive_{latest_display}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )

st.markdown(
    f"<div class='footer'>Makropetrol NEXUS v{APP_VERSION} · datos base persistidos en SQLite · recalculo dinámico · automatización por webhooks</div>",
    unsafe_allow_html=True,
)
